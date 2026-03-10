#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kengni Finance - Complete Financial Management & Trading Application
Version 2.0 - Enhanced with AI Analysis and Advanced Features
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import sqlite3
import os
from datetime import datetime, timedelta, date
import secrets
import random
import json
import pandas as pd
from io import BytesIO
import yfinance as yf
import numpy as np
import base64
from PIL import Image
import io
import urllib.request
import re
import smtplib
import threading
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import csv
from datetime import date as _date

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'kengni-finance-default-key-change-in-prod-2024')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'static/uploads'

# ── Configuration Gmail pour les rappels d'agenda ──────────────────────────────
GMAIL_CONFIG = {
    'sender_email':    'fabrice.kengni12@gmail.com',
    'sender_name':     'Kengni Finance — Agenda',
    'receiver_email':  'fabrice.kengni@icloud.com',
    'smtp_host':       'smtp.gmail.com',
    'smtp_port':       587,
    'smtp_password':   os.environ.get('GMAIL_APP_PASSWORD', ''),
}

# ── Types et couleurs des événements d'agenda ─────────────────────────────────
AGENDA_EVENT_COLORS = {
    'trading':   {'bg': '#00d4aa', 'border': '#00b894', 'icon': '📈', 'label': 'Trading'},
    'finance':   {'bg': '#4a9eff', 'border': '#2980b9', 'icon': '💰', 'label': 'Finance'},
    'formation': {'bg': '#a29bfe', 'border': '#6c5ce7', 'icon': '📚', 'label': 'Formation'},
    'personnel': {'bg': '#fd79a8', 'border': '#e84393', 'icon': '👤', 'label': 'Personnel'},
    'reunion':   {'bg': '#ffd700', 'border': '#f39c12', 'icon': '🤝', 'label': 'Réunion'},
    'revue':     {'bg': '#ff7675', 'border': '#d63031', 'icon': '🔍', 'label': 'Revue'},
}

# ── Informations de paiement Kengni Trading Academy ─────────────────────────
PAYMENT_INFO = {
    'orange_money': {'numero': '695 072 759', 'nom': 'Fabrice Kengni', 'label': 'Orange Money'},
    'mtn_money':    {'numero': '670 695 946', 'nom': 'Fabrice Kengni', 'label': 'MTN MoMo'},
    'paypal':       {'adresse': 'fabrice.kengni@icloud.com', 'label': 'PayPal'},
    'crypto':       {'adresse': 'fabrice.kengni@icloud.com', 'label': 'Crypto (via email)'},
}

FORMATION_PRICES = {
    'Débutant':       {'xaf': 25000,  'eur': 38},
    'Intermédiaire':  {'xaf': 50000,  'eur': 76},
    'Avancé':         {'xaf': 100000, 'eur': 152},
    'Pro / Mentoring':{'xaf': 200000, 'eur': 305},
}

# Add Python built-in functions to Jinja2 environment
# Custom Jinja2 filter: safely parse JSON
def _from_json_safe(s):
    try:
        import json as _j
        return _j.loads(s) if s else {}
    except Exception:
        return {}

app.jinja_env.filters['from_json_safe'] = _from_json_safe

app.jinja_env.globals.update({
    'abs': abs,
    'min': min,
    'max': max,
    'round': round,
    'int': int,
    'float': float,
    'len': len,
    'sum': sum
})

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database Configuration
DB_FILE = os.environ.get('DB_PATH', 'kengni_finance.db')

# Allowed extensions for image uploads
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Database Connection Helper
def get_db_connection():
    """Create and return database connection"""
    try:
        connection = sqlite3.connect(DB_FILE)
        connection.row_factory = sqlite3.Row
        return connection
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

# Initialize database with enhanced tables
def init_db():
    """Initialize database with all tables"""
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        
        # Users table with preferences
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            status TEXT DEFAULT 'active',
            preferred_currency TEXT DEFAULT 'EUR',
            timezone TEXT DEFAULT 'Europe/Paris',
            theme TEXT DEFAULT 'light',
            notifications_email INTEGER DEFAULT 1,
            notifications_app INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            last_login TEXT
        )
        ''')
        
        # Financial transactions with enhanced categories
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS financial_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('revenue', 'expense', 'receivable', 'credit', 'debt', 'investment', 'epargne')),
            category TEXT NOT NULL,
            subcategory TEXT,
            reason TEXT NOT NULL,
            usage TEXT,
            amount REAL NOT NULL,
            currency TEXT DEFAULT 'EUR',
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            payment_method TEXT,
            reference TEXT,
            status TEXT DEFAULT 'completed' CHECK(status IN ('completed', 'pending', 'cancelled')),
            notes TEXT,
            tags TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Trading positions
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS positions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            symbol TEXT NOT NULL,
            quantity REAL NOT NULL,
            avg_price REAL NOT NULL,
            current_price REAL NOT NULL,
            status TEXT DEFAULT 'open',
            stop_loss REAL,
            take_profit REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            closed_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Trading transactions
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            symbol TEXT NOT NULL,
            type TEXT NOT NULL,
            quantity REAL NOT NULL,
            price REAL NOT NULL,
            amount REAL NOT NULL,
            fees REAL DEFAULT 0,
            status TEXT DEFAULT 'completed',
            strategy TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Trading journal with images
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS trading_journal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            transaction_id INTEGER,
            symbol TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('buy', 'sell')),
            quantity REAL NOT NULL,
            entry_price REAL NOT NULL,
            exit_price REAL,
            profit_loss REAL,
            strategy TEXT,
            setup_description TEXT,
            emotions TEXT,
            mistakes TEXT,
            lessons_learned TEXT,
            notes TEXT,
            image_path TEXT,
            chart_analysis TEXT,
            market_conditions TEXT,
            risk_reward_ratio REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (transaction_id) REFERENCES transactions(id)
        )
        ''')
        
        # AI Analysis results
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS ai_analysis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            analysis_type TEXT NOT NULL CHECK(analysis_type IN ('financial', 'trading', 'psychological', 'strategy')),
            subject TEXT,
            insights TEXT NOT NULL,
            recommendations TEXT,
            warnings TEXT,
            confidence_score REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Trader performance scores
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS trader_scores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            overall_score REAL NOT NULL,
            discipline_score REAL,
            risk_management_score REAL,
            strategy_consistency_score REAL,
            emotional_control_score REAL,
            profitability_score REAL,
            monthly_trades INTEGER,
            win_rate REAL,
            profit_factor REAL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Psychological patterns detection
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS psychological_patterns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            pattern_type TEXT NOT NULL CHECK(pattern_type IN ('FOMO', 'revenge_trading', 'overtrading', 'overconfidence', 'fear', 'greed')),
            severity TEXT CHECK(severity IN ('low', 'medium', 'high', 'critical')),
            detected_date TEXT NOT NULL,
            description TEXT,
            evidence TEXT,
            recommendations TEXT,
            status TEXT DEFAULT 'active' CHECK(status IN ('active', 'resolved', 'monitoring')),
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Reports table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            report_type TEXT NOT NULL,
            period_start TEXT NOT NULL,
            period_end TEXT NOT NULL,
            revenue REAL DEFAULT 0,
            expenses REAL DEFAULT 0,
            profit REAL DEFAULT 0,
            profit_margin REAL DEFAULT 0,
            data TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        
        # Notifications
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('alert', 'warning', 'info', 'success')),
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            is_read INTEGER DEFAULT 0,
            action_url TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')

        # ── TABLE : Événements d'agenda ────────────────────────────────────────
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS agenda_events (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id          INTEGER NOT NULL,
            title            TEXT NOT NULL,
            description      TEXT,
            event_type       TEXT NOT NULL DEFAULT 'personnel'
                             CHECK(event_type IN ('trading','finance','formation','personnel','reunion','revue')),
            start_datetime   TEXT NOT NULL,
            end_datetime     TEXT NOT NULL,
            all_day          INTEGER DEFAULT 0,
            recurrence       TEXT DEFAULT 'none'
                             CHECK(recurrence IN ('none','daily','weekly','monthly')),
            reminder_minutes INTEGER DEFAULT 30,
            email_reminder   INTEGER DEFAULT 1,
            app_reminder     INTEGER DEFAULT 1,
            location         TEXT,
            notes            TEXT,
            linked_course_id INTEGER,
            status           TEXT DEFAULT 'active' CHECK(status IN ('active','cancelled','completed')),
            created_at       TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at       TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')

        # ── TABLE : Rappels envoyés (anti-doublon) ─────────────────────────────
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS agenda_reminders_sent (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id  INTEGER NOT NULL,
            sent_at   TEXT NOT NULL,
            method    TEXT NOT NULL CHECK(method IN ('email','app')),
            FOREIGN KEY (event_id) REFERENCES agenda_events(id)
        )
        ''')
        
        # Training courses
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL DEFAULT 'Sans titre',
            description TEXT,
            course_url TEXT,
            thumbnail_url TEXT,
            category TEXT DEFAULT 'Général',
            level TEXT DEFAULT 'debutant',
            day_of_week TEXT DEFAULT 'Non défini',
            scheduled_date TEXT,
            duration_minutes INTEGER DEFAULT 0,
            tags TEXT,
            is_published INTEGER DEFAULT 1,
            view_count INTEGER DEFAULT 0,
            participant_names TEXT DEFAULT '',
            analyses TEXT DEFAULT '',
            strategies TEXT DEFAULT '',
            position_images TEXT DEFAULT '[]',
            time_start TEXT DEFAULT '',
            time_end TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        # Migration: add columns if not exist
        for col, defn in [
            ('participant_names', 'TEXT DEFAULT ""'),
            ('analyses', 'TEXT DEFAULT ""'),
            ('strategies', 'TEXT DEFAULT ""'),
            ('position_images', 'TEXT DEFAULT "[]"'),
            ('time_start', 'TEXT DEFAULT ""'),
            ('time_end', 'TEXT DEFAULT ""'),
        ]:
            try:
                cursor.execute(f'ALTER TABLE training_courses ADD COLUMN {col} {defn}')
            except Exception:
                pass

        # ── TABLE : Inscriptions Kengni Trading Academy ──
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_leads (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name       TEXT NOT NULL,
            email           TEXT NOT NULL,
            whatsapp        TEXT NOT NULL,
            level_selected  TEXT NOT NULL,
            capital         TEXT,
            objective       TEXT,
            source          TEXT DEFAULT "Non renseigné",
            status          TEXT NOT NULL DEFAULT "Nouveau",
            notes           TEXT,
            user_id         INTEGER,
            created_at      TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')
        for col, defn in [('notes', 'TEXT'), ('user_id', 'INTEGER'), ('capital', 'TEXT'), ('objective', 'TEXT'),
                          ('payment_method', 'TEXT'), ('payment_ref', 'TEXT'), ('payment_status', "TEXT DEFAULT 'En attente'"),
                          ('amount_paid', 'REAL DEFAULT 0'), ('sincire_sent_at', 'TEXT')]:
            try:
                cursor.execute(f'ALTER TABLE training_leads ADD COLUMN {col} {defn}')
            except Exception:
                pass

        # ── TABLE : Bloc-Notes (entrées) ──────────────────────────────────────
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS bloc_notes (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            label       TEXT    NOT NULL,
            note_text   TEXT,
            amount      REAL,
            currency    TEXT    DEFAULT 'EUR',
            entry_type  TEXT    DEFAULT 'info'
                        CHECK(entry_type IN ('revenu','depense','tache','info')),
            entry_date  TEXT    NOT NULL,
            created_at  TEXT    DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')

        # ── TABLE : Bloc-Notes Mémo rapide ─────────────────────────────────────
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS bloc_notes_memo (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL UNIQUE,
            memo        TEXT    DEFAULT '',
            updated_at  TEXT    DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        ''')

        # Migration: add allowed_pages column if not exists
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN allowed_pages TEXT DEFAULT NULL")
        except Exception:
            pass

        # Migration: permissions granulaires boutique (add/edit/delete par utilisateur)
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN shop_permissions TEXT DEFAULT NULL")
        except Exception:
            pass

        # Migration: shop access pour tous les rôles (pas seulement shop_manager)
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN shop_access INTEGER DEFAULT 0")
        except Exception:
            pass

        # Table factures boutique
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS shop_invoices (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT    NOT NULL UNIQUE,
            order_id       INTEGER,
            type           TEXT    NOT NULL DEFAULT 'order',
            customer_name  TEXT    DEFAULT '',
            customer_phone TEXT    DEFAULT '',
            customer_city  TEXT    DEFAULT '',
            customer_email TEXT    DEFAULT '',
            items_json     TEXT    DEFAULT '[]',
            subtotal       REAL    DEFAULT 0,
            discount       REAL    DEFAULT 0,
            tax_rate       REAL    DEFAULT 0,
            tax_amount     REAL    DEFAULT 0,
            total          REAL    DEFAULT 0,
            pay_method     TEXT    DEFAULT '',
            status         TEXT    DEFAULT 'draft',
            notes          TEXT    DEFAULT '',
            created_by     INTEGER,
            created_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
            updated_at     TEXT    DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Table historique activité utilisateurs boutique (courbe évolution)
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS shop_activity_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            action      TEXT    NOT NULL,
            entity_type TEXT    DEFAULT '',
            entity_id   INTEGER DEFAULT 0,
            details     TEXT    DEFAULT '',
            created_at  TEXT    DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Migration: ajouter handled_by aux commandes (qui a géré)
        try:
            cursor.execute("ALTER TABLE shop_orders ADD COLUMN handled_by INTEGER DEFAULT NULL")
        except Exception:
            pass
        try:
            cursor.execute("ALTER TABLE shop_orders ADD COLUMN handled_at TEXT DEFAULT NULL")
        except Exception:
            pass

        # Check if default user exists
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE email = ?", ('fabrice.kengni@icloud.com',))
        if cursor.fetchone()[0] == 0:
            hashed_password = generate_password_hash(os.environ.get('ADMIN_PASSWORD', 'Kengni@fablo12'))
            cursor.execute('''
                INSERT INTO users (username, email, password, role, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', ('kengni', 'fabrice.kengni@icloud.com', hashed_password, 'admin', datetime.now().isoformat()))
        else:
            # Ensure admin always has the correct password (double sécurité)
            hashed_password = generate_password_hash(os.environ.get('ADMIN_PASSWORD', 'Kengni@fablo12'))
            cursor.execute(
                "UPDATE users SET password=? WHERE email=? AND role='admin'",
                (hashed_password, 'fabrice.kengni@icloud.com')
            )
        
        # ── Table annonces (Splash Screen post-2FA) ──
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS announcements (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            title             TEXT    NOT NULL,
            content           TEXT    NOT NULL,
            badge_label       TEXT    DEFAULT 'Annonce',
            badge_type        TEXT    DEFAULT 'default',
            start_date        TEXT    NOT NULL,
            end_date          TEXT    NOT NULL,
            auto_skip_seconds INTEGER DEFAULT 15,
            is_active         INTEGER DEFAULT 1,
            images            TEXT    DEFAULT '[]',
            author            TEXT    DEFAULT 'Admin',
            view_count        INTEGER DEFAULT 0,
            created_at        TEXT    NOT NULL,
            updated_at        TEXT    NOT NULL
        )
        ''')

        conn.commit()
        conn.close()
        print("✅ Database initialized successfully!")

# ── URL secrète admin ──
ADMIN_SECRET_TOKEN = os.environ.get('ADMIN_TOKEN', 'kengni-control-7749')
ADMIN_SECONDARY_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'Kengni@fablo12')


@app.context_processor
def inject_global_context():
    """Injecte des variables globales disponibles dans tous les templates"""
    ctx = {'training_total_nav': 0, 'user_allowed_pages': None, 'ALL_USER_PAGES': ALL_USER_PAGES}
    if 'user_id' in session:
        if session.get('role') not in ('admin', 'superadmin'):
            ctx['user_allowed_pages'] = get_user_allowed_pages(session['user_id'])
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT COUNT(*) as cnt FROM training_courses WHERE user_id = ?",
                    (session['user_id'],)
                )
                ctx['training_total_nav'] = cursor.fetchone()['cnt']
            except Exception:
                pass
            finally:
                conn.close()
    return ctx

# Authentication Decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin-only decorator — 404 pour tout le monde, l'admin reste invisible
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') not in ('admin', 'superadmin'):
            from flask import abort; abort(404)
        return f(*args, **kwargs)
    return decorated_function

# Page-level access decorator — checks user's allowed_pages JSON
# If allowed_pages is NULL (not set), user has access to everything.
# If set, the page key must be in the list.
ALL_USER_PAGES = [
    'dashboard', 'finances', 'history', 'portfolio', 'reports',
    'settings', 'trading', 'trading_journal', 'analysis', 'ai_assistant',
    'agenda', 'bloc_notes', 'notifications', 'inscription_trading'
]

def get_user_allowed_pages(user_id):
    """Returns the list of allowed pages for a user, or None (= all pages)."""
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT allowed_pages FROM users WHERE id=?", (user_id,))
        row = cursor.fetchone()
        conn.close()
        if row and row['allowed_pages']:
            try:
                return json.loads(row['allowed_pages'])
            except Exception:
                return None
    return None


def get_shop_perms(user_id):
    """Retourne les permissions boutique d'un utilisateur.
    {'add': bool, 'edit': bool, 'delete': bool, 'access': bool}
    Les admins/superadmin ont tout par defaut.
    Tout utilisateur avec shop_access=1 peut voir la boutique selon ses permissions.
    """
    role = session.get('role', 'user')
    if role in ('admin', 'superadmin'):
        return {'add': True, 'edit': True, 'delete': True, 'access': True}
    conn = get_db_connection()
    if conn:
        try:
            row = conn.execute(
                "SELECT shop_permissions, shop_access FROM users WHERE id=?", (user_id,)
            ).fetchone()
            conn.close()
            if row:
                has_access = bool(row['shop_access']) if 'shop_access' in row.keys() else False
                if row['shop_permissions']:
                    p = json.loads(row['shop_permissions'])
                    return {
                        'add':    bool(p.get('add',    False)),
                        'edit':   bool(p.get('edit',   False)),
                        'delete': bool(p.get('delete', False)),
                        'access': has_access or any([p.get('add'), p.get('edit'), p.get('delete')]),
                    }
                if has_access:
                    return {'add': False, 'edit': False, 'delete': False, 'access': True}
        except Exception:
            pass
    return {'add': False, 'edit': False, 'delete': False, 'access': False}


def can_shop(perm):
    """Vérifie si l'utilisateur connecté a la permission boutique donnée.
    perm : 'add' | 'edit' | 'delete'
    """
    role = session.get('role', 'user')
    if role in ('admin', 'superadmin'):
        return True
    uid = session.get('user_id')
    if not uid:
        return False
    return get_shop_perms(uid).get(perm, False)

def page_access_required(page_key):
    """Decorator: if user's allowed_pages is set and page_key not in it, show locked page."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            # Admins always have full access
            if session.get('role') in ('admin', 'superadmin'):
                return f(*args, **kwargs)
            allowed = get_user_allowed_pages(session['user_id'])
            if allowed is not None and page_key not in allowed:
                return render_template('locked_page.html', page_key=page_key), 403
            return f(*args, **kwargs)
        return decorated
    return decorator



def analyze_trading_psychology(user_id):
    """Analyze psychological trading patterns"""
    conn = get_db_connection()
    patterns = []
    
    if conn:
        cursor = conn.cursor()
        
        # Get recent transactions
        cursor.execute("""
            SELECT * FROM transactions
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 50
        """, (user_id,))
        transactions = [dict(row) for row in cursor.fetchall()]
        
        # Get journal entries
        cursor.execute("""
            SELECT * FROM trading_journal
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 20
        """, (user_id,))
        journal_entries = [dict(row) for row in cursor.fetchall()]
        
        # Analyze patterns
        
        # 1. Overtrading detection
        recent_24h = sum(1 for t in transactions if datetime.fromisoformat(t['created_at']) > datetime.now() - timedelta(hours=24))
        if recent_24h > 10:
            patterns.append({
                'type': 'overtrading',
                'severity': 'high' if recent_24h > 20 else 'medium',
                'description': f'Vous avez effectué {recent_24h} transactions en 24h',
                'recommendation': 'Prenez du recul. Le overtrading augmente les frais et diminue la qualité des décisions.'
            })
        
        # 2. FOMO detection (buying after big moves)
        buy_after_loss = 0
        for i in range(1, min(len(transactions), 10)):
            if transactions[i]['type'] == 'buy' and i > 0:
                prev = transactions[i-1]
                if prev['type'] == 'sell' and prev['amount'] < 0:  # Previous was a loss
                    buy_after_loss += 1
        
        if buy_after_loss >= 3:
            patterns.append({
                'type': 'FOMO',
                'severity': 'high',
                'description': 'Tendance à acheter immédiatement après des pertes',
                'recommendation': 'Attendez 30 minutes avant toute nouvelle transaction après une perte.'
            })
        
        # 3. Revenge trading
        consecutive_losses = 0
        max_consecutive = 0
        for t in transactions:
            if t['type'] == 'sell' and t['amount'] < 0:
                consecutive_losses += 1
                max_consecutive = max(max_consecutive, consecutive_losses)
            else:
                consecutive_losses = 0
        
        if max_consecutive >= 3:
            patterns.append({
                'type': 'revenge_trading',
                'severity': 'critical',
                'description': f'{max_consecutive} pertes consécutives détectées',
                'recommendation': 'Arrêtez de trader après 2 pertes consécutives. Analysez vos erreurs.'
            })
        
        # 4. Emotional patterns from journal
        emotional_keywords = {
            'fear': ['peur', 'anxieux', 'stressé', 'inquiet', 'nerveux'],
            'greed': ['avidité', 'cupide', 'trop confiant', 'sûr de moi'],
            'overconfidence': ['facile', 'certain', 'évident', 'garanti']
        }
        
        for entry in journal_entries:
            if entry.get('emotions'):
                emotions_text = entry['emotions'].lower()
                for emotion, keywords in emotional_keywords.items():
                    if any(kw in emotions_text for kw in keywords):
                        patterns.append({
                            'type': emotion,
                            'severity': 'medium',
                            'description': f'Émotion détectée: {emotion}',
                            'recommendation': 'Identifiée dans votre journal. Restez objectif.'
                        })
        
        # Save patterns to database
        for pattern in patterns:
            cursor.execute("""
                INSERT INTO psychological_patterns 
                (user_id, pattern_type, severity, detected_date, description, recommendations)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                pattern['type'],
                pattern['severity'],
                datetime.now().isoformat(),
                pattern['description'],
                pattern['recommendation']
            ))
        
        conn.commit()
        conn.close()
    
    return patterns

def calculate_trader_score(user_id):
    """Calculate comprehensive trader score (0-100)"""
    conn = get_db_connection()
    score_data = {
        'overall_score': 50,
        'discipline_score': 50,
        'risk_management_score': 50,
        'strategy_consistency_score': 50,
        'emotional_control_score': 50,
        'profitability_score': 50,
        'details': {}
    }
    
    if conn:
        cursor = conn.cursor()
        
        # Get transactions
        cursor.execute("""
            SELECT * FROM transactions
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 100
        """, (user_id,))
        transactions = [dict(row) for row in cursor.fetchall()]
        
        if not transactions:
            conn.close()
            return score_data
        
        # 1. Profitability Score (30% weight)
        total_profit = sum(t['amount'] for t in transactions if t['type'] == 'sell')
        total_invested = sum(abs(t['amount']) for t in transactions if t['type'] == 'buy')
        
        if total_invested > 0:
            roi = (total_profit / total_invested) * 100
            score_data['profitability_score'] = min(100, max(0, 50 + roi))
        
        wins = sum(1 for t in transactions if t['type'] == 'sell' and t['amount'] > 0)
        total_sells = sum(1 for t in transactions if t['type'] == 'sell')
        win_rate = (wins / total_sells * 100) if total_sells > 0 else 0
        score_data['details']['win_rate'] = round(win_rate, 2)
        
        # 2. Risk Management Score (25% weight)
        risk_score = 50
        
        # Check for stop losses
        cursor.execute("""
            SELECT COUNT(*) as with_sl FROM positions
            WHERE user_id = ? AND stop_loss IS NOT NULL
        """, (user_id,))
        positions_with_sl = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) as total FROM positions
            WHERE user_id = ?
        """, (user_id,))
        total_positions = cursor.fetchone()[0]
        
        if total_positions > 0:
            sl_percentage = (positions_with_sl / total_positions) * 100
            risk_score += (sl_percentage - 50) * 0.5
        
        # Check for position sizing consistency
        amounts = [abs(t['amount']) for t in transactions if t['type'] == 'buy']
        if len(amounts) > 5:
            avg_amount = np.mean(amounts)
            std_amount = np.std(amounts)
            cv = (std_amount / avg_amount) if avg_amount > 0 else 0
            if cv < 0.3:  # Good consistency
                risk_score += 20
            elif cv > 0.8:  # Poor consistency
                risk_score -= 20
        
        score_data['risk_management_score'] = min(100, max(0, risk_score))
        
        # 3. Discipline Score (20% weight)
        discipline_score = 50
        
        # Check for overtrading
        recent_24h = sum(1 for t in transactions if datetime.fromisoformat(t['created_at']) > datetime.now() - timedelta(hours=24))
        if recent_24h > 15:
            discipline_score -= 30
        elif recent_24h < 5:
            discipline_score += 20
        
        # Check for revenge trading patterns
        cursor.execute("""
            SELECT COUNT(*) FROM psychological_patterns
            WHERE user_id = ? AND pattern_type = 'revenge_trading' AND status = 'active'
        """, (user_id,))
        revenge_patterns = cursor.fetchone()[0]
        if revenge_patterns > 0:
            discipline_score -= 20
        
        score_data['discipline_score'] = min(100, max(0, discipline_score))
        
        # 4. Strategy Consistency (15% weight)
        cursor.execute("""
            SELECT strategy, COUNT(*) as count
            FROM transactions
            WHERE user_id = ? AND strategy IS NOT NULL
            GROUP BY strategy
        """, (user_id,))
        strategies = cursor.fetchall()
        
        strategy_score = 50
        if strategies:
            # Reward using consistent strategies
            max_strategy_count = max(s[1] for s in strategies)
            total_with_strategy = sum(s[1] for s in strategies)
            consistency = (max_strategy_count / total_with_strategy) * 100 if total_with_strategy > 0 else 0
            strategy_score = min(100, consistency)
        
        score_data['strategy_consistency_score'] = strategy_score
        
        # 5. Emotional Control (10% weight)
        cursor.execute("""
            SELECT COUNT(*) FROM psychological_patterns
            WHERE user_id = ? AND status = 'active'
        """, (user_id,))
        active_patterns = cursor.fetchone()[0]
        
        emotional_score = 100 - (active_patterns * 15)
        score_data['emotional_control_score'] = min(100, max(0, emotional_score))
        
        # Calculate overall score (weighted average)
        score_data['overall_score'] = round(
            score_data['profitability_score'] * 0.30 +
            score_data['risk_management_score'] * 0.25 +
            score_data['discipline_score'] * 0.20 +
            score_data['strategy_consistency_score'] * 0.15 +
            score_data['emotional_control_score'] * 0.10,
            2
        )
        
        # Save to database
        cursor.execute("""
            INSERT INTO trader_scores 
            (user_id, date, overall_score, discipline_score, risk_management_score, 
             strategy_consistency_score, emotional_control_score, profitability_score,
             monthly_trades, win_rate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id,
            datetime.now().isoformat(),
            score_data['overall_score'],
            score_data['discipline_score'],
            score_data['risk_management_score'],
            score_data['strategy_consistency_score'],
            score_data['emotional_control_score'],
            score_data['profitability_score'],
            len(transactions),
            win_rate
        ))
        
        conn.commit()
        conn.close()
    
    return score_data

def analyze_financial_report(data):
    """AI-powered financial report analysis"""
    try:
        insights = {
            'summary': "Analyse automatique du rapport financier",
            'recommendations': [],
            'risks': [],
            'opportunities': [],
            'anomalies': []
        }
        
        if 'revenue' in data and 'expenses' in data:
            profit_margin = ((data['revenue'] - data['expenses']) / data['revenue'] * 100) if data['revenue'] > 0 else 0
            
            if profit_margin > 20:
                insights['recommendations'].append("✅ Excellente marge bénéficiaire. Envisagez d'investir dans la croissance.")
                insights['opportunities'].append("Capacité d'investissement disponible")
            elif profit_margin < 10:
                insights['recommendations'].append("⚠️ Marge bénéficiaire faible. Optimisez vos dépenses.")
                insights['risks'].append("Risque de rentabilité")
            
            if data['expenses'] > data['revenue']:
                insights['risks'].append("🚨 Dépenses supérieures aux revenus - attention critique!")
                insights['recommendations'].append("Action immédiate requise: réduire les dépenses de " + 
                    f"{round((data['expenses'] - data['revenue']) / data['revenue'] * 100, 2)}%")
            
            # Anomaly detection
            if data['expenses'] > data['revenue'] * 1.5:
                insights['anomalies'].append("Dépenses anormalement élevées détectées")
        
        return insights
    except Exception as e:
        return {'error': str(e)}

def analyze_trade_image(image_path, trade_data):
    """Analyze trading chart image and provide insights"""
    insights = {
        'setup_quality': 'N/A',
        'entry_timing': 'N/A',
        'risk_reward': 'N/A',
        'recommendations': []
    }
    
    try:
        # Basic analysis based on trade data
        if trade_data.get('risk_reward_ratio'):
            rr = trade_data['risk_reward_ratio']
            if rr >= 2:
                insights['risk_reward'] = 'Excellent'
                insights['recommendations'].append('✅ Bon ratio risque/récompense')
            elif rr >= 1:
                insights['risk_reward'] = 'Acceptable'
                insights['recommendations'].append('⚠️ Ratio risque/récompense minimum atteint')
            else:
                insights['risk_reward'] = 'Mauvais'
                insights['recommendations'].append('❌ Ratio risque/récompense insuffisant')
        
        # Analyze based on profit/loss
        if trade_data.get('profit_loss'):
            pl = trade_data['profit_loss']
            if pl > 0:
                insights['recommendations'].append('✅ Trade gagnant - analysez ce qui a fonctionné')
            else:
                insights['recommendations'].append('📝 Trade perdant - identifiez les erreurs')
        
        # Entry timing analysis
        if trade_data.get('strategy'):
            insights['setup_quality'] = 'Défini'
            insights['recommendations'].append(f'Strategy utilisée: {trade_data["strategy"]}')
        
    except Exception as e:
        insights['error'] = str(e)
    
    return insights

def trading_recommendation(symbol, timeframe='1mo'):
    """AI trading recommendations based on market data"""
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period=timeframe)
        
        if hist.empty:
            return {'error': 'Données non disponibles'}
        
        current_price = hist['Close'].iloc[-1]
        sma_20 = hist['Close'].rolling(window=20).mean().iloc[-1]
        sma_50 = hist['Close'].rolling(window=50).mean().iloc[-1] if len(hist) >= 50 else sma_20
        
        rsi = calculate_rsi(hist['Close'])
        
        # Volume analysis
        avg_volume = hist['Volume'].mean()
        recent_volume = hist['Volume'].iloc[-1]
        volume_ratio = recent_volume / avg_volume if avg_volume > 0 else 1
        
        recommendation = {
            'symbol': symbol,
            'current_price': round(current_price, 2),
            'sma_20': round(sma_20, 2),
            'sma_50': round(sma_50, 2),
            'rsi': round(rsi, 2),
            'volume_ratio': round(volume_ratio, 2),
            'signal': 'NEUTRE',
            'confidence': 50,
            'analysis': [],
            'entry_points': [],
            'stop_loss': 0,
            'take_profit': 0
        }
        
        # Trend analysis
        if current_price > sma_20 and sma_20 > sma_50:
            recommendation['analysis'].append("📈 Tendance haussière confirmée")
            trend_score = 20
        elif current_price < sma_20 and sma_20 < sma_50:
            recommendation['analysis'].append("📉 Tendance baissière confirmée")
            trend_score = -20
        else:
            recommendation['analysis'].append("➡️ Tendance neutre/consolidation")
            trend_score = 0
        
        # RSI analysis
        if rsi > 70:
            recommendation['analysis'].append(f"🔴 RSI: {round(rsi, 2)} - Surachat détecté")
            rsi_score = -15
        elif rsi < 30:
            recommendation['analysis'].append(f"🟢 RSI: {round(rsi, 2)} - Survente détectée")
            rsi_score = 15
        else:
            recommendation['analysis'].append(f"🟡 RSI: {round(rsi, 2)} - Zone neutre")
            rsi_score = 0
        
        # Volume analysis
        if volume_ratio > 1.5:
            recommendation['analysis'].append(f"📊 Volume élevé ({round(volume_ratio, 2)}x) - Signal fort")
            volume_score = 10
        else:
            volume_score = 0
        
        # Generate signal
        total_score = trend_score + rsi_score + volume_score
        
        if total_score > 20:
            recommendation['signal'] = 'ACHAT FORT'
            recommendation['confidence'] = min(90, 50 + total_score)
            recommendation['entry_points'].append(round(current_price * 0.98, 2))
            recommendation['stop_loss'] = round(current_price * 0.95, 2)
            recommendation['take_profit'] = round(current_price * 1.10, 2)
        elif total_score > 5:
            recommendation['signal'] = 'ACHAT'
            recommendation['confidence'] = min(80, 50 + total_score)
            recommendation['entry_points'].append(round(current_price * 0.99, 2))
            recommendation['stop_loss'] = round(current_price * 0.96, 2)
            recommendation['take_profit'] = round(current_price * 1.08, 2)
        elif total_score < -20:
            recommendation['signal'] = 'VENTE FORTE'
            recommendation['confidence'] = min(90, 50 + abs(total_score))
            recommendation['entry_points'].append(round(current_price * 1.02, 2))
            recommendation['stop_loss'] = round(current_price * 1.05, 2)
            recommendation['take_profit'] = round(current_price * 0.90, 2)
        elif total_score < -5:
            recommendation['signal'] = 'VENTE'
            recommendation['confidence'] = min(80, 50 + abs(total_score))
            recommendation['entry_points'].append(round(current_price * 1.01, 2))
            recommendation['stop_loss'] = round(current_price * 1.04, 2)
            recommendation['take_profit'] = round(current_price * 0.92, 2)
        else:
            recommendation['signal'] = 'ATTENDRE'
            recommendation['confidence'] = 50
            recommendation['analysis'].append("⏸️ Pas de signal clair - attendre un meilleur setup")
        
        return recommendation
    except Exception as e:
        return {'error': str(e)}

def calculate_rsi(prices, period=14):
    """Calculate Relative Strength Index"""
    delta = prices.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.iloc[-1] if not rsi.empty else 50

@app.route('/verify-token', methods=['GET', 'POST'])
def verify_token_page():
    """Vérification du token 2FA"""
    email = request.args.get('email', session.get('pending_2fa_email', ''))
    
    if request.method == 'POST':
        entered_token = request.form.get('token', '').strip()
        stored_token  = session.get('pending_2fa_token', '')
        expires_str   = session.get('pending_2fa_expires', '')
        
        # Vérifier expiration
        expired = False
        if expires_str:
            try:
                expires = datetime.fromisoformat(expires_str)
                expired = datetime.now() > expires
            except Exception:
                expired = True
        
        if expired:
            # Nettoyer session
            for k in list(session.keys()):
                if k.startswith('pending_2fa_'):
                    session.pop(k, None)
            return render_template('verify_token.html',
                email=email, error='Token expiré. Veuillez vous reconnecter.',
                token=None, message=None)
        
        if entered_token == stored_token:
            # Token correct — compléter la session
            is_admin_login = session.pop('pending_2fa_is_admin_login', False)
            session['user_id']  = session.pop('pending_2fa_user_id',  None)
            session['username'] = session.pop('pending_2fa_username', '')
            session['email']    = session.pop('pending_2fa_email',    '')
            session['theme']    = session.pop('pending_2fa_theme',    'dark')
            session['role']     = session.pop('pending_2fa_role',     'user')
            session.pop('pending_2fa_token',   None)
            session.pop('pending_2fa_expires', None)
            session['admin_secondary_verified'] = False  # Reset secondary check on new login
            session.permanent = False  # Session expire a la fermeture du navigateur
            
            if is_admin_login:
                return redirect(url_for('admin_secondary_verify'))
            # Splash screen annonces (si annonce active) → sinon dashboard direct
            return redirect(url_for('show_announcement'))
        else:
            return render_template('verify_token.html',
                email=email,
                token=session.get('pending_2fa_token'),
                error='Code incorrect. Vérifiez et réessayez.',
                message=None)
    
    # GET — afficher la page avec le token (démo)
    pending_token = session.get('pending_2fa_token')
    if not pending_token:
        return redirect(url_for('login'))
    
    return render_template('verify_token.html',
        email=email,
        token=pending_token,
        error=None,
        message='Un code de vérification a été généré. Entrez-le ci-dessous pour continuer.')


def create_notification(user_id, notif_type, title, message):
    """Helper to create a notification"""
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO notifications (user_id, type, title, message)
                VALUES (?, ?, ?, ?)
            """, (user_id, notif_type, title, message))
            conn.commit()
            conn.close()
    except Exception:
        pass

# Routes

@app.route('/')
def index():
    """Landing page"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or request.form
        email = data.get('email')
        password = data.get('password')
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
            user = cursor.fetchone()
            
            if user and check_password_hash(user['password'], password):
                # Générer le token 2FA
                token_2fa = str(random.randint(100000, 999999))
                
                # Stocker les infos en session en attente de vérification
                session['pending_2fa_token']    = token_2fa
                session['pending_2fa_user_id']  = user['id']
                session['pending_2fa_username'] = user['username']
                session['pending_2fa_email']    = user['email']
                session['pending_2fa_theme']    = user['theme']
                session['pending_2fa_role']     = user['role']
                session['pending_2fa_expires']  = (datetime.now() + timedelta(minutes=5)).isoformat()
                
                # Update last login
                cursor.execute("UPDATE users SET last_login = ? WHERE id = ?", 
                             (datetime.now().isoformat(), user['id']))
                conn.commit()
                conn.close()
                
                if request.is_json:
                    return jsonify({'success': True, 'redirect': url_for('verify_token_page', email=email)})
                return redirect(url_for('verify_token_page', email=email))
            
            conn.close()
        
        if request.is_json:
            return jsonify({'success': False, 'message': 'Email ou mot de passe incorrect'}), 401
        return render_template('login.html', error='Email ou mot de passe incorrect')
    
    return render_template('login.html')

@app.route('/api/login-flyers', methods=['GET'])
def login_flyers():
    """
    Endpoint PUBLIC (sans login) pour la page de connexion.
    Retourne la liste des images du journal de trading + les flyers statiques,
    afin que le carrousel de la page login se synchronise automatiquement.
    """
    items = []

    # 1. Images uploadées dans le journal de trading (toutes, pas filtrées par user)
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT tj.image_path, tj.symbol, tj.type, tj.profit_loss, tj.strategy,
                       tj.date, tj.entry_price, tj.exit_price
                FROM trading_journal tj
                WHERE tj.image_path IS NOT NULL AND tj.image_path != ''
                ORDER BY tj.created_at DESC
                LIMIT 20
            """)
            rows = cursor.fetchall()
            conn.close()
            for row in rows:
                row = dict(row)
                # Normalise le chemin pour url_for
                img = row['image_path'].replace('static/', '').replace('\\', '/')
                pl = row.get('profit_loss')
                pl_str = (f"+{pl:.0f}€" if pl and pl > 0 else f"{pl:.0f}€") if pl is not None else None
                items.append({
                    'type': 'journal',
                    'img_url': f"/static/{img}",
                    'name': f"{row['symbol']} — {row['date']}",
                    'spec': row.get('strategy') or ('Achat' if row['type'] == 'buy' else 'Vente'),
                    'price': None,
                    'promo': pl_str,
                    'link': None,
                    'badge': 'TRADE' if (pl is None or pl >= 0) else 'PERTE',
                    'badge_color': '#00ff88' if (pl is None or pl >= 0) else '#ff4757'
                })
    except Exception as e:
        print(f"login_flyers journal error: {e}")

    # 2. Produits statiques (flyers) avec liens
    static_products = [
        {
            'img': '71JOBadJCL__SL1500___1_.jpg',
            'name': 'PC Gamer High-End',
            'price': '950k',
            'promo': '849k',
            'spec': 'RTX 4060 | 32GB RAM',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20PC%20Gamer%20High-End'
        },
        {
            'img': 'access-point-cisco-CAP2602I.jpg',
            'name': 'Cisco AP Pro',
            'price': '85k',
            'promo': '70k',
            'spec': 'Dual Band | PoE',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20Cisco%20AP%20Pro'
        },
        {
            'img': 'bat.jpg',
            'name': 'Batterie 12V 9Ah',
            'price': '25k',
            'promo': '18k',
            'spec': 'Cycle Profond',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20la%20Batterie%2012V'
        },
        {
            'img': 'mg-858.jpg',
            'name': 'Routeur MG-858',
            'price': '45k',
            'promo': '39k',
            'spec': '4G LTE Ultra-Fast',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20Routeur%20MG-858'
        },
        {
            'img': 'rb750r2_01-1.jpg',
            'name': 'MikroTik hEX lite',
            'price': '40k',
            'promo': '32k',
            'spec': '5 Ports Fast Ethernet',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20MikroTik%20hEX%20lite'
        },
        {
            'img': 'render.png',
            'name': 'Workstation Render',
            'price': '1.5M',
            'promo': '1.2M',
            'spec': 'Dual Xeon | 128GB RAM',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20la%20Workstation%20Render'
        },
        {
            'img': 'Y58.jpg',
            'name': 'Switch Industriel',
            'price': '120k',
            'promo': '95k',
            'spec': 'Giga 8 Ports Métal',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20Switch%20Industriel'
        },
        {
            'img': 'seche.jpg',
            'name': 'Produit Séché',
            'price': '35k',
            'promo': '28k',
            'spec': 'Qualité Premium',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé'
        },
        {
            'img': 't1000yellow_elqj3whe7jtl75gy.webp',
            'name': 'T1000 Yellow',
            'price': '60k',
            'promo': '49k',
            'spec': 'Design Industriel',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20T1000%20Yellow'
        },
        {
            'img': 'TL-WR841NEU14_0-288x202x86-L-7022505730_normal_1524475444511q.jpg',
            'name': 'TP-Link WR841N',
            'price': '18k',
            'promo': '14k',
            'spec': 'WiFi 300Mbps',
            'link': 'https://wa.me/237XXXXXXXXX?text=Je%20suis%20intéressé%20par%20le%20TP-Link%20WR841N'
        },
    ]
    for p in static_products:
        flyer_path = os.path.join(app.static_folder, 'flyers', p['img'])
        if not os.path.exists(flyer_path):
            continue  # Skip missing flyer images silently
        items.append({
            'type': 'product',
            'img_url': f"/static/flyers/{p['img']}",
            'name': p['name'],
            'spec': p['spec'],
            'price': p['price'],
            'promo': p['promo'],
            'link': p['link'],
            'badge': 'PROMO',
            'badge_color': '#00d4aa'
        })

    return jsonify({'success': True, 'items': items})


@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration"""
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        confirm_password = data.get('confirm_password')
        preferred_currency = data.get('preferred_currency', 'EUR')
        
        # Validation
        errors = []
        
        if not username or len(username) < 3:
            errors.append("Le nom d'utilisateur doit contenir au moins 3 caractères")
        
        if not email or '@' not in email:
            errors.append("Email invalide")
        
        if not password or len(password) < 6:
            errors.append("Le mot de passe doit contenir au moins 6 caractères")
        
        if password != confirm_password:
            errors.append("Les mots de passe ne correspondent pas")
        
        if errors:
            if request.is_json:
                return jsonify({'success': False, 'errors': errors}), 400
            return render_template('register.html', error=', '.join(errors))
        
        # Check if user already exists
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM users WHERE email = ?", (email,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                conn.close()
                if request.is_json:
                    return jsonify({'success': False, 'message': 'Cet email est déjà utilisé'}), 400
                return render_template('register.html', error='Cet email est déjà utilisé')
            
            # Create new user
            try:
                hashed_password = generate_password_hash(password)
                cursor.execute("""
                    INSERT INTO users 
                    (username, email, password, preferred_currency, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, (username, email, hashed_password, preferred_currency, datetime.now().isoformat()))
                
                conn.commit()
                user_id = cursor.lastrowid
                conn.close()
                
                # Auto-login after registration
                session['user_id'] = user_id
                session['username'] = username
                session['email'] = email
                session['theme'] = 'dark'
                session['role'] = 'user'
                
                if request.is_json:
                    return jsonify({'success': True, 'redirect': url_for('dashboard')})
                return redirect(url_for('dashboard'))
                
            except Exception as e:
                conn.close()
                if request.is_json:
                    return jsonify({'success': False, 'message': str(e)}), 500
                return render_template('register.html', error=f"Erreur lors de la création du compte: {str(e)}")
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    """User logout"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@page_access_required('dashboard')
def dashboard():
    """Main dashboard"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    stats = {
        'net_worth': 0,
        'monthly_cashflow': 0,
        'expense_ratio': 0,
        'savings_rate': 0,
        'total_revenue': 0,
        'total_expenses': 0,
        'trader_score': 0,
        'total_pnl': 0,
        'total_pnl_percent': 0,
        'total': 0,
        'total_closed': 0,
        'wins': 0,
        'win_rate': 0,
        'profit_factor': 0,
        'best': {'profit_loss': 0},
        'worst': {'profit_loss': 0}
    }
    
    if conn:
        cursor = conn.cursor()
        
        # Get financial stats
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN type = 'revenue' THEN amount ELSE 0 END) as total_revenue,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as total_expenses
            FROM financial_transactions
            WHERE user_id = ? AND date >= date('now', '-30 days')
        """, (user_id,))
        
        result = cursor.fetchone()
        if result:
            stats['total_revenue'] = result['total_revenue'] or 0
            stats['total_expenses'] = result['total_expenses'] or 0
            stats['monthly_cashflow'] = stats['total_revenue'] - stats['total_expenses']
            
            if stats['total_revenue'] > 0:
                stats['expense_ratio'] = (stats['total_expenses'] / stats['total_revenue']) * 100
                stats['savings_rate'] = 100 - stats['expense_ratio']
        
        # Get trading value
        cursor.execute("""
            SELECT SUM(quantity * current_price) as portfolio_value
            FROM positions
            WHERE user_id = ? AND status = 'open'
        """, (user_id,))
        
        result = cursor.fetchone()
        portfolio_value = result['portfolio_value'] or 0
        
        # Calculate total PnL from positions
        cursor.execute('''
            SELECT SUM((current_price - avg_price) * quantity) as total_pnl,
                   SUM(avg_price * quantity) as total_cost
            FROM positions
            WHERE user_id = ? AND status = 'open'
        ''', (user_id,))
        pnl_result = cursor.fetchone()
        if pnl_result and pnl_result['total_pnl'] is not None:
            stats['total_pnl'] = round(pnl_result['total_pnl'], 2)
            total_cost = pnl_result['total_cost'] or 0
            if total_cost > 0:
                stats['total_pnl_percent'] = round((stats['total_pnl'] / total_cost) * 100, 2)
        
        stats['net_worth'] = stats['monthly_cashflow'] + portfolio_value
        
        # Trading stats from journal
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN exit_price IS NOT NULL THEN 1 ELSE 0 END) as total_closed,
                SUM(CASE WHEN profit_loss > 0 THEN 1 ELSE 0 END) as wins,
                SUM(CASE WHEN profit_loss > 0 THEN profit_loss ELSE 0 END) as gross_profit,
                SUM(CASE WHEN profit_loss < 0 THEN ABS(profit_loss) ELSE 0 END) as gross_loss,
                MAX(profit_loss) as best,
                MIN(profit_loss) as worst
            FROM trading_journal
            WHERE user_id = ?
        """, (user_id,))
        tj = cursor.fetchone()
        if tj and tj['total']:
            stats['total'] = tj['total'] or 0
            stats['total_closed'] = tj['total_closed'] or 0
            stats['wins'] = tj['wins'] or 0
            stats['best'] = {'profit_loss': round(tj['best'] or 0, 2)}
            stats['worst'] = {'profit_loss': round(tj['worst'] or 0, 2)}
            if stats['total_closed'] > 0:
                stats['win_rate'] = round((stats['wins'] / stats['total_closed']) * 100, 1)
            gross_profit = tj['gross_profit'] or 0
            gross_loss = tj['gross_loss'] or 0
            if gross_loss > 0:
                stats['profit_factor'] = round(gross_profit / gross_loss, 2)

        # Get latest trader score
        cursor.execute("""
            SELECT overall_score FROM trader_scores
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 1
        """, (user_id,))
        
        result = cursor.fetchone()
        if result:
            stats['trader_score'] = result['overall_score']
        
        # Get recent transactions
        cursor.execute("""
            SELECT * FROM transactions
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 10
        """, (user_id,))
        recent_transactions = [dict(row) for row in cursor.fetchall()]
        
        # Get unread notifications
        cursor.execute("""
            SELECT COUNT(*) as unread FROM notifications
            WHERE user_id = ? AND is_read = 0
        """, (user_id,))
        unread_notifications = cursor.fetchone()['unread']
        
        # === Données pour graphique Performance (6 derniers mois) ===
        cursor.execute("""
            SELECT 
                strftime('%Y-%m', date) as month,
                SUM(CASE WHEN type='revenue' THEN amount ELSE 0 END) as revenus,
                SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as depenses
            FROM financial_transactions
            WHERE user_id = ? AND date >= date('now', '-6 months')
            GROUP BY month
            ORDER BY month
        """, (user_id,))
        perf_rows = cursor.fetchall()
        month_names = {'01':'Jan','02':'Fév','03':'Mar','04':'Avr','05':'Mai',
                       '06':'Juin','07':'Juil','08':'Août','09':'Sep','10':'Oct','11':'Nov','12':'Déc'}
        chart_labels   = [month_names.get(r['month'][-2:], r['month']) for r in perf_rows]
        chart_revenus  = [round(r['revenus'] or 0, 2) for r in perf_rows]
        chart_depenses = [round(r['depenses'] or 0, 2) for r in perf_rows]
        chart_solde    = [round((r['revenus'] or 0) - (r['depenses'] or 0), 2) for r in perf_rows]

        # === Données pour graphique Répartition par catégorie ===
        cursor.execute("""
            SELECT category, SUM(amount) as total
            FROM financial_transactions
            WHERE user_id = ? AND type='expense'
            GROUP BY category
            ORDER BY total DESC
            LIMIT 6
        """, (user_id,))
        cat_rows = cursor.fetchall()
        donut_labels = [r['category'] for r in cat_rows]
        donut_data   = [round(r['total'] or 0, 2) for r in cat_rows]

        # === Formations / Training ===
        cursor.execute("""
            SELECT id, title, category, level, scheduled_date, duration_minutes,
                   participant_names, time_start, time_end, created_at
            FROM training_courses
            WHERE user_id = ?
            ORDER BY scheduled_date DESC, created_at DESC
            LIMIT 5
        """, (user_id,))
        recent_trainings = [dict(r) for r in cursor.fetchall()]

        # Stats formations
        cursor.execute("SELECT COUNT(*) as total FROM training_courses WHERE user_id = ?", (user_id,))
        training_total = cursor.fetchone()['total']

        cursor.execute("SELECT SUM(duration_minutes) as total_min FROM training_courses WHERE user_id = ?", (user_id,))
        training_total_min = cursor.fetchone()['total_min'] or 0

        cursor.execute("""
            SELECT COUNT(*) as cnt FROM training_courses
            WHERE user_id = ? AND scheduled_date >= date('now', '-30 days')
        """, (user_id,))
        training_this_month = cursor.fetchone()['cnt']

        conn.close()

        return render_template('dashboard.html',
                             stats=stats,
                             transactions=recent_transactions,
                             unread_notifications=unread_notifications,
                             user_role=session.get('role','user'),
                             chart_labels=chart_labels,
                             chart_revenus=chart_revenus,
                             chart_depenses=chart_depenses,
                             chart_solde=chart_solde,
                             donut_labels=donut_labels,
                             donut_data=donut_data,
                             recent_trainings=recent_trainings,
                             training_total=training_total,
                             training_total_min=training_total_min,
                             training_this_month=training_this_month)

    return render_template('dashboard.html', stats=stats, transactions=[], unread_notifications=0,
                           user_role=session.get('role','user'),
                           chart_labels=[], chart_revenus=[], chart_depenses=[],
                           chart_solde=[], donut_labels=[], donut_data=[],
                           recent_trainings=[], training_total=0,
                           training_total_min=0, training_this_month=0)
@app.route('/finances')
@page_access_required('finances')
def finances():
    """Gestion financière avancée avec filtres, statistiques et graphiques"""
    user_id = session['user_id']
    filter_cat = request.args.get('category', '')
    filter_month = request.args.get('month', '')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Requête avec filtres dynamiques
    query = "SELECT * FROM financial_transactions WHERE user_id = ?"
    params = [user_id]
    
    if filter_cat:
        query += " AND category = ?"
        params.append(filter_cat)
    if filter_month:
        query += " AND strftime('%Y-%m', date) = ?"
        params.append(filter_month)
        
    query += " ORDER BY date DESC, time DESC"
    cursor.execute(query, tuple(params))
    transactions = [dict(row) for row in cursor.fetchall()]
    
    # Résumé financier
    total_rev = sum(t['amount'] for t in transactions if t['type'] in ('revenue', 'receivable', 'credit'))
    total_exp = sum(t['amount'] for t in transactions if t['type'] in ('expense', 'debt'))
    balance = total_rev - total_exp
    savings_rate = max((balance / total_rev * 100) if total_rev > 0 else 0, 0)

    summary = {
        'total_revenue':  total_rev,
        'total_expenses': total_exp,
        'net_balance':    balance,
        'balance':        balance,
        'savings_rate':   savings_rate,
        'period':         filter_month if filter_month else "Global"
    }
    
    # Données pour Chart.js (30 derniers jours — par jour)
    cursor.execute("""
        SELECT strftime('%Y-%m-%d', date) as day,
               SUM(CASE WHEN type='revenue' THEN amount ELSE 0 END) as rev,
               SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as exp
        FROM financial_transactions
        WHERE user_id = ? AND date >= date('now', '-30 days')
        GROUP BY day ORDER BY day ASC
    """, (user_id,))
    chart_raw = [dict(row) for row in cursor.fetchall()]
    
    # Liste des catégories pour le menu déroulant
    cursor.execute("SELECT DISTINCT category FROM financial_transactions WHERE user_id = ?", (user_id,))
    categories = [row[0] for row in cursor.fetchall()]
    conn.close()
    
    return render_template('finances.html', 
                           transactions=transactions, 
                           summary=summary, 
                           categories=categories,
                           chart_data=json.dumps(chart_raw[::-1]))

@app.route('/api/add-transaction', methods=['POST'])
@login_required
def add_transaction():
    """Route API pour ajouter une transaction avec image sécurisée"""
    try:
        user_id  = session['user_id']
        t_type   = request.form.get('type')
        amount   = float(request.form.get('amount', 0))
        category = request.form.get('category')
        # Compatibilité champ reason OU description
        reason   = request.form.get('reason') or request.form.get('description') or ''
        # Compatibilité champ date OU transaction_date
        t_date   = request.form.get('transaction_date') or request.form.get('date') or datetime.now().strftime('%Y-%m-%d')
        t_time   = request.form.get('time') or datetime.now().strftime('%H:%M:%S')
        currency        = request.form.get('currency', 'EUR')
        payment_method  = request.form.get('payment_method', '')
        tags            = ','.join(request.form.getlist('tags'))
        notes           = request.form.get('notes', '')
        emotional_state = ','.join(request.form.getlist('emotional_state'))
        energy_level    = request.form.get('energy_level', '3')

        # Enrichir les notes avec le contexte psychologique si renseigné
        if emotional_state:
            notes = f"{notes} [Émotions: {emotional_state}] [Énergie: {energy_level}/5]".strip()

        # Gestion de l'image justificative
        img_tag = ""
        if 'receipt_image' in request.files:
            file = request.files['receipt_image']
            if file and file.filename != '':
                filename = secure_filename(f"rec_{user_id}_{datetime.now().strftime('%m%d_%H%M%S')}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                img_tag = f" [IMG:{filename}]"

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO financial_transactions
            (user_id, type, amount, category, reason, date, time, status, currency, payment_method, tags, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (user_id, t_type, amount, category, f"{reason}{img_tag}",
              t_date, t_time, 'Terminé', currency, payment_method, tags, notes))
        conn.commit()
        conn.close()
        flash('Transaction et justificatif enregistrés !', 'success')
    except Exception as e:
        flash(f'Erreur : {str(e)}', 'error')
    return redirect(url_for('finances'))

@app.route('/delete-transaction/<int:id>', methods=['POST'])
@login_required
def delete_transaction(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    # Correction: suppression dans financial_transactions et non transactions
    cursor.execute("DELETE FROM financial_transactions WHERE id = ? AND user_id = ?", (id, session['user_id']))
    conn.commit()
    conn.close()
    return redirect(url_for('finances'))


@app.route('/delete-journal/<int:id>', methods=['POST'])
@login_required
def delete_journal(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM trading_journal WHERE id = ? AND user_id = ?", (id, session['user_id']))
    conn.commit()
    conn.close()
    return redirect(url_for('trading_journal'))

@app.route('/api/financial-transaction', methods=['POST'])
@login_required
def add_financial_transaction():
    """Add new financial transaction"""
    data = request.get_json(force=True, silent=True)
    user_id = session['user_id']
    
    required_fields = ['type', 'category', 'reason', 'amount', 'date', 'time']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO financial_transactions 
                (user_id, type, category, subcategory, reason, usage, amount, currency, 
                 date, time, payment_method, reference, status, notes, tags)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                data['type'],
                data['category'],
                data.get('subcategory'),
                data['reason'],
                data.get('usage'),
                float(data['amount']),
                data.get('currency', 'EUR'),
                data['date'],
                data['time'],
                data.get('payment_method'),
                data.get('reference'),
                data.get('status', 'completed'),
                data.get('notes'),
                data.get('tags')
            ))
            
            conn.commit()
            transaction_id = cursor.lastrowid
            
            # Create notification for large transactions
            if float(data['amount']) > 1000:
                cursor.execute("""
                    INSERT INTO notifications (user_id, type, title, message)
                    VALUES (?, 'info', 'Transaction importante', ?)
                """, (user_id, f"Transaction de {data['amount']}€ enregistrée"))
                conn.commit()
            
            conn.close()
            return jsonify({'success': True, 'id': transaction_id})
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500

@app.route('/journal')
@app.route('/journal-alias')
@page_access_required('trading_journal')
def trading_journal():
    """Trading journal with images"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    entries = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM trading_journal
            WHERE user_id = ?
            ORDER BY date DESC, time DESC
        """, (user_id,))
        entries = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('trading_journal.html', entries=entries)

@app.route('/api/journal-entry', methods=['POST'])
@login_required
def add_journal_entry():
    """Add trading journal entry with optional image"""
    user_id = session['user_id']
    
    # Handle file upload
    image_path = None
    if 'image' in request.files:
        file = request.files['image']
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            image_path = filepath
    
    # Get form data
    data = request.form if not request.is_json else request.get_json(force=True, silent=True)
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO trading_journal 
                (user_id, symbol, date, time, type, quantity, entry_price, exit_price, 
                 profit_loss, strategy, setup_description, emotions, mistakes, 
                 lessons_learned, notes, image_path, market_conditions, risk_reward_ratio)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                data.get('symbol'),
                data.get('date'),
                data.get('time'),
                data.get('type'),
                float(data.get('quantity', 0)),
                float(data.get('entry_price', 0)),
                float(data.get('exit_price', 0)) if data.get('exit_price') else None,
                float(data.get('profit_loss', 0)) if data.get('profit_loss') else None,
                data.get('strategy'),
                data.get('setup_description'),
                data.get('emotions'),
                data.get('mistakes'),
                data.get('lessons_learned'),
                data.get('notes'),
                image_path,
                data.get('market_conditions'),
                float(data.get('risk_reward_ratio', 0)) if data.get('risk_reward_ratio') else None
            ))
            
            entry_id = cursor.lastrowid
            
            # Analyze the trade if image provided
            if image_path:
                trade_data = {
                    'profit_loss': float(data.get('profit_loss', 0)) if data.get('profit_loss') else None,
                    'risk_reward_ratio': float(data.get('risk_reward_ratio', 0)) if data.get('risk_reward_ratio') else None,
                    'strategy': data.get('strategy')
                }
                analysis = analyze_trade_image(image_path, trade_data)
                
                # Save analysis
                cursor.execute("""
                    INSERT INTO ai_analysis (user_id, analysis_type, subject, insights)
                    VALUES (?, 'trading', ?, ?)
                """, (user_id, f"Journal Entry #{entry_id}", json.dumps(analysis)))
            
            conn.commit()
            conn.close()
            
            if request.is_json:
                return jsonify({'success': True, 'id': entry_id})
            return redirect(url_for('trading_journal'))
        
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500

@app.route('/trading')
@page_access_required('trading')
def trading():
    """Trading interface"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    positions = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM positions
            WHERE user_id = ? AND status = 'open'
            ORDER BY created_at DESC
        """, (user_id,))
        positions = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('trading.html', positions=positions)

@app.route('/api/execute-trade', methods=['POST'])
@login_required
def execute_trade():
    """Execute trade - accepte JSON et form-data"""
    if request.is_json:
        data = request.get_json(force=True, silent=True)
    else:
        data = request.form.to_dict()
    user_id = session['user_id']
    
    if not data:
        return jsonify({'success': False, 'message': 'Données manquantes'}), 400
    
    required_fields = ['symbol', 'type', 'quantity', 'price']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    
    symbol = data['symbol'].upper()
    trade_type = data['type']
    quantity = float(data['quantity'])
    price = float(data['price'])
    fees = float(data.get('fees', 0))
    strategy = data.get('strategy')
    
    amount = quantity * price
    
    if trade_type == 'sell':
        amount = amount - fees
    else:
        amount = -(amount + fees)
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            # Insert transaction
            cursor.execute("""
                INSERT INTO transactions (user_id, symbol, type, quantity, price, amount, fees, strategy, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (user_id, symbol, trade_type, quantity, price, amount, fees, strategy, datetime.now().isoformat()))
            
            # Update positions
            if trade_type == 'buy':
                cursor.execute("""
                    SELECT * FROM positions 
                    WHERE user_id = ? AND symbol = ? AND status = 'open'
                """, (user_id, symbol))
                existing_position = cursor.fetchone()
                
                if existing_position:
                    new_quantity = existing_position['quantity'] + quantity
                    new_avg_price = ((existing_position['quantity'] * existing_position['avg_price']) + 
                                   (quantity * price)) / new_quantity
                    
                    cursor.execute("""
                        UPDATE positions 
                        SET quantity = ?, avg_price = ?, updated_at = ?
                        WHERE user_id = ? AND symbol = ? AND status = 'open'
                    """, (new_quantity, new_avg_price, datetime.now().isoformat(), user_id, symbol))
                else:
                    cursor.execute("""
                        INSERT INTO positions (user_id, symbol, quantity, avg_price, current_price, status)
                        VALUES (?, ?, ?, ?, ?, 'open')
                    """, (user_id, symbol, quantity, price, price))
            else:  # sell
                cursor.execute("""
                    UPDATE positions 
                    SET quantity = quantity - ?, updated_at = ?
                    WHERE user_id = ? AND symbol = ? AND status = 'open'
                """, (quantity, datetime.now().isoformat(), user_id, symbol))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Transaction exécutée avec succès'})
        
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500

@app.route('/portfolio')
@page_access_required('portfolio')
def portfolio():
    """Portfolio management with enhanced structure"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    positions = []
    portfolio_stats = {
        'total_value': 0,
        'total_cost': 0,
        'total_pnl': 0,
        'total_pnl_percent': 0,
        'best_performer': None,
        'worst_performer': None
    }
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM positions
            WHERE user_id = ? AND status = 'open'
            ORDER BY (quantity * current_price) DESC
        """, (user_id,))
        positions = [dict(row) for row in cursor.fetchall()]
        
        # Calculate portfolio statistics
        for pos in positions:
            pos['market_value'] = pos['quantity'] * pos['current_price']
            pos['cost_basis'] = pos['quantity'] * pos['avg_price']
            pos['pnl'] = pos['market_value'] - pos['cost_basis']
            pos['pnl_percent'] = (pos['pnl'] / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
            
            portfolio_stats['total_value'] += pos['market_value']
            portfolio_stats['total_cost'] += pos['cost_basis']
            portfolio_stats['total_pnl'] += pos['pnl']
        
        if portfolio_stats['total_cost'] > 0:
            portfolio_stats['total_pnl_percent'] = (portfolio_stats['total_pnl'] / 
                                                   portfolio_stats['total_cost'] * 100)
        
        # Find best and worst performers
        if positions:
            positions_sorted = sorted(positions, key=lambda x: x['pnl_percent'], reverse=True)
            portfolio_stats['best_performer'] = positions_sorted[0]
            portfolio_stats['worst_performer'] = positions_sorted[-1]
        
        conn.close()
    
    return render_template('portfolio.html', 
                         positions=positions, 
                         stats=portfolio_stats)

@app.route('/api/add-position', methods=['POST'])
@login_required
def add_position():
    """Add new portfolio position"""
    data = request.get_json(force=True, silent=True)
    user_id = session['user_id']
    
    required_fields = ['symbol', 'quantity', 'avg_price']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    
    try:
        # Valider les données
        quantity = float(data['quantity'])
        avg_price = float(data['avg_price'])
        
        if quantity <= 0 or avg_price <= 0:
            return jsonify({'success': False, 'message': 'La quantité et le prix doivent être positifs'}), 400
        
        # Obtenir le prix actuel avec yfinance
        current_price = avg_price
        try:
            ticker = yf.Ticker(data['symbol'])
            hist = ticker.history(period='1d')
            if not hist.empty:
                current_price = float(hist['Close'].iloc[-1])
        except:
            pass  # Utiliser avg_price si la récupération échoue
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    INSERT INTO positions 
                    (user_id, symbol, asset_type, quantity, avg_price, current_price, 
                     status, platform, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    user_id,
                    data['symbol'].upper(),
                    data.get('asset_type', 'stock'),
                    quantity,
                    avg_price,
                    current_price,
                    'open',
                    data.get('platform', 'Manual'),
                    data.get('notes', '')
                ))
                
                conn.commit()
                position_id = cursor.lastrowid
                conn.close()
                
                return jsonify({'success': True, 'id': position_id})
            except sqlite3.IntegrityError as e:
                conn.close()
                return jsonify({'success': False, 'message': 'Cette position existe déjà'}), 400
            except Exception as e:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'message': str(e)}), 500
        
        return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500
    except ValueError:
        return jsonify({'success': False, 'message': 'Valeurs numériques invalides'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export-portfolio')
@login_required
def export_portfolio():
    """Export portfolio to various formats"""
    user_id = session['user_id']
    export_format = request.args.get('format', 'json')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM positions
        WHERE user_id = ? AND status = 'open'
        ORDER BY symbol
    """, (user_id,))
    positions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Calculer les valeurs
    for pos in positions:
        pos['market_value'] = pos['quantity'] * pos['current_price']
        pos['cost_basis'] = pos['quantity'] * pos['avg_price']
        pos['pnl'] = pos['market_value'] - pos['cost_basis']
        pos['pnl_percent'] = (pos['pnl'] / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
    
    if export_format == 'json':
        return jsonify({'success': True, 'data': positions})
    
    elif export_format == 'excel':
        try:
            df = pd.DataFrame(positions)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Portfolio')
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'portfolio_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    
    elif export_format == 'csv':
        try:
            df = pd.DataFrame(positions)
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'portfolio_{datetime.now().strftime("%Y%m%d")}.csv'
            )
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            output   = BytesIO()
            now      = datetime.now()
            doc_info = {
                'title'        : 'Rapport Portfolio',
                'type'         : 'PORTFOLIO',
                'period_start' : now.strftime('%Y-%m-01'),
                'period_end'   : now.strftime('%Y-%m-%d'),
                'generated_at' : now,
            }

            def _brand_page(c, doc):
                _kf_draw_branded_page(c, doc.pagesize[0], doc.pagesize[1], doc_info)

            doc = SimpleDocTemplate(
                output, pagesize=A4,
                topMargin=36 * mm, bottomMargin=20 * mm,
                leftMargin=18 * mm, rightMargin=18 * mm
            )
            elements = []
            styles   = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'KFTitle', parent=styles['Heading1'],
                fontSize=18, textColor=colors.HexColor('#1a1a2e'),
                spaceAfter=12, alignment=1
            )
            elements.append(Paragraph('Rapport Portfolio — Kengni Finance', title_style))
            elements.append(Spacer(1, 8))

            date_style = ParagraphStyle(
                'KFDate', parent=styles['Normal'],
                fontSize=9, textColor=colors.grey, alignment=1
            )
            elements.append(Paragraph(
                f'Généré le {now.strftime("%d/%m/%Y à %H:%M")}', date_style))
            elements.append(Spacer(1, 20))

            # Tableau de données
            data = [['Symbole', 'Qté', 'Prix Moy.', 'Prix Act.', 'P&L', 'P&L %']]
            total_value = 0
            total_cost  = 0

            for pos in positions:
                data.append([
                    pos['symbol'],
                    f"{pos['quantity']:.2f}",
                    f"{pos['avg_price']:.2f}€",
                    f"{pos['current_price']:.2f}€",
                    f"{pos['pnl']:.2f}€",
                    f"{pos['pnl_percent']:.2f}%"
                ])
                total_value += pos['market_value']
                total_cost  += pos['cost_basis']

            total_pnl         = total_value - total_cost
            total_pnl_percent = (total_pnl / total_cost * 100) if total_cost > 0 else 0
            data.append(['TOTAL', '', '', '', f'{total_pnl:.2f}€',
                         f'{total_pnl_percent:.2f}%'])

            col_w = [1.2*inch, 0.9*inch, 1*inch, 1.1*inch, 1*inch, 1*inch]
            table = Table(data, colWidths=col_w)
            table.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#00d4aa')),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  11),
                ('BOTTOMPADDING', (0, 0),  (-1, 0),  10),
                ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID',          (0, 0),  (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2),
                 [colors.white, colors.HexColor('#f9f9f9')]),
                ('FONTSIZE',      (0, 1),  (-1, -1), 9),
            ]))
            elements.append(table)

            doc.build(elements, onFirstPage=_brand_page, onLaterPages=_brand_page)
            output.seek(0)

            return send_file(
                output, mimetype='application/pdf', as_attachment=True,
                download_name=f'portfolio_{now.strftime("%Y%m%d")}.pdf'
            )
        except ImportError:
            return jsonify({'success': False,
                            'message': 'ReportLab non installé. Installez avec: pip install reportlab'}), 500
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export-finances')
@login_required
def export_finances():
    """Export financial transactions to various formats"""
    user_id = session['user_id']
    export_format = request.args.get('format', 'json')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM financial_transactions
        WHERE user_id = ?
        ORDER BY date DESC, time DESC
        LIMIT 1000
    """, (user_id,))
    transactions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    if export_format == 'json':
        return jsonify({'success': True, 'data': transactions})
    
    elif export_format == 'excel':
        try:
            df = pd.DataFrame(transactions)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Transactions')
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'finances_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    
    elif export_format == 'csv':
        try:
            df = pd.DataFrame(transactions)
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'finances_{datetime.now().strftime("%Y%m%d")}.csv'
            )
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            output = BytesIO()
            now    = datetime.now()

            # Déduire la période réelle depuis les transactions
            dates_sorted = sorted([t['date'] for t in transactions if t.get('date')]) if transactions else []
            p_start = dates_sorted[0]  if dates_sorted else now.strftime('%Y-%m-01')
            p_end   = dates_sorted[-1] if dates_sorted else now.strftime('%Y-%m-%d')

            doc_info = {
                'title'        : 'Rapport Transactions Financières',
                'type'         : 'FINANCES',
                'period_start' : p_start,
                'period_end'   : p_end,
                'generated_at' : now,
            }

            def _brand_page(c, doc):
                _kf_draw_branded_page(c, doc.pagesize[0], doc.pagesize[1], doc_info)

            doc = SimpleDocTemplate(
                output, pagesize=A4,
                topMargin=36 * mm, bottomMargin=20 * mm,
                leftMargin=18 * mm, rightMargin=18 * mm
            )
            elements = []
            styles   = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'KFTitle', parent=styles['Heading1'],
                fontSize=18, textColor=colors.HexColor('#1a1a2e'),
                spaceAfter=12, alignment=1
            )
            elements.append(Paragraph('Transactions Financières — Kengni Finance', title_style))
            elements.append(Spacer(1, 8))

            date_style = ParagraphStyle(
                'KFDate', parent=styles['Normal'],
                fontSize=9, textColor=colors.grey, alignment=1
            )
            elements.append(Paragraph(
                f'Généré le {now.strftime("%d/%m/%Y à %H:%M")}', date_style))
            elements.append(Spacer(1, 16))

            # Limiter à 50 transactions pour le PDF
            limited_transactions = transactions[:50]
            data          = [['Date', 'Type', 'Catégorie', 'Raison', 'Montant']]
            total_revenue = 0
            total_expense = 0

            for trans in limited_transactions:
                amount = float(trans['amount'])
                if trans['type'] == 'revenue':
                    total_revenue += amount
                    amount_str = f'+{amount:.2f}€'
                else:
                    total_expense += amount
                    amount_str = f'-{amount:.2f}€'

                data.append([
                    trans['date'],
                    trans['type'].capitalize(),
                    trans['category'][:15],
                    trans['reason'][:22],
                    amount_str
                ])

            balance = total_revenue - total_expense
            data.append(['', '', '', 'SOLDE', f'{balance:.2f}€'])

            col_w = [1.1*inch, 0.9*inch, 1.2*inch, 1.6*inch, 1*inch]
            table = Table(data, colWidths=col_w)
            table.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#00d4aa')),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  11),
                ('BOTTOMPADDING', (0, 0),  (-1, 0),  10),
                ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID',          (0, 0),  (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2),
                 [colors.white, colors.HexColor('#f9f9f9')]),
                ('FONTSIZE',      (0, 1),  (-1, -1), 9),
            ]))
            elements.append(table)

            if len(transactions) > 50:
                note_style = ParagraphStyle(
                    'KFNote', parent=styles['Normal'],
                    fontSize=8, textColor=colors.grey, alignment=1
                )
                elements.append(Spacer(1, 14))
                elements.append(Paragraph(
                    f'Note : Affichage limité à 50 transactions sur {len(transactions)} au total.',
                    note_style))

            doc.build(elements, onFirstPage=_brand_page, onLaterPages=_brand_page)
            output.seek(0)

            return send_file(
                output, mimetype='application/pdf', as_attachment=True,
                download_name=f'finances_{now.strftime("%Y%m%d")}.pdf'
            )
        except ImportError:
            return jsonify({'success': False,
                            'message': 'ReportLab non installé. Installez avec: pip install reportlab'}), 500
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    return jsonify({'success': False, 'message': 'Format non supporté'}), 400

@app.route('/api/analyze-portfolio')
@login_required
def analyze_portfolio():
    """Analyze portfolio with AI insights"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM positions
        WHERE user_id = ? AND status = 'open'
    """, (user_id,))
    positions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    if not positions:
        return jsonify({
            'success': True,
            'analysis': 'Aucune position dans le portfolio pour l\'instant.'
        })
    
    # Calculer les statistiques
    total_value = 0
    total_cost = 0
    best_performer = None
    worst_performer = None
    max_pnl_percent = float('-inf')
    min_pnl_percent = float('inf')
    
    for pos in positions:
        market_value = pos['quantity'] * pos['current_price']
        cost_basis = pos['quantity'] * pos['avg_price']
        pnl_percent = ((market_value - cost_basis) / cost_basis * 100) if cost_basis > 0 else 0
        
        total_value += market_value
        total_cost += cost_basis
        
        if pnl_percent > max_pnl_percent:
            max_pnl_percent = pnl_percent
            best_performer = pos['symbol']
        
        if pnl_percent < min_pnl_percent:
            min_pnl_percent = pnl_percent
            worst_performer = pos['symbol']
    
    total_pnl = total_value - total_cost
    total_pnl_percent = (total_pnl / total_cost * 100) if total_cost > 0 else 0
    
    # Générer l'analyse
    analysis = f"""📊 Analyse de votre Portfolio

💰 Valeur totale: {total_value:.2f}XAF
📈 P&L total: {total_pnl:+.2f}XAF ({total_pnl_percent:+.2f}%)
📦 Nombre de positions: {len(positions)}

🌟 Meilleure performance: {best_performer} ({max_pnl_percent:+.2f}%)
⚠️ Moins bonne performance: {worst_performer} ({min_pnl_percent:+.2f}%)

💡 Recommandations:
- {'Excellent rendement!' if total_pnl_percent > 10 else 'Continuez à diversifier votre portfolio'}
- {'Pensez à prendre des bénéfices sur ' + best_performer if max_pnl_percent > 20 else 'Surveillez les opportunités de renforcement'}
- {'Analysez ' + worst_performer + ' pour décider de conserver ou liquider' if min_pnl_percent < -10 else 'Portfolio bien équilibré'}
"""
    
    return jsonify({
        'success': True,
        'analysis': analysis,
        'stats': {
            'total_value': total_value,
            'total_pnl': total_pnl,
            'total_pnl_percent': total_pnl_percent,
            'positions_count': len(positions),
            'best_performer': best_performer,
            'worst_performer': worst_performer
        }
    })

@app.route('/api/ai-analyze-finances', methods=['GET', 'POST'])
@login_required
def ai_analyze_finances():
    """AI analysis of financial transactions"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500
    
    cursor = conn.cursor()
    
    # Get last 30 days transactions
    cursor.execute("""
        SELECT 
            type,
            category,
            SUM(amount) as total,
            COUNT(*) as count
        FROM financial_transactions
        WHERE user_id = ? AND date >= date('now', '-30 days')
        GROUP BY type, category
        ORDER BY total DESC
    """, (user_id,))
    
    categories = cursor.fetchall()
    
    cursor.execute("""
        SELECT 
            SUM(CASE WHEN type = 'revenue' THEN amount ELSE 0 END) as total_revenue,
            SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as total_expenses
        FROM financial_transactions
        WHERE user_id = ? AND date >= date('now', '-30 days')
    """, (user_id,))
    
    totals = cursor.fetchone()
    conn.close()
    
    total_revenue = totals['total_revenue'] or 0
    total_expenses = totals['total_expenses'] or 0
    balance = total_revenue - total_expenses
    
    # Generate analysis
    analysis = f"""📊 Analyse Financière des 30 derniers jours

💰 Revenus: {total_revenue:.2f}XAF
💸 Dépenses: {total_expenses:.2f}XAF
📈 Solde: {balance:+.2f}XAF

📋 Répartition par catégorie:
"""
    
    for cat in categories[:5]:
        analysis += f"\n- {cat['category']}: {cat['total']:.2f}€ ({cat['count']} transactions)"
    
    analysis += f"""

💡 Recommandations:
- {'Excellente gestion!' if balance > 0 else 'Attention aux dépenses'}
- {'Augmentez votre épargne de ' + str(int(balance * 0.2)) + '€' if balance > 500 else 'Réduisez vos dépenses non essentielles'}
- Taux d'épargne: {(balance/total_revenue*100 if total_revenue > 0 else 0):.1f}%
"""
    
    return jsonify({
        'success': True,
        'analysis': analysis
    })


@app.route('/ai-assistant')
@page_access_required('ai_assistant')
def ai_assistant():
    """AI Assistant conversational page"""
    return render_template('ai_assistant.html')

@app.route('/api/ai-chat', methods=['POST'])
@login_required
def ai_chat():
    """AI conversational assistant endpoint"""
    data = request.get_json(force=True, silent=True)
    user_id = session['user_id']
    question = data.get('question', '').lower()
    
    conn = get_db_connection()
    response = {
        'answer': '',
        'data': None,
        'charts': []
    }
    
    if conn:
        cursor = conn.cursor()
        
        # Analyze question and provide intelligent response
        if 'pourquoi' in question and ('perdu' in question or 'perte' in question):
            # Why did I lose money this month?
            cursor.execute("""
                SELECT 
                    symbol,
                    SUM(CASE WHEN type = 'sell' THEN amount ELSE 0 END) as total_sell,
                    SUM(CASE WHEN type = 'buy' THEN amount ELSE 0 END) as total_buy,
                    COUNT(*) as trade_count
                FROM transactions
                WHERE user_id = ? AND created_at >= date('now', '-30 days')
                GROUP BY symbol
                HAVING (total_sell + total_buy) < 0
                ORDER BY (total_sell + total_buy) ASC
            """, (user_id,))
            
            losing_trades = [dict(row) for row in cursor.fetchall()]
            
            if losing_trades:
                total_loss = sum(t['total_sell'] + t['total_buy'] for t in losing_trades)
                response['answer'] = f"Vous avez perdu {abs(total_loss):.2f}€ ce mois-ci. "
                response['answer'] += f"Les principales pertes proviennent de: "
                response['answer'] += ", ".join([f"{t['symbol']} ({t['total_sell'] + t['total_buy']:.2f}€)" 
                                                for t in losing_trades[:3]])
                response['data'] = losing_trades
            else:
                response['answer'] = "Vous n'avez pas enregistré de pertes ce mois-ci. Bravo!"
        
        elif 'stratégie' in question and ('rentable' in question or 'meilleur' in question):
            # Which strategy is most profitable?
            cursor.execute("""
                SELECT 
                    strategy,
                    COUNT(*) as trade_count,
                    SUM(amount) as total_profit,
                    AVG(amount) as avg_profit,
                    SUM(CASE WHEN amount > 0 THEN 1 ELSE 0 END) as wins,
                    SUM(CASE WHEN amount < 0 THEN 1 ELSE 0 END) as losses
                FROM transactions
                WHERE user_id = ? AND strategy IS NOT NULL AND type = 'sell'
                GROUP BY strategy
                ORDER BY total_profit DESC
            """, (user_id,))
            
            strategies = [dict(row) for row in cursor.fetchall()]
            
            if strategies:
                best = strategies[0]
                win_rate = (best['wins'] / best['trade_count'] * 100) if best['trade_count'] > 0 else 0
                
                response['answer'] = f"Votre meilleure stratégie est '{best['strategy']}' avec:\n"
                response['answer'] += f"• Profit total: {best['total_profit']:.2f}€\n"
                response['answer'] += f"• {best['trade_count']} trades\n"
                response['answer'] += f"• Taux de réussite: {win_rate:.1f}%\n"
                response['answer'] += f"• Profit moyen: {best['avg_profit']:.2f}€"
                response['data'] = strategies
            else:
                response['answer'] = "Vous n'avez pas encore de données de stratégie enregistrées."
        
        elif 'score' in question or 'performance' in question:
            # What's my trader score?
            score_data = calculate_trader_score(user_id)
            
            response['answer'] = f"Votre score de trader est: {score_data['overall_score']:.1f}/100\n\n"
            response['answer'] += "Détails:\n"
            response['answer'] += f"• Rentabilité: {score_data['profitability_score']:.1f}/100\n"
            response['answer'] += f"• Gestion du risque: {score_data['risk_management_score']:.1f}/100\n"
            response['answer'] += f"• Discipline: {score_data['discipline_score']:.1f}/100\n"
            response['answer'] += f"• Cohérence stratégique: {score_data['strategy_consistency_score']:.1f}/100\n"
            response['answer'] += f"• Contrôle émotionnel: {score_data['emotional_control_score']:.1f}/100"
            
            if score_data['overall_score'] < 50:
                response['answer'] += "\n\n⚠️ Votre score est faible. Concentrez-vous sur la discipline et la gestion du risque."
            elif score_data['overall_score'] < 70:
                response['answer'] += "\n\n📈 Bon début ! Travaillez sur la cohérence de vos stratégies."
            else:
                response['answer'] += "\n\n✅ Excellent score ! Continuez ainsi!"
            
            response['data'] = score_data
        
        elif 'combien' in question and ('gagn' in question or 'perdu' in question):
            # How much did I make/lose?
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) as total_gains,
                    SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END) as total_losses,
                    SUM(amount) as net_profit
                FROM transactions
                WHERE user_id = ? AND type = 'sell'
            """, (user_id,))
            
            result = cursor.fetchone()
            
            if result and result['total_gains']:
                response['answer'] = f"Résultats de trading:\n"
                response['answer'] += f"• Gains totaux: {result['total_gains']:.2f}€\n"
                response['answer'] += f"• Pertes totales: {result['total_losses']:.2f}€\n"
                response['answer'] += f"• Profit net: {result['net_profit']:.2f}€"
                
                if result['net_profit'] > 0:
                    response['answer'] += "\n\n✅ Vous êtes profitable!"
                else:
                    response['answer'] += "\n\n⚠️ Vous êtes en perte. Analysez vos trades."
            else:
                response['answer'] = "Vous n'avez pas encore de trades fermés."
        
        elif 'problème' in question or 'erreur' in question:
            # What are my problems?
            patterns = analyze_trading_psychology(user_id)
            
            if patterns:
                response['answer'] = f"J'ai détecté {len(patterns)} problèmes:\n\n"
                for i, pattern in enumerate(patterns[:5], 1):
                    response['answer'] += f"{i}. {pattern['type'].upper()} ({pattern['severity']})\n"
                    response['answer'] += f"   {pattern['description']}\n"
                    response['answer'] += f"   💡 {pattern['recommendation']}\n\n"
                response['data'] = patterns
            else:
                response['answer'] = "Aucun problème majeur détecté. Continuez votre bon travail!"
        
        elif 'conseil' in question or 'recommandation' in question:
            # Give me advice
            patterns = analyze_trading_psychology(user_id)
            score_data = calculate_trader_score(user_id)
            
            response['answer'] = "Recommandations personnalisées:\n\n"
            
            if score_data['discipline_score'] < 60:
                response['answer'] += "1. 📋 Discipline: Créez un plan de trading et suivez-le strictement\n"
            
            if score_data['risk_management_score'] < 60:
                response['answer'] += "2. 🛡️ Risque: Utilisez toujours des stop-loss (max 2% par trade)\n"
            
            if score_data['emotional_control_score'] < 60:
                response['answer'] += "3. 🧘 Émotions: Prenez une pause après 2 pertes consécutives\n"
            
            if patterns:
                response['answer'] += f"4. ⚠️ Attention: Vous montrez des signes de {patterns[0]['type']}\n"
            
            response['answer'] += "\n💡 Conseil du jour: Tenez un journal de trading détaillé pour identifier vos patterns."
        
        elif 'finance' in question or 'dépense' in question or 'revenu' in question or 'solde' in question:
            # Financial summary
            cursor.execute("""
                SELECT
                    SUM(CASE WHEN type IN ('revenue','receivable','credit') THEN amount ELSE 0 END) as rev,
                    SUM(CASE WHEN type IN ('expense','debt') THEN amount ELSE 0 END) as exp,
                    COUNT(*) as cnt
                FROM financial_transactions
                WHERE user_id = ? AND date >= date('now', '-30 days')
            """, (user_id,))
            row = cursor.fetchone()
            rev = row['rev'] or 0
            exp = row['exp'] or 0
            sol = rev - exp
            pct = (sol/rev*100) if rev > 0 else 0
            response['answer'] = f"📊 Résumé Financier (30 derniers jours)\n\n"
            response['answer'] += f"💰 Revenus  : {rev:.2f}€ ({rev*655.96:.0f} XAF)\n"
            response['answer'] += f"💸 Dépenses : {exp:.2f}€ ({exp*655.96:.0f} XAF)\n"
            response['answer'] += f"📈 Solde Net: {sol:+.2f}€  |  Taux épargne: {pct:.1f}%\n\n"
            if sol > 0:
                response['answer'] += "✅ Vous êtes en positif ce mois. Continuez ainsi!"
            else:
                response['answer'] += "⚠️ Vos dépenses dépassent vos revenus. Réduisez les dépenses non-essentielles."

        elif 'objectif' in question or 'goal' in question or 'target' in question:
            response['answer'] = "🎯 Objectifs recommandés basés sur votre profil:\n\n"
            response['answer'] += "1. 💰 Taux d'épargne minimum : 20% de vos revenus\n"
            response['answer'] += "2. 🛡️ Stop-loss systématique sur chaque trade (max 2% du capital)\n"
            response['answer'] += "3. 📓 Journal de trading après chaque session\n"
            response['answer'] += "4. 🧘 Pause obligatoire après 2 pertes consécutives\n"
            response['answer'] += "5. 📊 Win rate cible : ≥ 55% avec Risk/Reward ≥ 1:2"

        elif 'formation' in question or 'cours' in question or 'training' in question:
            cursor.execute("SELECT COUNT(*) as cnt, SUM(duration_minutes) as dur FROM training_courses WHERE user_id=?", (user_id,))
            row = cursor.fetchone()
            cnt = row['cnt'] or 0
            dur = row['dur'] or 0
            response['answer'] = f"🎓 Votre parcours de formation:\n\n"
            response['answer'] += f"📚 {cnt} sessions enregistrées\n"
            response['answer'] += f"⏱ {dur} minutes de formation ({dur//60}h{dur%60}min)\n\n"
            if cnt < 5:
                response['answer'] += "💡 Conseil: Augmentez la fréquence de vos formations.\nVisez au moins 1 session par jour."
            else:
                response['answer'] += "✅ Bonne régularité dans votre formation!"

        else:
            response['answer'] = "Je suis votre assistant IA trading et finance. Voici ce que je peux analyser :\n\n"
            response['answer'] += "📊 **Trading**\n"
            response['answer'] += "• 'Pourquoi j'ai perdu ce mois-ci?'\n"
            response['answer'] += "• 'Quelle est ma meilleure stratégie?'\n"
            response['answer'] += "• 'Quel est mon score de trader?'\n"
            response['answer'] += "• 'Quels sont mes problèmes psychologiques?'\n"
            response['answer'] += "• 'Combien j'ai gagné/perdu?'\n\n"
            response['answer'] += "💰 **Finance**\n"
            response['answer'] += "• 'Montre mon solde financier'\n"
            response['answer'] += "• 'Analyse mes dépenses'\n\n"
            response['answer'] += "🎓 **Formation**\n"
            response['answer'] += "• 'Combien de cours j'ai fait?'\n\n"
            response['answer'] += "🎯 **Objectifs**\n"
            response['answer'] += "• 'Quels sont mes objectifs?'\n"
            response['answer'] += "• 'Donne-moi des conseils'"
        
        conn.close()
    
    return jsonify(response)

@app.route('/analysis')
@page_access_required('analysis')
def analysis():
    """Analysis and insights page"""
    user_id = session['user_id']
    
    # Calculate scores and patterns
    trader_score = calculate_trader_score(user_id)
    patterns = analyze_trading_psychology(user_id)
    
    conn = get_db_connection()
    recent_analyses = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM ai_analysis
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 10
        """, (user_id,))
        recent_analyses = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('analysis.html', 
                         trader_score=trader_score,
                         patterns=patterns,
                         analyses=recent_analyses)

@app.route('/api/analyze-finances', methods=['POST'])
@login_required
def analyze_finances():
    """Analyze financial data"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        
        # Get financial data
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN type = 'revenue' THEN amount ELSE 0 END) as revenue,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expenses
            FROM financial_transactions
            WHERE user_id = ? AND date >= date('now', '-30 days')
        """, (user_id,))
        
        result = cursor.fetchone()
        data = {
            'revenue': result['revenue'] or 0,
            'expenses': result['expenses'] or 0
        }
        
        insights = analyze_financial_report(data)
        
        # Save analysis
        cursor.execute("""
            INSERT INTO ai_analysis (user_id, analysis_type, subject, insights)
            VALUES (?, 'financial', 'Monthly Report', ?)
        """, (user_id, json.dumps(insights)))
        
        conn.commit()
        conn.close()
        
        return jsonify(insights)
    
    return jsonify({'error': 'Database connection failed'}), 500

@app.route('/api/trading-recommendation/<symbol>')
@login_required
def get_trading_recommendation(symbol):
    """Get AI trading recommendation for a symbol"""
    recommendation = trading_recommendation(symbol.upper())
    return jsonify(recommendation)

@app.route('/settings')
@page_access_required('settings')
def settings():
    """User settings page"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    user_settings = {}
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        if user:
            user_settings = dict(user)
        conn.close()
    
    return render_template('settings.html', settings=user_settings, user_role=session.get('role','user'))

@app.route('/api/update-settings', methods=['POST'])
@login_required
def update_settings():
    """Update user settings"""
    data = request.get_json(force=True, silent=True)
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            updates = []
            params = []
            
            allowed_fields = ['username', 'email', 'preferred_currency', 'timezone', 
                            'theme', 'notifications_email', 'notifications_app']
            
            for field in allowed_fields:
                if field in data:
                    updates.append(f"{field} = ?")
                    params.append(data[field])
            
            if 'password' in data and data['password']:
                updates.append("password = ?")
                params.append(generate_password_hash(data['password']))
            
            if updates:
                params.append(user_id)
                query = f"UPDATE users SET {', '.join(updates)} WHERE id = ?"
                cursor.execute(query, params)
                conn.commit()
                
                # Update session theme
                if 'theme' in data:
                    session['theme'] = data['theme']
            
            conn.close()
            return jsonify({'success': True, 'message': 'Paramètres mis à jour'})
        
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return jsonify({'success': False, 'message': 'Erreur de connexion'}), 500

@app.route('/notifications')
@page_access_required('notifications')
def notifications():
    """Notifications page"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    notifications_list = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM notifications
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 50
        """, (user_id,))
        notifications_list = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('notifications.html', notifications=notifications_list)

@app.route('/api/mark-notification-read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Mark notification as read"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE notifications SET is_read = 1
            WHERE id = ? AND user_id = ?
        """, (notification_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    return jsonify({'success': False}), 500

@app.route('/reports')
@page_access_required('reports')
def reports():
    """Reports page"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    reports_list = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM reports
            WHERE user_id = ?
            ORDER BY created_at DESC
        """, (user_id,))
        reports_list = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('reports.html', reports=reports_list)

@app.route('/api/generate-report', methods=['POST'])
@login_required
def generate_report():
    """Generate financial report"""
    data = request.get_json(force=True, silent=True)
    user_id = session['user_id']
    
    report_type = data.get('type', 'monthly')
    period_start = data.get('start')
    period_end = data.get('end')
    
    # Générer des dates par défaut si non fournies
    if not period_start or not period_end:
        today = datetime.now()
        if report_type == 'monthly':
            period_start = today.replace(day=1).strftime('%Y-%m-%d')
            period_end = today.strftime('%Y-%m-%d')
        elif report_type == 'yearly':
            period_start = today.replace(month=1, day=1).strftime('%Y-%m-%d')
            period_end = today.strftime('%Y-%m-%d')
        else:  # weekly
            period_start = (today - timedelta(days=7)).strftime('%Y-%m-%d')
            period_end = today.strftime('%Y-%m-%d')
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        
        try:
            # Get financial data for period
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN type = 'revenue' THEN amount ELSE 0 END) as revenue,
                    SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expenses
                FROM financial_transactions
                WHERE user_id = ? AND date BETWEEN ? AND ?
            """, (user_id, period_start, period_end))
            
            result = cursor.fetchone()
            revenue = result['revenue'] or 0
            expenses = result['expenses'] or 0
            profit = revenue - expenses
            profit_margin = (profit / revenue * 100) if revenue > 0 else 0
            
            # Create report
            cursor.execute("""
                INSERT INTO reports 
                (user_id, title, report_type, period_start, period_end, revenue, expenses, profit, profit_margin)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                f"Rapport {report_type} - {period_start} à {period_end}",
                report_type,
                period_start,
                period_end,
                revenue,
                expenses,
                profit,
                profit_margin
            ))
            
            conn.commit()
            report_id = cursor.lastrowid
            conn.close()
            
            return jsonify({'success': True, 'report_id': report_id})
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    return jsonify({'success': False, 'message': 'Erreur de connexion à la base de données'}), 500

# ═══════════════════════════════════════════════════════════════════
#  HELPER PARTAGÉ : en-tête, filigrane, QR code, pied de page PDF
#  Appelé sur chaque page de chaque export PDF Kengni Finance
# ═══════════════════════════════════════════════════════════════════

def _kf_draw_branded_page(c, width, height, doc_info=None):
    """
    Dessine sur la page courante (canvas ReportLab) :
      - Logo k-ni en filigrane central (semi-transparent)
      - Bande d'en-tête verte : logo petit (centre), infos doc (gauche), infos proprio (droite)
      - QR code en bas à gauche (URL inscription + WhatsApp)
      - Pied de page certifié avec logo miniature

    doc_info : dict optionnel {
        'title'        : str,
        'type'         : str,
        'period_start' : str,
        'period_end'   : str,
        'generated_at' : datetime
    }
    """
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    import io as _io

    doc_info = doc_info or {}
    now      = doc_info.get('generated_at', datetime.now())

    # ── Chemin du logo ────────────────────────────────────────────────
    _LOGO_PATH = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'static', 'img', 'logo.jpeg'
    )

    def _logo_reader():
        """Retourne un ImageReader du logo, ou None si introuvable."""
        try:
            if os.path.exists(_LOGO_PATH):
                return ImageReader(_LOGO_PATH)
        except Exception:
            pass
        return None

    logo = _logo_reader()

    # ═══════════════════════════════════════════════════════════════
    # 1. FILIGRANE : logo centré, semi-transparent (fond de page)
    # ═══════════════════════════════════════════════════════════════
    if logo:
        c.saveState()
        c.setFillAlpha(0.06)           # très discret comme l'image de référence
        wm_size = 110 * mm
        wm_x    = (width  - wm_size) / 2
        wm_y    = (height - wm_size) / 2
        c.drawImage(logo, wm_x, wm_y, wm_size, wm_size,
                    mask='auto', preserveAspectRatio=True)
        c.restoreState()
    else:
        # Fallback texte si logo absent
        c.saveState()
        c.setFillColor(colors.HexColor('#00d4aa'))
        c.setFillAlpha(0.04)
        c.setFont('Helvetica-Bold', 68)
        c.translate(width / 2, height / 2)
        c.rotate(40)
        for offset in (-120, 0, 120):
            c.drawCentredString(0, offset, 'KENGNI FINANCE')
        c.restoreState()

    # ═══════════════════════════════════════════════════════════════
    # 2. EN-TÊTE : bande verte (30 mm)
    # ═══════════════════════════════════════════════════════════════
    hdr_h = 32 * mm
    c.setFillColor(colors.HexColor('#051a0f'))   # vert très sombre
    c.rect(0, height - hdr_h, width, hdr_h, fill=True, stroke=False)

    # Séparateur doré en bas du header
    c.setStrokeColor(colors.HexColor('#00d4aa'))
    c.setLineWidth(1.8)
    c.line(0, height - hdr_h, width, height - hdr_h)

    # ── Logo dans l'en-tête (centré, carré) ──────────────────────────
    if logo:
        logo_hdr_size = 24 * mm
        logo_hdr_x    = (width - logo_hdr_size) / 2
        logo_hdr_y    = height - hdr_h + (hdr_h - logo_hdr_size) / 2
        # Fond blanc circulaire derrière le logo
        c.setFillColor(colors.white)
        cx = logo_hdr_x + logo_hdr_size / 2
        cy = logo_hdr_y + logo_hdr_size / 2
        c.circle(cx, cy, logo_hdr_size / 2 + 1.5, fill=True, stroke=False)
        c.drawImage(logo, logo_hdr_x, logo_hdr_y,
                    logo_hdr_size, logo_hdr_size,
                    mask='auto', preserveAspectRatio=True)

    # ── Infos document (gauche) ───────────────────────────────────────
    c.setFillColor(colors.white)
    y_top      = height - 7 * mm
    doc_type   = str(doc_info.get('type',  'RAPPORT')).upper()
    doc_title  = str(doc_info.get('title', 'Rapport Kengni Finance'))[:42]
    p_start    = doc_info.get('period_start', '')
    p_end      = doc_info.get('period_end',   '')
    period_str = f"{p_start} → {p_end}" if p_start else 'Tous'

    c.setFont('Helvetica-Bold', 9.5)
    c.drawString(12 * mm, y_top, 'k-ni chez Htech-training')
    c.setFont('Helvetica', 7)
    c.drawString(12 * mm, y_top - 5 * mm,  f'Type     : {doc_type}')
    c.drawString(12 * mm, y_top - 9.5 * mm, f'Période  : {period_str}')
    c.drawString(12 * mm, y_top - 14 * mm,  doc_title)
    c.drawString(12 * mm, y_top - 18.5 * mm,
                 f"Édité le : {now.strftime('%d/%m/%Y à %H:%M')}")

    # ── Infos propriétaire (droite) ───────────────────────────────────
    c.setFont('Helvetica-Bold', 9.5)
    c.drawRightString(width - 12 * mm, y_top, 'Fabrice Kengni Nzoyem')
    c.setFont('Helvetica', 7)
    c.drawRightString(width - 12 * mm, y_top - 5 * mm,   'WhatsApp : +237 695 072 759')
    c.drawRightString(width - 12 * mm, y_top - 9.5 * mm,  'kengni.pythonanywhere.com')
    c.drawRightString(width - 12 * mm, y_top - 14 * mm,   'Kengni Trading Academy')
    c.drawRightString(width - 12 * mm, y_top - 18.5 * mm,
                      f"Réf. : KF-{now.strftime('%Y%m%d%H%M')}")

    # Ligne verte fine sous le header (séparation corps)
    c.setStrokeColor(colors.HexColor('#00d4aa'))
    c.setLineWidth(0.8)
    c.line(12 * mm, height - hdr_h - 4 * mm,
           width - 12 * mm, height - hdr_h - 4 * mm)

    # ═══════════════════════════════════════════════════════════════
    # 3. QR CODE — bas à gauche (comme l'attestation DGI)
    # ═══════════════════════════════════════════════════════════════
    try:
        import qrcode as _qrcode

        qr_data = (
            'https://kengni.pythonanywhere.com/inscription-trading'
            '?ref=kf_doc&wa=237695072759'
        )
        qr = _qrcode.QRCode(
            version=2, box_size=5, border=2,
            error_correction=_qrcode.constants.ERROR_CORRECT_M
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color='#051a0f', back_color='white')
        qr_buf = _io.BytesIO()
        qr_img.save(qr_buf, format='PNG')
        qr_buf.seek(0)

        qr_size = 28 * mm
        qr_x    = 12 * mm
        qr_y    = 19 * mm

        # Fond blanc + bordure verte légère
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor('#00d4aa'))
        c.setLineWidth(0.6)
        c.roundRect(qr_x - 1.5, qr_y - 1.5,
                    qr_size + 3, qr_size + 3, 3,
                    fill=True, stroke=True)

        c.drawImage(ImageReader(qr_buf), qr_x, qr_y,
                    qr_size, qr_size, mask='auto')

        # Légende + URL sous le QR (style DGI)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont('Helvetica-Bold', 5)
        c.drawCentredString(qr_x + qr_size / 2, qr_y - 3.5 * mm,
                            'Scanner pour s\'inscrire')
        c.setFont('Helvetica', 4.5)
        c.drawCentredString(qr_x + qr_size / 2, qr_y - 6.5 * mm,
                            'kengni.pythonanywhere.com/inscription-trading')
    except Exception:
        pass   # qrcode non installé → ignoré silencieusement

    # ═══════════════════════════════════════════════════════════════
    # 4. PIED DE PAGE avec logo miniature
    # ═══════════════════════════════════════════════════════════════
    ftr_h = 17 * mm
    c.setFillColor(colors.HexColor('#f0f4f0'))
    c.rect(0, 0, width, ftr_h, fill=True, stroke=False)
    c.setStrokeColor(colors.HexColor('#00d4aa'))
    c.setLineWidth(1)
    c.line(0, ftr_h, width, ftr_h)

    # Logo miniature dans le pied de page (à droite)
    if logo:
        logo_ftr_size = 11 * mm
        c.setFillColor(colors.white)
        c.circle(width - 14 * mm, ftr_h / 2,
                 logo_ftr_size / 2 + 1, fill=True, stroke=False)
        c.drawImage(logo,
                    width - 14 * mm - logo_ftr_size / 2,
                    ftr_h / 2 - logo_ftr_size / 2,
                    logo_ftr_size, logo_ftr_size,
                    mask='auto', preserveAspectRatio=True)

    c.setFillColor(colors.HexColor('#444444'))
    c.setFont('Helvetica-Bold', 6.5)
    c.drawCentredString(width / 2, ftr_h - 4 * mm,
        'Document certifié — Kengni Finance v2.1 — © 2025 Tous droits réservés')
    c.setFont('Helvetica', 6)
    c.drawCentredString(width / 2, ftr_h - 7.5 * mm,
        'k-ni chez Htech-training  ·  +237 695 072 759  ·  WhatsApp : +237 695 072 759')
    c.setFont('Helvetica', 5.5)
    c.drawCentredString(width / 2, ftr_h - 11 * mm,
        f"Généré le {now.strftime('%d/%m/%Y à %H:%M')}  ·  Document confidentiel  ·  kengni.pythonanywhere.com")


@app.route('/api/download-report/<int:report_id>', methods=['GET'])
@login_required
def download_report(report_id):
    """Télécharger un rapport en PDF avec filigrane, QR code et en-tête pro"""
    user_id = session['user_id']
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB error'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM reports WHERE id = ? AND user_id = ?", (report_id, user_id))
        report = cursor.fetchone()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

    if not report:
        return jsonify({'error': 'Rapport introuvable'}), 404
    report = dict(report)

    try:
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm

        buffer  = BytesIO()
        c       = pdf_canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        now      = datetime.now()
        doc_info = {
            'title'        : report.get('title', 'Rapport Kengni Finance'),
            'type'         : report.get('report_type', 'financial'),
            'period_start' : report.get('period_start', ''),
            'period_end'   : report.get('period_end',   ''),
            'generated_at' : now,
        }

        # ── Branding (watermark + header + QR + footer) ───────────────
        _kf_draw_branded_page(c, width, height, doc_info)

        # ── Corps du document (décalé sous le header) ─────────────────
        # Zone utile : de (hdr_h + 8mm) jusqu'à (ftr_h + 5mm)
        hdr_h  = 30 * mm
        content_top = height - hdr_h - 12 * mm

        # Titre du rapport
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.setFont('Helvetica-Bold', 15)
        c.drawCentredString(width / 2, content_top,
                            report.get('title', 'Rapport Kengni Finance'))

        # Ligne séparatrice
        c.setStrokeColor(colors.HexColor('#00d4aa'))
        c.setLineWidth(2)
        c.line(40, content_top - 8, width - 40, content_top - 8)

        # Données financières
        revenue  = float(report.get('revenue')  or 0)
        expenses = float(report.get('expenses') or 0)
        profit   = float(report.get('profit')   or revenue - expenses)
        margin   = float(report.get('profit_margin') or
                         (profit / revenue * 100 if revenue > 0 else 0))

        rows = [
            ('💰 Revenus Total',       f'{revenue:,.2f} €',  '#00c853'),
            ('💸 Dépenses Total',      f'{expenses:,.2f} €', '#d50000'),
            ('📈 Profit / Perte',      f'{profit:+,.2f} €',
             '#00c853' if profit >= 0 else '#d50000'),
            ('📊 Marge Bénéficiaire',  f'{margin:.1f} %',    '#1565c0'),
        ]

        y = content_top - 30
        for label, value, color in rows:
            c.setFillColor(colors.HexColor('#f5f5f5'))
            c.roundRect(40, y - 10, width - 80, 32, 4, fill=True, stroke=False)
            c.setFillColor(colors.HexColor('#333333'))
            c.setFont('Helvetica-Bold', 12)
            c.drawString(58, y + 7, label)
            c.setFillColor(colors.HexColor(color))
            c.setFont('Helvetica-Bold', 13)
            c.drawRightString(width - 58, y + 7, value)
            y -= 44

        # Note de génération
        c.setFillColor(colors.HexColor('#888888'))
        c.setFont('Helvetica', 8)
        c.drawCentredString(width / 2, y - 10,
            f"Rapport généré le {now.strftime('%d/%m/%Y à %H:%M')}")

        c.save()
        buffer.seek(0)
        filename = (f"rapport_{report.get('report_type','custom')}_"
                    f"{report.get('period_start','').replace('-','')}.pdf")
        return send_file(buffer, as_attachment=True,
                         download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': f'Erreur PDF: {str(e)}'}), 500


@app.route('/report/<int:report_id>')
@login_required
def view_report(report_id):
    """Alias vers download_report"""
    return redirect(url_for('download_report', report_id=report_id))


@app.route('/history')
@page_access_required('history')
def history():
    """Transaction history"""
    user_id = session['user_id']
    
    conn = get_db_connection()
    transactions = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM transactions
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 200
        """, (user_id,))
        transactions = [dict(row) for row in cursor.fetchall()]
        conn.close()
    
    return render_template('history.html', transactions=transactions)

@app.route('/delete-journal-entry/<int:id>', methods=['POST'])
@login_required
def delete_journal_entry(id):
    if session.get('role') not in ('admin', 'superadmin'):
        flash('Suppression réservée aux administrateurs', 'danger')
        return redirect(url_for('trading_journal'))
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM trading_journal WHERE id = ? AND user_id = ?",
                          (id, session['user_id']))
            conn.commit()
            conn.close()
            flash('Entrée supprimée', 'success')
            return redirect(url_for('trading_journal'))
    except Exception as e:
        flash(f'Erreur: {e}', 'danger')
        return redirect(url_for('trading_journal'))


@app.route('/api/delete-financial-transaction/<int:id>', methods=['DELETE', 'POST'])
@login_required
def delete_financial_transaction(id):
    """Delete a financial transaction"""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM financial_transactions WHERE id = ? AND user_id = ?",
                          (id, session['user_id']))
            conn.commit()
            conn.close()
            
            create_notification(session['user_id'], 'success', 
                              'Transaction supprimée', 
                              f'La transaction #{id} a été supprimée')
            
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delete-trade/<int:id>', methods=['DELETE', 'POST'])
@login_required
def delete_trade(id):
    """Delete a trade"""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM transactions WHERE id = ? AND user_id = ?",
                          (id, session['user_id']))
            conn.commit()
            conn.close()
            
            create_notification(session['user_id'], 'success', 
                              'Trade supprimé', 
                              f'Le trade #{id} a été supprimé')
            
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
        
@app.route('/image-spam')
@admin_required
def image_spam():
    return render_template('image_spam_manager.html')

@app.route('/api/delete-position/<int:id>', methods=['DELETE', 'POST'])
@login_required
def delete_position(id):
    """Delete a position"""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM positions WHERE id = ? AND user_id = ?",
                          (id, session['user_id']))
            conn.commit()
            conn.close()
            
            create_notification(session['user_id'], 'success', 
                              'Position supprimée', 
                              f'La position #{id} a été supprimée')
            
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



# ══════════════════════════════════════════════════════════════
# PANNEAU ADMIN SECRET
# ══════════════════════════════════════════════════════════════

@app.route(f'/{ADMIN_SECRET_TOKEN}')
def admin_secret_entry():
    if 'user_id' in session and session.get('role') in ('admin', 'superadmin'):
        return redirect(url_for('admin_panel'))
    return render_template('admin_login.html', token=ADMIN_SECRET_TOKEN)

@app.route(f'/{ADMIN_SECRET_TOKEN}/auth', methods=['POST'])
def admin_auth():
    data = request.get_json(force=True, silent=True) or request.form
    email    = data.get('email','').strip()
    password = data.get('password','').strip()
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE email=? AND role IN ('admin','superadmin')", (email,))
        user = cursor.fetchone()
        if user and check_password_hash(user['password'], password):
            # 2FA for admin
            token_2fa = str(random.randint(100000, 999999))
            session['pending_2fa_token']    = token_2fa
            session['pending_2fa_user_id']  = user['id']
            session['pending_2fa_username'] = user['username']
            session['pending_2fa_email']    = user['email']
            session['pending_2fa_theme']    = user['theme']
            session['pending_2fa_role']     = user['role']
            session['pending_2fa_expires']  = (datetime.now() + timedelta(minutes=5)).isoformat()
            session['pending_2fa_is_admin_login'] = True
            cursor.execute("UPDATE users SET last_login=? WHERE id=?", (datetime.now().isoformat(), user['id']))
            conn.commit(); conn.close()
            if request.is_json: return jsonify({'success': True, 'redirect': url_for('verify_token_page', email=user['email'])})
            return redirect(url_for('verify_token_page', email=user['email']))
        conn.close()
    if request.is_json: return jsonify({'success': False, 'message': 'Identifiants incorrects'}), 401
    from flask import abort; abort(404)

# ════════════════════════════════════════════════════════════
# SPLASH SCREEN ANNONCES — Routes
# Visualisation : tous utilisateurs connectés
# Administration : admin uniquement
# ════════════════════════════════════════════════════════════

@app.route('/announcement')
@login_required
def show_announcement():
    """Splash screen affiché juste après la validation du token 2FA."""
    today = date.today().isoformat()
    conn  = get_db_connection()
    ann   = None
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM announcements
            WHERE is_active = 1
              AND start_date <= ?
              AND end_date   >= ?
            ORDER BY created_at DESC
            LIMIT 1
        """, (today, today))
        row = cursor.fetchone()
        if row:
            ann = dict(row)
            cursor.execute("UPDATE announcements SET view_count = view_count + 1 WHERE id = ?", (ann['id'],))
            conn.commit()
            try:
                ann['images_list'] = json.loads(ann.get('images') or '[]')
            except Exception:
                ann['images_list'] = []
        conn.close()

    if not ann:
        return redirect(url_for('dashboard'))

    return render_template('announcement.html',
                           announcement=ann,
                           auto_skip_seconds=ann.get('auto_skip_seconds', 15))


@app.route('/admin/announcements')
@admin_required
def admin_announcements():
    """Dashboard admin — liste et gestion des annonces."""
    if not session.get('admin_secondary_verified'):
        return redirect(url_for('admin_secondary_verify'))
    today = date.today().isoformat()
    conn  = get_db_connection()
    announcements = []
    stats = {'total': 0, 'active': 0, 'upcoming': 0, 'expired': 0}
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM announcements ORDER BY created_at DESC")
        rows = cursor.fetchall()
        for row in rows:
            ann = dict(row)
            ann['is_currently_active'] = (
                ann['is_active'] == 1 and ann['start_date'] <= today <= ann['end_date']
            )
            ann['is_upcoming'] = ann['start_date'] > today
            ann['is_expired']  = ann['end_date'] < today
            try:
                ann['images_count'] = len(json.loads(ann.get('images') or '[]'))
            except Exception:
                ann['images_count'] = 0
            announcements.append(ann)
        stats['total']    = len(announcements)
        stats['active']   = sum(1 for a in announcements if a['is_currently_active'])
        stats['upcoming'] = sum(1 for a in announcements if a['is_upcoming'])
        stats['expired']  = sum(1 for a in announcements if a['is_expired'])
        conn.close()
    return render_template('admin_announcements.html', announcements=announcements, stats=stats)


@app.route('/admin/announcements/create', methods=['POST'])
@admin_required
def admin_announcement_create():
    try:
        title      = request.form.get('title', '').strip()
        content    = request.form.get('content', '').strip()
        start_date = request.form.get('start_date', '').strip()
        end_date   = request.form.get('end_date', '').strip()
        if not all([title, content, start_date, end_date]):
            return jsonify({'success': False, 'message': 'Champs obligatoires manquants'}), 400
        badge_label       = request.form.get('badge_label', 'Annonce').strip()
        badge_type        = request.form.get('badge_type', 'default').strip()
        auto_skip_seconds = int(request.form.get('auto_skip_seconds', 15))
        is_active         = int(request.form.get('is_active', 1))
        author            = session.get('username', 'Admin')
        now               = datetime.now().isoformat()
        images            = _save_announcement_images(request.files.getlist('images'))
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Erreur base de données'}), 500
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO announcements
                (title, content, badge_label, badge_type, start_date, end_date,
                 auto_skip_seconds, is_active, images, author, view_count, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,0,?,?)
        """, (title, content, badge_label, badge_type, start_date, end_date,
              auto_skip_seconds, is_active, json.dumps(images), author, now, now))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Annonce créée avec succès'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/announcements/update/<int:ann_id>', methods=['POST'])
@admin_required
def admin_announcement_update(ann_id):
    try:
        title      = request.form.get('title', '').strip()
        content    = request.form.get('content', '').strip()
        start_date = request.form.get('start_date', '').strip()
        end_date   = request.form.get('end_date', '').strip()
        if not all([title, content, start_date, end_date]):
            return jsonify({'success': False, 'message': 'Champs obligatoires manquants'}), 400
        badge_label       = request.form.get('badge_label', 'Annonce').strip()
        badge_type        = request.form.get('badge_type', 'default').strip()
        auto_skip_seconds = int(request.form.get('auto_skip_seconds', 15))
        is_active         = int(request.form.get('is_active', 1))
        now               = datetime.now().isoformat()
        keep_images = []
        try:
            keep_images = json.loads(request.form.get('keep_images', '[]'))
        except Exception:
            pass
        new_images  = _save_announcement_images(request.files.getlist('images'))
        all_images  = keep_images + new_images
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Erreur base de données'}), 500
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE announcements SET
                title=?, content=?, badge_label=?, badge_type=?,
                start_date=?, end_date=?, auto_skip_seconds=?,
                is_active=?, images=?, updated_at=?
            WHERE id=?
        """, (title, content, badge_label, badge_type, start_date, end_date,
              auto_skip_seconds, is_active, json.dumps(all_images), now, ann_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Annonce modifiée avec succès'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/announcements/toggle/<int:ann_id>', methods=['POST'])
@admin_required
def admin_announcement_toggle(ann_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT is_active FROM announcements WHERE id=?", (ann_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Annonce introuvable'}), 404
        new_state = 0 if row['is_active'] else 1
        cursor.execute("UPDATE announcements SET is_active=?, updated_at=? WHERE id=?",
                       (new_state, datetime.now().isoformat(), ann_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'is_active': new_state})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/announcements/delete/<int:ann_id>', methods=['POST'])
@admin_required
def admin_announcement_delete(ann_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT images FROM announcements WHERE id=?", (ann_id,))
        row = cursor.fetchone()
        if row:
            try:
                for img_path in json.loads(row['images'] or '[]'):
                    full = os.path.join(app.static_folder, img_path.replace('/static/', ''))
                    if os.path.exists(full):
                        os.remove(full)
            except Exception:
                pass
            cursor.execute("DELETE FROM announcements WHERE id=?", (ann_id,))
            conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def _save_announcement_images(file_list):
    """Sauvegarde les images uploadées pour les annonces."""
    saved = []
    upload_folder = os.path.join(app.static_folder, 'uploads', 'announcements')
    os.makedirs(upload_folder, exist_ok=True)
    allowed = {'png', 'jpg', 'jpeg', 'webp', 'gif'}
    for f in file_list:
        if f and f.filename and '.' in f.filename:
            ext = f.filename.rsplit('.', 1)[1].lower()
            if ext in allowed:
                filename = f"ann_{int(datetime.now().timestamp()*1000)}_{len(saved)}.{ext}"
                f.save(os.path.join(upload_folder, filename))
                saved.append(f"/static/uploads/announcements/{filename}")
    return saved


# ════════════════════════════════════════════════════════════
# FIN SPLASH SCREEN ANNONCES
# ════════════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_panel():
    # Double sécurité admin — vérifier le mot de passe secondaire
    if not session.get('admin_secondary_verified'):
        return redirect(url_for('admin_secondary_verify'))
    conn = get_db_connection()
    users, stats = [], {}
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id,username,email,role,status,shop_access,shop_permissions,created_at,last_login FROM users ORDER BY created_at DESC")
        users = [dict(r) for r in cursor.fetchall()]
        cursor.execute("SELECT COUNT(*) FROM users");               stats['total_users']      = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM users WHERE status='active'"); stats['active_users'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM users WHERE role IN ('admin','superadmin')"); stats['admins'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM financial_transactions"); stats['total_transactions'] = cursor.fetchone()[0]
        conn.close()
    return render_template('admin.html', users=users, stats=stats,
                           current_role=session.get('role'), token=ADMIN_SECRET_TOKEN)

@app.route('/admin/secondary-verify', methods=['GET', 'POST'])
@admin_required
def admin_secondary_verify():
    """Double sécurité admin — mot de passe secondaire Kengni@fablo12"""
    error = None
    if request.method == 'POST':
        pwd = (request.get_json(force=True, silent=True) or request.form).get('secondary_password', '')
        # Compteur de tentatives
        session['admin_sec_attempts'] = session.get('admin_sec_attempts', 0) + 1
        if session['admin_sec_attempts'] > 3:
            # Trop de tentatives — déconnexion forcée
            session.clear()
            if request.is_json:
                return jsonify({'success': False, 'message': 'Trop de tentatives — déconnexion'}), 403
            flash('Trop de tentatives incorrectes. Session terminée.', 'danger')
            return redirect(url_for('login'))
        if pwd == ADMIN_SECONDARY_PASSWORD:
            session['admin_secondary_verified'] = True
            session['admin_sec_attempts'] = 0
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('admin_panel')})
            return redirect(url_for('admin_panel'))
        else:
            remaining = 3 - session['admin_sec_attempts']
            error = f'Mot de passe secondaire incorrect. {remaining} tentative(s) restante(s).'
            if request.is_json:
                return jsonify({'success': False, 'message': error}), 401
    return render_template('admin_secondary.html', error=error)

@app.route('/admin/create-user', methods=['POST'])
@admin_required
def admin_create_user():
    data = request.get_json(force=True, silent=True)
    username,email,password = data.get('username','').strip(), data.get('email','').strip(), data.get('password','').strip()
    role, status = data.get('role','user'), data.get('status','active')
    shop_access = 1 if data.get('shop_access', False) else 0
    shop_perms = data.get('shop_permissions', {'add': False, 'edit': False, 'delete': False})
    allowed = ['viewer','user','editor','admin']
    if session.get('role')=='superadmin': allowed.append('superadmin')
    if not all([username,email,password]): return jsonify({'success':False,'message':'Tous les champs sont requis'}),400
    if role not in allowed: return jsonify({'success':False,'message':'Rôle non autorisé'}),403
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE email=?", (email,))
        if cursor.fetchone(): conn.close(); return jsonify({'success':False,'message':'Email déjà utilisé'}),409
        cursor.execute("INSERT INTO users (username,email,password,role,status,shop_access,shop_permissions,created_at) VALUES (?,?,?,?,?,?,?,?)",
                       (username,email,generate_password_hash(password),role,status,shop_access,json.dumps(shop_perms),datetime.now().isoformat()))
        conn.commit(); new_id=cursor.lastrowid; conn.close()
        return jsonify({'success':True,'message':f'Compte créé (ID {new_id})','id':new_id})
    return jsonify({'success':False,'message':'Erreur DB'}),500

@app.route('/admin/update-user/<int:user_id>', methods=['POST'])
@admin_required
def admin_update_user(user_id):
    data = request.get_json(force=True, silent=True)
    role, status = data.get('role'), data.get('status')
    shop_access = 1 if data.get('shop_access', False) else 0
    shop_perms = data.get('shop_permissions')
    allowed = ['viewer','user','editor','admin']
    if session.get('role')=='superadmin': allowed.append('superadmin')
    if role and role not in allowed: return jsonify({'success':False,'message':'Rôle non autorisé'}),403
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        if role:   cursor.execute("UPDATE users SET role=?,updated_at=? WHERE id=?",   (role,   datetime.now().isoformat(), user_id))
        if status: cursor.execute("UPDATE users SET status=?,updated_at=? WHERE id=?", (status, datetime.now().isoformat(), user_id))
        cursor.execute("UPDATE users SET shop_access=?,shop_permissions=?,updated_at=? WHERE id=?",
                       (shop_access, json.dumps(shop_perms) if shop_perms else '{}', datetime.now().isoformat(), user_id))
        conn.commit(); conn.close()
        return jsonify({'success':True,'message':'Utilisateur mis à jour'})
    return jsonify({'success':False,'message':'Erreur DB'}),500

@app.route('/admin/reset-password/<int:user_id>', methods=['POST'])
@admin_required
def admin_reset_password(user_id):
    data = request.get_json(force=True, silent=True)
    password = data.get('password','').strip()
    if len(password)<6: return jsonify({'success':False,'message':'Mot de passe trop court'}),400
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET password=?,updated_at=? WHERE id=?",
                       (generate_password_hash(password), datetime.now().isoformat(), user_id))
        conn.commit(); conn.close()
        return jsonify({'success':True,'message':'Mot de passe réinitialisé'})
    return jsonify({'success':False,'message':'Erreur DB'}),500

@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    if user_id==session['user_id']: return jsonify({'success':False,'message':'Impossible de supprimer votre propre compte'}),400
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit(); conn.close()
        return jsonify({'success':True,'message':'Utilisateur supprimé'})
    return jsonify({'success':False,'message':'Erreur DB'}),500

@app.route('/admin/update-permissions/<int:user_id>', methods=['POST'])
@admin_required
def admin_update_permissions(user_id):
    """Set allowed pages for a user. Empty list = blocked everywhere, None = all access."""
    data = request.get_json(force=True, silent=True) or {}
    pages = data.get('pages')  # list or None
    if pages is not None:
        # Validate
        pages = [p for p in pages if p in ALL_USER_PAGES]
        pages_json = json.dumps(pages)
    else:
        pages_json = None  # NULL = unrestricted
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET allowed_pages=?,updated_at=? WHERE id=?",
                       (pages_json, datetime.now().isoformat(), user_id))
        conn.commit(); conn.close()
        return jsonify({'success': True, 'message': 'Permissions mises à jour'})
    return jsonify({'success': False, 'message': 'Erreur DB'}), 500

@app.route('/admin/get-shop-permissions/<int:user_id>')
@admin_required
def admin_get_shop_permissions(user_id):
    conn = get_db_connection()
    if conn:
        row = conn.execute("SELECT shop_access, shop_permissions FROM users WHERE id=?", (user_id,)).fetchone()
        conn.close()
        if row:
            perms = json.loads(row['shop_permissions']) if row['shop_permissions'] else {}
            return jsonify({'access': bool(row['shop_access']), 'add': perms.get('add', False),
                            'edit': perms.get('edit', False), 'delete': perms.get('delete', False)})
    return jsonify({'access': False, 'add': False, 'edit': False, 'delete': False})

@app.route('/admin/get-permissions/<int:user_id>')
@admin_required
def admin_get_permissions(user_id):
    """Get allowed pages for a user."""
    allowed = get_user_allowed_pages(user_id)
    return jsonify({'success': True, 'pages': allowed})


@app.route('/api/admin/users')
@admin_required
def api_admin_users():
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id,username,email,role,status,last_login FROM users ORDER BY username")
        users = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'success':True,'users':users})
    return jsonify({'success':False}),500


# ══════════════════════════════════════════════════════════════
# MODULE TRAINING — Gestion des cours de formation
# ══════════════════════════════════════════════════════════════

def detect_thumbnail(url):
    """Détecte automatiquement la miniature selon la plateforme."""
    if not url:
        return '/static/img/course_default.svg'
    url_lower = url.lower()
    if 'claude.ai' in url_lower:
        return '/static/img/claude_thumb.svg'
    if 'chat.openai.com' in url_lower or 'chatgpt.com' in url_lower:
        return '/static/img/chatgpt_thumb.svg'
    yt = re.search(r'(?:v=|youtu\.be/)([A-Za-z0-9_-]{11})', url)
    if yt:
        return f'https://img.youtube.com/vi/{yt.group(1)}/hqdefault.jpg'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            html = resp.read(30000).decode('utf-8', errors='ignore')
        og = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\'](.*?)["\']', html)
        if og:
            return og.group(1)
        og2 = re.search(r'<meta[^>]+content=["\'](.*?)["\'][^>]+property=["\']og:image["\']', html)
        if og2:
            return og2.group(1)
    except Exception:
        pass
    return '/static/img/course_default.svg'


@app.route('/training')
def training():
    """Page de gestion des cours de formation — contenu visible par tous."""
    empty_days = {d: [] for d in ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi','Dimanche','Non défini']}
    conn = get_db_connection()
    if not conn:
        return render_template('training.html', courses_by_day=empty_days, stats={})
    cursor = conn.cursor()
    # Charger TOUS les cours de TOUS les utilisateurs + nom de l'auteur
    cursor.execute('''
        SELECT tc.*, u.username AS author_name, u.email AS author_email
        FROM training_courses tc
        LEFT JOIN users u ON tc.user_id = u.id
        ORDER BY tc.day_of_week, tc.created_at DESC
    ''')
    courses = [dict(r) for r in cursor.fetchall()]
    conn.close()

    days_order = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche', 'Non défini']
    courses_by_day = {d: [] for d in days_order}
    for c in courses:
        day = c.get('day_of_week', 'Non défini')
        if day not in days_order:
            day = 'Non défini'
        # Parse position_images JSON
        try:
            c['position_images'] = json.loads(c.get('position_images') or '[]')
        except Exception:
            c['position_images'] = []
        courses_by_day[day].append(c)

    stats = {
        'total': len(courses),
        'published': sum(1 for c in courses if c['is_published']),
        'scheduled': sum(1 for c in courses if c.get('scheduled_date')),
        'total_duration': sum((c['duration_minutes'] or 0) for c in courses),
    }
    return render_template('training.html', courses_by_day=courses_by_day, stats=stats)


@app.route('/training/add', methods=['POST'])
@login_required
def training_add():
    url = request.form.get('course_url', '').strip()
    thumbnail = request.form.get('thumbnail_url', '').strip() or detect_thumbnail(url)

    # Handle position images upload
    position_images = []
    for key in sorted(request.files.keys()):
        if key.startswith('position_img_') and not key.endswith('_caption'):
            file = request.files[key]
            if file and file.filename:
                fname = secure_filename(f"pos_{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(fpath)
                position_images.append(f'/static/uploads/{fname}')

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO training_courses
        (user_id, title, description, course_url, thumbnail_url, category, level,
         day_of_week, scheduled_date, duration_minutes, tags, is_published,
         participant_names, analyses, strategies, position_images, time_start, time_end, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        session['user_id'],
        request.form.get('title', 'Sans titre'),
        request.form.get('description', ''),
        url,
        thumbnail,
        request.form.get('category', 'Général'),
        request.form.get('level', 'debutant'),
        request.form.get('day_of_week', 'Non défini'),
        request.form.get('scheduled_date', ''),
        int(request.form.get('duration_minutes', 0) or 0),
        request.form.get('tags', ''),
        1 if request.form.get('is_published') else 0,
        request.form.get('participants', ''),
        request.form.get('analyses', ''),
        request.form.get('strategies', ''),
        json.dumps(position_images),
        request.form.get('time_start', ''),
        request.form.get('time_end', ''),
        datetime.now().isoformat()
    ))
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': new_id, 'thumbnail': thumbnail})


@app.route('/training/update/<int:cid>', methods=['POST'])
@login_required
def training_update(cid):
    url = request.form.get('course_url', '').strip()
    thumbnail = request.form.get('thumbnail_url', '').strip() or detect_thumbnail(url)

    # Handle new position images
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute('SELECT position_images FROM training_courses WHERE id=? AND user_id=?', (cid, session['user_id']))
    row = cursor.fetchone()
    try:
        existing_images = json.loads(row['position_images'] if row else '[]') or []
    except Exception:
        existing_images = []

    # Remove deleted images
    imgs_to_delete = request.form.getlist('delete_images')
    existing_images = [img for img in existing_images if img not in imgs_to_delete]

    # Add new images
    for key in sorted(request.files.keys()):
        if key.startswith('position_img_') and not key.endswith('_caption'):
            file = request.files[key]
            if file and file.filename:
                fname = secure_filename(f"pos_{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(fpath)
                existing_images.append(f'/static/uploads/{fname}')

    cursor.execute('''
        UPDATE training_courses SET
            title=?, description=?, course_url=?, thumbnail_url=?,
            category=?, level=?, day_of_week=?, scheduled_date=?,
            duration_minutes=?, tags=?, is_published=?,
            participant_names=?, analyses=?, strategies=?, position_images=?,
            time_start=?, time_end=?, updated_at=?
        WHERE id=? AND user_id=?
    ''', (
        request.form.get('title'), request.form.get('description'), url, thumbnail,
        request.form.get('category'), request.form.get('level'), request.form.get('day_of_week'),
        request.form.get('scheduled_date'), int(request.form.get('duration_minutes', 0) or 0),
        request.form.get('tags'), 1 if request.form.get('is_published') else 0,
        request.form.get('participants', ''),
        request.form.get('analyses', ''),
        request.form.get('strategies', ''),
        json.dumps(existing_images),
        request.form.get('time_start', ''),
        request.form.get('time_end', ''),
        datetime.now().isoformat(), cid, session['user_id']
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'thumbnail': thumbnail})


@app.route('/training/delete/<int:cid>', methods=['POST', 'DELETE'])
@login_required
def training_delete(cid):
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute('DELETE FROM training_courses WHERE id=? AND user_id=?', (cid, session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/training/fetch-thumb', methods=['POST'])
@login_required
def training_fetch_thumb():
    data = request.get_json(force=True, silent=True) or {}
    url = data.get('url', '')
    thumb = detect_thumbnail(url)
    return jsonify({'thumbnail': thumb})



# ══════════════════════════════════════════════════════════════
# MODULE KENGNI TRADING ACADEMY — Inscriptions & Gestion Leads
# ══════════════════════════════════════════════════════════════

@app.route('/inscription-trading', methods=['GET'])
def training_registration():
    """Page d'inscription publique à la Kengni Trading Academy."""
    success = request.args.get('success')
    wa      = request.args.get('wa', '')
    return render_template('inscription_trading.html', success=success, wa=wa)


@app.route('/inscription-trading', methods=['POST'])
def register_trading_lead():
    """Enregistre un nouveau lead avec vérification de doublon."""
    from urllib.parse import quote

    full_name      = request.form.get('full_name', '').strip()
    email          = request.form.get('email', '').strip().lower()
    whatsapp       = request.form.get('whatsapp', '').strip()
    level_selected = request.form.get('level_selected', '').strip()
    capital        = request.form.get('capital', '').strip()
    objective      = request.form.get('objective', '').strip()
    source         = request.form.get('source', 'Non renseigné').strip()

    # Validation serveur
    errors = []
    if not full_name or len(full_name) < 2:
        errors.append("Le nom complet est requis.")
    if not email or '@' not in email:
        errors.append("Adresse email invalide.")
    if not whatsapp or len(whatsapp.replace(' ', '')) < 8:
        errors.append("Numéro WhatsApp requis.")
    if not level_selected:
        errors.append("Veuillez choisir un niveau de formation.")

    if errors:
        for err in errors:
            flash(err, 'error')
        return redirect(url_for('training_registration'))

    conn = get_db_connection()
    if not conn:
        flash("Erreur de base de données. Veuillez réessayer.", 'error')
        return redirect(url_for('training_registration'))

    cursor = conn.cursor()

    # Vérification doublon
    cursor.execute(
        "SELECT id FROM training_leads WHERE email=? AND level_selected=?",
        (email, level_selected)
    )
    if cursor.fetchone():
        conn.close()
        flash(f"Vous êtes déjà inscrit(e) à la formation {level_selected} avec cet email. Notre équipe vous contactera bientôt !", 'error')
        return redirect(url_for('training_registration'))

    # Liaison user connecté si disponible
    user_id = session.get('user_id')

    cursor.execute('''
        INSERT INTO training_leads
        (full_name, email, whatsapp, level_selected, capital, objective, source, status, user_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, "Nouveau", ?, ?)
    ''', (
        full_name, email, whatsapp, level_selected,
        capital or None, objective or None, source,
        user_id, datetime.now().isoformat()
    ))
    conn.commit()
    conn.close()

    flash(f"Inscription confirmée ! Nous vous contacterons sur WhatsApp très bientôt. 🎉", 'success')
    return redirect(url_for('training_registration', success=1, wa=whatsapp))


@app.route('/admin/leads')
@login_required
@admin_required
def admin_leads():
    """Panneau admin : liste chronologique de tous les leads."""
    conn = get_db_connection()
    leads = []
    stats = {'total': 0, 'nouveau': 0, 'contacte': 0, 'inscrit': 0, 'paye': 0}
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM training_leads ORDER BY created_at DESC")
        leads = [dict(r) for r in cursor.fetchall()]
        conn.close()
        stats['total']    = len(leads)
        stats['nouveau']  = sum(1 for l in leads if l['status'] == 'Nouveau')
        stats['contacte'] = sum(1 for l in leads if l['status'] == 'Contacté')
        stats['inscrit']  = sum(1 for l in leads if l['status'] == 'Inscrit')
        stats['paye']     = sum(1 for l in leads if l['status'] == 'Payé')
    return render_template('admin_leads.html', leads=leads, stats=stats)


@app.route('/admin/leads/<int:lead_id>/status', methods=['POST'])
@login_required
@admin_required
def update_lead_status(lead_id):
    """Met à jour le statut d'un lead."""
    new_status = request.form.get('status', '').strip()
    if new_status not in ['Nouveau', 'Contacté', 'Inscrit', 'Payé']:
        flash("Statut invalide.", 'error')
        return redirect(url_for('admin_leads'))
    conn = get_db_connection()
    if conn:
        conn.execute("UPDATE training_leads SET status=? WHERE id=?", (new_status, lead_id))
        conn.commit()
        conn.close()
        flash(f"Statut mis à jour : {new_status}", 'success')
    return redirect(url_for('admin_leads'))


@app.route('/admin/leads/<int:lead_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lead(lead_id):
    """Supprime un lead."""
    conn = get_db_connection()
    if conn:
        conn.execute("DELETE FROM training_leads WHERE id=?", (lead_id,))
        conn.commit()
        conn.close()
        flash("Lead supprimé.", 'success')
    return redirect(url_for('admin_leads'))


def _send_sincire_email(lead: dict) -> bool:
    """Envoie l'email 'sincire' (invitation à payer) au prospect."""
    cfg      = GMAIL_CONFIG
    level    = lead.get('level_selected', 'Formation')
    name     = lead.get('full_name', 'Cher(e) prospect(e)')
    prospect_email = lead.get('email', '')
    prices   = FORMATION_PRICES.get(level, {'xaf': 50000, 'eur': 76})
    pay      = PAYMENT_INFO

    if not prospect_email:
        return False

    html = f'''<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/></head>
<body style="margin:0;padding:0;background:#0a0f1a;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:600px;margin:0 auto;padding:24px;">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#0d1b2a,#1a2a3a);border-radius:18px 18px 0 0;padding:36px 32px;text-align:center;border-bottom:3px solid #00d4aa;">
    <div style="font-size:3rem;margin-bottom:10px;">🎓</div>
    <h1 style="color:#fff;margin:0;font-size:22px;font-weight:800;">Kengni Trading Academy</h1>
    <p style="color:#00d4aa;margin:8px 0 0;font-size:14px;font-weight:600;">Votre place est réservée — Finalisez votre inscription</p>
  </div>

  <!-- Body -->
  <div style="background:#111827;padding:32px;border-radius:0 0 18px 18px;border:1px solid #1e2a3a;border-top:none;">

    <p style="color:#e0e0e0;font-size:15px;line-height:1.7;margin:0 0 20px;">
      Bonjour <strong style="color:#00d4aa;">{name}</strong>,<br><br>
      Merci pour votre intérêt pour la formation <strong style="color:#fff;">"{level}"</strong> !
      Votre dossier a été examiné et nous sommes ravis de vous confirmer que votre place est réservée.<br><br>
      Pour <strong style="color:#ffd700;">finaliser votre inscription</strong>, il vous suffit de procéder au règlement selon l'un des moyens ci-dessous.
    </p>

    <!-- Prix -->
    <div style="background:linear-gradient(135deg,rgba(0,212,170,.15),rgba(0,212,170,.05));border:1px solid rgba(0,212,170,.3);border-radius:12px;padding:20px;text-align:center;margin-bottom:24px;">
      <div style="font-size:.8rem;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;">Montant à régler — {level}</div>
      <div style="font-size:2rem;font-weight:800;color:#00d4aa;">{prices['xaf']:,} FCFA</div>
      <div style="font-size:.9rem;color:#888;margin-top:4px;">≈ {prices['eur']} EUR</div>
    </div>

    <!-- Méthodes de paiement -->
    <h3 style="color:#fff;font-size:15px;font-weight:700;margin:0 0 14px;">💳 Modes de paiement acceptés</h3>

    <!-- OM -->
    <div style="background:#0d1b2a;border-radius:12px;padding:16px 18px;margin-bottom:10px;display:flex;align-items:center;gap:14px;border-left:4px solid #ff6b00;">
      <div style="font-size:1.8rem;flex-shrink:0;">🟠</div>
      <div>
        <div style="color:#ff6b00;font-weight:700;font-size:13px;">Orange Money</div>
        <div style="color:#fff;font-size:1.1rem;font-weight:800;letter-spacing:1px;">{pay['orange_money']['numero']}</div>
        <div style="color:#888;font-size:12px;">Au nom de : {pay['orange_money']['nom']}</div>
      </div>
    </div>

    <!-- MTN -->
    <div style="background:#0d1b2a;border-radius:12px;padding:16px 18px;margin-bottom:10px;display:flex;align-items:center;gap:14px;border-left:4px solid #ffd700;">
      <div style="font-size:1.8rem;flex-shrink:0;">🟡</div>
      <div>
        <div style="color:#ffd700;font-weight:700;font-size:13px;">MTN Mobile Money</div>
        <div style="color:#fff;font-size:1.1rem;font-weight:800;letter-spacing:1px;">{pay['mtn_money']['numero']}</div>
        <div style="color:#888;font-size:12px;">Au nom de : {pay['mtn_money']['nom']}</div>
      </div>
    </div>

    <!-- PayPal -->
    <div style="background:#0d1b2a;border-radius:12px;padding:16px 18px;margin-bottom:10px;display:flex;align-items:center;gap:14px;border-left:4px solid #009cde;">
      <div style="font-size:1.8rem;flex-shrink:0;">🔵</div>
      <div>
        <div style="color:#009cde;font-weight:700;font-size:13px;">PayPal</div>
        <div style="color:#fff;font-size:1rem;font-weight:700;">{pay['paypal']['adresse']}</div>
        <div style="color:#888;font-size:12px;">Envoyer en "Amis &amp; Famille" pour éviter les frais</div>
      </div>
    </div>

    <!-- Crypto -->
    <div style="background:#0d1b2a;border-radius:12px;padding:16px 18px;margin-bottom:24px;display:flex;align-items:center;gap:14px;border-left:4px solid #f7931a;">
      <div style="font-size:1.8rem;flex-shrink:0;">₿</div>
      <div>
        <div style="color:#f7931a;font-weight:700;font-size:13px;">Crypto (BTC, USDT, ETH…)</div>
        <div style="color:#fff;font-size:1rem;font-weight:700;">{pay['crypto']['adresse']}</div>
        <div style="color:#888;font-size:12px;">Contactez-nous par email pour recevoir l'adresse de wallet</div>
      </div>
    </div>

    <!-- Instructions -->
    <div style="background:rgba(255,215,0,.07);border:1px solid rgba(255,215,0,.25);border-radius:12px;padding:16px 18px;margin-bottom:24px;">
      <p style="color:#ffd700;font-weight:700;font-size:13px;margin:0 0 8px;">📋 Après votre paiement</p>
      <ol style="color:#aaa;font-size:13px;margin:0;padding-left:18px;line-height:2;">
        <li>Envoyez la capture d'écran de votre paiement sur WhatsApp</li>
        <li>Numéro WhatsApp : <strong style="color:#fff;">+237 695 072 759</strong></li>
        <li>Votre accès à la formation sera activé sous 24h</li>
      </ol>
    </div>

    <!-- CTA -->
    <div style="text-align:center;margin-bottom:16px;">
      <a href="https://wa.me/237695072759?text=Bonjour%2C%20j%27ai%20effectu%C3%A9%20mon%20paiement%20pour%20la%20formation%20{level.replace(' ','%20')}"
         style="background:linear-gradient(135deg,#00d4aa,#00ff88);color:#000;font-weight:800;font-size:14px;padding:14px 32px;border-radius:10px;text-decoration:none;display:inline-block;box-shadow:0 4px 16px rgba(0,212,170,.4);">
        📲 Confirmer mon paiement sur WhatsApp
      </a>
    </div>

    <!-- Footer -->
    <div style="border-top:1px solid #1e2a3a;padding-top:16px;text-align:center;margin-top:8px;">
      <p style="color:#444;font-size:11px;margin:0;">Kengni Trading Academy · fabrice.kengni12@gmail.com</p>
      <p style="color:#333;font-size:10px;margin:4px 0 0;">Cet email vous a été envoyé suite à votre inscription sur notre plateforme.</p>
    </div>
  </div>
</div>
</body></html>'''

    text = (f"Bonjour {name},\n\n"
            f"Votre place pour la formation \"{level}\" est réservée !\n\n"
            f"MONTANT : {prices['xaf']:,} FCFA (≈ {prices['eur']} EUR)\n\n"
            f"PAIEMENT :\n"
            f"• Orange Money : {pay['orange_money']['numero']}\n"
            f"• MTN MoMo : {pay['mtn_money']['numero']}\n"
            f"• PayPal / Crypto : {pay['paypal']['adresse']}\n\n"
            f"Après paiement, envoyez la capture sur WhatsApp : +237 695 072 759\n\n"
            f"— Kengni Trading Academy")

    if not cfg.get('smtp_password'):
        print("[Sincire] ❌ GMAIL_PASSWORD manquant — configurez la variable d'environnement sur PythonAnywhere")
        return False

    msg = MIMEMultipart('alternative')
    msg['Subject'] = f"🎓 Finalisez votre inscription — {level} | Kengni Trading Academy"
    msg['From']    = f"Kengni Trading Academy <{cfg['sender_email']}>"
    msg['To']      = prospect_email
    msg['Reply-To'] = cfg['sender_email']
    msg.attach(MIMEText(text, 'plain', 'utf-8'))
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    for attempt in range(1, 4):  # 3 tentatives
        try:
            with smtplib.SMTP(cfg['smtp_host'], cfg['smtp_port'], timeout=15) as s:
                s.ehlo()
                s.starttls()
                s.ehlo()
                s.login(cfg['sender_email'], cfg['smtp_password'])
                s.sendmail(cfg['sender_email'], prospect_email, msg.as_string())
            print(f"[Sincire] ✅ Email envoyé à {prospect_email} (tentative {attempt})")
            return True
        except smtplib.SMTPAuthenticationError:
            print("[Sincire] ❌ Authentification Gmail échouée — vérifiez le mot de passe d'application sur PythonAnywhere")
            return False  # Inutile de réessayer si le mot de passe est faux
        except smtplib.SMTPException as e:
            print(f"[Sincire] ⚠️ Tentative {attempt}/3 échouée : {e}")
            if attempt < 3:
                time.sleep(2 * attempt)
        except Exception as e:
            print(f"[Sincire] ❌ Erreur inattendue : {e}")
            return False

    print(f"[Sincire] ❌ Échec après 3 tentatives pour {prospect_email}")
    return False


@app.route('/admin/leads/<int:lead_id>/sincire', methods=['POST'])
@login_required
@admin_required
def sincire_lead(lead_id):
    """Envoie l'email de sinciration (invitation paiement) au prospect."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM training_leads WHERE id=?", (lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        return jsonify({'success': False, 'error': 'Lead introuvable'}), 404

    lead_dict = dict(lead)
    ok = _send_sincire_email(lead_dict)
    if ok:
        now_iso = datetime.now().isoformat()
        cursor.execute(
            "UPDATE training_leads SET sincire_sent_at=?, status='Contacté' WHERE id=?",
            (now_iso, lead_id)
        )
        conn.commit()
        conn.close()
        return jsonify({
            'success': True,
            'message': f"✅ Email sincire envoyé à {lead_dict.get('email')} !",
            'sent_at': now_iso[:16].replace('T', ' à '),
        })
    else:
        conn.close()
        return jsonify({
            'success': False,
            'error': "Échec de l'envoi — vérifiez la config Gmail.",
        }), 500


@app.route('/admin/leads/<int:lead_id>/update-payment', methods=['POST'])
@login_required
@admin_required
def update_lead_payment(lead_id):
    """Met à jour les infos de paiement d'un lead."""
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    conn.execute('''
        UPDATE training_leads
        SET payment_method=?, payment_ref=?, payment_status=?, amount_paid=?
        WHERE id=?
    ''', (
        data.get('payment_method',''),
        data.get('payment_ref',''),
        data.get('payment_status','En attente'),
        float(data.get('amount_paid', 0) or 0),
        lead_id
    ))
    conn.commit(); conn.close()
    return jsonify({'success': True})


@app.route('/api/export-leads')
@login_required
@admin_required
def export_leads():
    """Export CSV ou Excel des leads via Pandas."""
    fmt  = request.args.get('format', 'csv').lower()
    conn = get_db_connection()
    leads = []
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM training_leads ORDER BY created_at DESC")
        leads = [dict(r) for r in cursor.fetchall()]
        conn.close()

    data = [{
        'ID':        l['id'],
        'Nom':       l['full_name'],
        'Email':     l['email'],
        'WhatsApp':  l['whatsapp'],
        'Formation': l['level_selected'],
        'Capital':   l.get('capital') or '',
        'Objectif':  l.get('objective') or '',
        'Source':    l.get('source') or '',
        'Statut':    l['status'],
        'Date':      l['created_at'],
        'Lien WA':   f"https://wa.me/{l['whatsapp'].replace(' ','').replace('+','')}",
    } for l in leads]

    df = pd.DataFrame(data) if data else pd.DataFrame()
    output = BytesIO()
    ts = datetime.now().strftime('%Y%m%d_%H%M')

    if fmt == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Leads Trading')
            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                ws = writer.sheets['Leads Trading']
                fill = PatternFill("solid", fgColor="0D3B2D")
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = Font(bold=True, color="00C77A")
                    cell.alignment = Alignment(horizontal='center')
                for col in ws.columns:
                    w = max(len(str(c.value or '')) for c in col) + 4
                    ws.column_dimensions[col[0].column_letter].width = min(w, 40)
            except Exception:
                pass
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f'leads_trading_{ts}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f'leads_trading_{ts}.csv',
                         mimetype='text/csv')


@app.route('/api/leads/stats')
@login_required
@admin_required
def leads_stats_api():
    """Statistiques leads en JSON."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB'}), 500
    cursor = conn.cursor()
    cursor.execute("SELECT status, COUNT(*) as cnt FROM training_leads GROUP BY status")
    rows = {r['status']: r['cnt'] for r in cursor.fetchall()}
    cursor.execute("SELECT level_selected, COUNT(*) as cnt FROM training_leads GROUP BY level_selected")
    by_level = {r['level_selected']: r['cnt'] for r in cursor.fetchall()}
    total = sum(rows.values())
    paye  = rows.get('Payé', 0)
    conn.close()
    return jsonify({
        'total':           total,
        'par_statut':      rows,
        'par_formation':   by_level,
        'taux_conversion': round(paye / total * 100, 1) if total else 0,
    })


# ══════════════════════════════════════════════════════════════
# MODULE AGENDA — Email Gmail + Scheduler + Routes
# ══════════════════════════════════════════════════════════════

def _build_agenda_email_html(event: dict, minutes_before: int) -> str:
    """Construit le corps HTML de l'email de rappel."""
    etype  = event.get('event_type', 'personnel')
    cfg    = AGENDA_EVENT_COLORS.get(etype, AGENDA_EVENT_COLORS['personnel'])
    color  = cfg['bg']
    icon   = cfg['icon']
    label  = cfg['label']
    title  = event.get('title', '(Sans titre)')
    desc   = event.get('description') or ''
    loc    = event.get('location') or ''
    notes  = event.get('notes') or ''
    start  = (event.get('start_datetime') or '')[:16].replace('T', ' à ')
    end    = (event.get('end_datetime')   or '')[:16].replace('T', ' à ')

    if minutes_before >= 1440:
        remind_txt = f"{minutes_before // 1440} jour(s)"
    elif minutes_before >= 60:
        remind_txt = f"{minutes_before // 60} heure(s)"
    else:
        remind_txt = f"{minutes_before} minute(s)"

    loc_row   = f'<tr><td style="padding:8px 0;color:#888;font-size:13px;width:100px;">📍 Lieu</td><td style="padding:8px 0;color:#e0e0e0;font-size:13px;">{loc}</td></tr>' if loc else ''
    notes_blk = f'<div style="background:#1a2a3a;border-radius:8px;padding:16px;margin-top:18px;"><p style="color:#888;font-size:11px;margin:0 0 6px;">📝 Notes</p><p style="color:#ccc;font-size:13px;margin:0;">{notes}</p></div>' if notes else ''
    desc_blk  = f'<p style="color:#aaa;margin:4px 0 0;font-size:14px;">{desc}</p>' if desc else ''

    return f'''<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/></head>
<body style="margin:0;padding:0;background:#0a0f1a;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:580px;margin:0 auto;padding:24px;">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#0d1b2a,#1a2a3a);border-radius:16px 16px 0 0;padding:32px 28px;text-align:center;border-bottom:3px solid {color};">
    <div style="font-size:42px;margin-bottom:10px;">⏰</div>
    <h1 style="color:#fff;margin:0;font-size:22px;font-weight:800;letter-spacing:.3px;">Rappel d'Agenda</h1>
    <p style="color:{color};margin:8px 0 0;font-size:14px;font-weight:600;">{icon} {label} · dans {remind_txt}</p>
  </div>

  <!-- Body -->
  <div style="background:#111827;padding:28px;border-radius:0 0 16px 16px;border:1px solid #1e2a3a;border-top:none;">

    <!-- Event card -->
    <div style="background:linear-gradient(135deg,{color}18,{color}08);border:1px solid {color}44;border-left:5px solid {color};border-radius:10px;padding:20px;margin-bottom:22px;">
      <h2 style="color:#fff;margin:0 0 4px;font-size:20px;font-weight:700;">{title}</h2>
      {desc_blk}
    </div>

    <!-- Details table -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:4px;">
      <tr><td style="padding:8px 0;color:#888;font-size:13px;width:100px;">🕐 Début</td><td style="padding:8px 0;color:#e0e0e0;font-size:13px;font-weight:600;">{start}</td></tr>
      <tr><td style="padding:8px 0;color:#888;font-size:13px;">🏁 Fin</td><td style="padding:8px 0;color:#e0e0e0;font-size:13px;">{end}</td></tr>
      {loc_row}
    </table>

    {notes_blk}

    <!-- CTA Button -->
    <div style="text-align:center;margin:26px 0 16px;">
      <a href="http://localhost:5001/agenda" style="background:linear-gradient(135deg,{color},{color}bb);color:#000;font-weight:800;font-size:14px;padding:14px 36px;border-radius:10px;text-decoration:none;display:inline-block;letter-spacing:.5px;box-shadow:0 4px 16px {color}44;">
        📅 Ouvrir mon Agenda
      </a>
    </div>

    <!-- Footer -->
    <div style="border-top:1px solid #1e2a3a;padding-top:14px;text-align:center;">
      <p style="color:#333;font-size:11px;margin:0;">Kengni Finance · Rappel automatique · fabricekengni90@gmail.com</p>
    </div>
  </div>
</div>
</body></html>'''


def _send_agenda_email(event: dict, minutes_before: int) -> bool:
    """Envoie un email de rappel via Gmail SMTP avec retry automatique."""
    cfg = GMAIL_CONFIG

    if not cfg.get('smtp_password'):
        print("[Agenda] ❌ GMAIL_PASSWORD manquant — configurez la variable d'environnement sur PythonAnywhere")
        return False

    h = f"{'%dh' % (minutes_before//60) if minutes_before >= 60 else '%dmin' % minutes_before}"
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f"⏰ Rappel dans {h} : {event['title']}"
    msg['From']    = f"{cfg['sender_name']} <{cfg['sender_email']}>"
    msg['To']      = cfg['receiver_email']

    text = (f"RAPPEL — {event['title']}\n"
            f"Début   : {(event.get('start_datetime') or '')[:16]}\n"
            f"Fin     : {(event.get('end_datetime') or '')[:16]}\n"
            f"Lieu    : {event.get('location') or 'Non précisé'}\n\n"
            f"{event.get('description') or ''}\n\n---\nKengni Finance")
    msg.attach(MIMEText(text, 'plain', 'utf-8'))
    msg.attach(MIMEText(_build_agenda_email_html(event, minutes_before), 'html', 'utf-8'))

    for attempt in range(1, 4):  # 3 tentatives
        try:
            with smtplib.SMTP(cfg['smtp_host'], cfg['smtp_port'], timeout=15) as s:
                s.ehlo()
                s.starttls()
                s.ehlo()
                s.login(cfg['sender_email'], cfg['smtp_password'])
                s.sendmail(cfg['sender_email'], cfg['receiver_email'], msg.as_string())
            print(f"[Agenda] ✅ Email envoyé : {event['title']} ({minutes_before}min avant, tentative {attempt})")
            return True
        except smtplib.SMTPAuthenticationError:
            print("[Agenda] ❌ Auth Gmail échouée — vérifiez le mot de passe d'application sur PythonAnywhere")
            return False
        except smtplib.SMTPException as e:
            print(f"[Agenda] ⚠️ Tentative {attempt}/3 échouée : {e}")
            if attempt < 3:
                time.sleep(2 * attempt)
        except Exception as e:
            print(f"[Agenda] ❌ Erreur inattendue : {e}")
            return False

    print(f"[Agenda] ❌ Échec après 3 tentatives pour : {event['title']}")
    return False


def _agenda_check_reminders():
    """Vérifie et envoie les rappels toutes les 60 secondes (thread background)."""
    while True:
        try:
            now = datetime.now()
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            window = (now + timedelta(hours=48)).isoformat()

            cursor.execute('''
                SELECT * FROM agenda_events
                WHERE status = 'active'
                  AND start_datetime BETWEEN ? AND ?
                  AND (email_reminder = 1 OR app_reminder = 1)
            ''', (now.isoformat(), window))
            events = [dict(r) for r in cursor.fetchall()]

            for ev in events:
                try:
                    start_dt  = datetime.fromisoformat(ev['start_datetime'])
                    remind_at = start_dt - timedelta(minutes=ev['reminder_minutes'])
                    diff = abs((now - remind_at).total_seconds())
                    if diff > 90:
                        continue

                    # Anti-doublon
                    cursor.execute('''
                        SELECT id FROM agenda_reminders_sent
                        WHERE event_id = ? AND sent_at >= ?
                    ''', (ev['id'], (now - timedelta(minutes=3)).isoformat()))
                    if cursor.fetchone():
                        continue

                    # Email
                    if ev['email_reminder']:
                        ok = _send_agenda_email(ev, ev['reminder_minutes'])
                        if ok:
                            cursor.execute('INSERT INTO agenda_reminders_sent (event_id,sent_at,method) VALUES (?,?,?)',
                                           (ev['id'], now.isoformat(), 'email'))

                    # Notification in-app
                    if ev['app_reminder']:
                        h = ev['reminder_minutes']
                        label = f"{h//60}h" if h >= 60 else f"{h}min"
                        cursor.execute('''
                            INSERT INTO notifications (user_id, type, title, message, action_url, created_at)
                            VALUES (?,?,?,?,?,?)
                        ''', (ev['user_id'], 'warning' if h <= 15 else 'info',
                              f"⏰ Rappel dans {label} : {ev['title']}",
                              f"Votre événement commence à {ev['start_datetime'][11:16]}.",
                              '/agenda', now.isoformat()))
                        cursor.execute('INSERT INTO agenda_reminders_sent (event_id,sent_at,method) VALUES (?,?,?)',
                                       (ev['id'], now.isoformat(), 'app'))

                    conn.commit()
                except Exception as ex:
                    print(f"[Agenda] Erreur traitement event #{ev.get('id')}: {ex}")

            conn.close()
        except Exception as e:
            print(f"[Agenda Scheduler] Erreur : {e}")
        time.sleep(60)


def start_agenda_scheduler():
    """Lance le scheduler en thread daemon."""
    t = threading.Thread(target=_agenda_check_reminders, daemon=True, name='AgendaScheduler')
    t.start()
    print("✅ Agenda Scheduler démarré (Gmail → iCloud, toutes les 60s)")


# ── ROUTES AGENDA ──────────────────────────────────────────────────────────────

@app.route('/agenda')
@page_access_required('agenda')
def agenda():
    """Page principale de l'agenda."""
    user_id = session['user_id']
    now     = datetime.now()
    conn    = get_db_connection()
    events, next_event, stats_raw = [], None, {}

    if conn:
        cursor = conn.cursor()
        start_w = (now - timedelta(days=60)).strftime('%Y-%m-%d')
        end_w   = (now + timedelta(days=90)).strftime('%Y-%m-%d')

        cursor.execute('''
            SELECT * FROM agenda_events
            WHERE user_id=? AND status!='cancelled'
              AND DATE(start_datetime) BETWEEN ? AND ?
            ORDER BY start_datetime ASC
        ''', (user_id, start_w, end_w))
        events = [dict(r) for r in cursor.fetchall()]

        cursor.execute('''
            SELECT * FROM agenda_events
            WHERE user_id=? AND status='active' AND start_datetime >= ?
            ORDER BY start_datetime ASC LIMIT 1
        ''', (user_id, now.isoformat()))
        row = cursor.fetchone()
        next_event = dict(row) if row else None

        cursor.execute('''
            SELECT event_type, COUNT(*) as cnt FROM agenda_events
            WHERE user_id=? AND status='active' AND DATE(start_datetime)>=DATE('now')
            GROUP BY event_type
        ''', (user_id,))
        stats_raw = {r['event_type']: r['cnt'] for r in cursor.fetchall()}
        conn.close()

    # Préparer pour FullCalendar
    fc_events = []
    for e in events:
        cfg = AGENDA_EVENT_COLORS.get(e['event_type'], AGENDA_EVENT_COLORS['personnel'])
        fc_events.append({
            'id': e['id'], 'title': e['title'],
            'start': e['start_datetime'], 'end': e['end_datetime'],
            'allDay': bool(e['all_day']),
            'backgroundColor': cfg['bg'], 'borderColor': cfg['border'],
            'extendedProps': {
                'type': e['event_type'], 'icon': cfg['icon'],
                'description': e['description'] or '',
                'location': e['location'] or '',
                'notes': e['notes'] or '',
                'reminder': e['reminder_minutes'],
            }
        })

    return render_template('agenda.html',
        events_json  = json.dumps(fc_events),
        events       = events,
        next_event   = next_event,
        stats        = stats_raw,
        event_colors = AGENDA_EVENT_COLORS,
        now          = now,
    )


@app.route('/api/agenda/events', methods=['POST'])
@login_required
def agenda_create_event():
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}

    title      = (data.get('title') or '').strip()
    event_type = data.get('event_type', 'personnel')
    start_dt   = data.get('start_datetime', '')
    end_dt     = data.get('end_datetime', '')

    if not title or not start_dt or not end_dt:
        return jsonify({'success': False, 'error': 'Champs requis manquants'}), 400
    if event_type not in AGENDA_EVENT_COLORS:
        event_type = 'personnel'

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO agenda_events
        (user_id,title,description,event_type,start_datetime,end_datetime,
         all_day,recurrence,reminder_minutes,email_reminder,app_reminder,
         location,notes,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        user_id, title,
        data.get('description',''), event_type, start_dt, end_dt,
        int(data.get('all_day', 0)), data.get('recurrence','none'),
        int(data.get('reminder_minutes', 30)),
        int(data.get('email_reminder', 1)), int(data.get('app_reminder', 1)),
        data.get('location',''), data.get('notes',''),
        datetime.now().isoformat(), datetime.now().isoformat()
    ))
    event_id = cursor.lastrowid
    cursor.execute('''
        INSERT INTO notifications (user_id,type,title,message,action_url,created_at)
        VALUES (?,?,?,?,?,?)
    ''', (user_id, 'success',
          f"📅 Événement créé : {title}",
          f'Planifié le {start_dt[:16].replace("T"," à ")}. Rappel dans {data.get("reminder_minutes",30)} min.',
          '/agenda', datetime.now().isoformat()))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'event_id': event_id})


@app.route('/api/agenda/events')
@login_required
def agenda_get_events():
    user_id = session['user_id']
    start   = request.args.get('start', (datetime.now() - timedelta(days=30)).isoformat())
    end     = request.args.get('end',   (datetime.now() + timedelta(days=90)).isoformat())
    conn    = get_db_connection()
    result  = []
    if conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM agenda_events
            WHERE user_id=? AND status!='cancelled'
              AND start_datetime BETWEEN ? AND ?
            ORDER BY start_datetime ASC
        ''', (user_id, start[:10], end[:10]))
        for e in cursor.fetchall():
            e = dict(e)
            cfg = AGENDA_EVENT_COLORS.get(e['event_type'], AGENDA_EVENT_COLORS['personnel'])
            result.append({
                'id': e['id'], 'title': e['title'],
                'start': e['start_datetime'], 'end': e['end_datetime'],
                'allDay': bool(e['all_day']),
                'backgroundColor': cfg['bg'], 'borderColor': cfg['border'],
                'extendedProps': {'type': e['event_type'], 'icon': cfg['icon'],
                                  'description': e['description'] or '',
                                  'location': e['location'] or '',
                                  'reminder': e['reminder_minutes']}
            })
        conn.close()
    return jsonify(result)


@app.route('/api/agenda/events/<int:event_id>', methods=['PUT', 'PATCH'])
@login_required
def agenda_update_event(event_id):
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}
    conn    = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM agenda_events WHERE id=? AND user_id=?', (event_id, user_id))
    if not cursor.fetchone():
        conn.close(); return jsonify({'success': False, 'error': 'Non trouvé'}), 404

    fields = ['title','description','event_type','start_datetime','end_datetime',
              'all_day','recurrence','reminder_minutes','email_reminder',
              'app_reminder','location','notes','status']
    sets, vals = [], []
    for f in fields:
        if f in data:
            sets.append(f'{f}=?'); vals.append(data[f])
    if sets:
        vals += [datetime.now().isoformat(), event_id, user_id]
        cursor.execute(f"UPDATE agenda_events SET {', '.join(sets)}, updated_at=? WHERE id=? AND user_id=?", vals)
        conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/agenda/events/<int:event_id>', methods=['DELETE'])
@login_required
def agenda_delete_event(event_id):
    user_id = session['user_id']
    conn    = get_db_connection()
    if conn:
        conn.execute("UPDATE agenda_events SET status='cancelled' WHERE id=? AND user_id=?", (event_id, user_id))
        conn.commit(); conn.close()
    return jsonify({'success': True})


@app.route('/api/agenda/today')
@login_required
def agenda_today():
    user_id = session['user_id']
    today   = datetime.now().strftime('%Y-%m-%d')
    conn    = get_db_connection()
    events  = []
    if conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM agenda_events
            WHERE user_id=? AND status='active' AND DATE(start_datetime)=?
            ORDER BY start_datetime ASC
        ''', (user_id, today))
        events = [dict(r) for r in cursor.fetchall()]
        conn.close()
    return jsonify({'events': events, 'count': len(events), 'date': today})


@app.route('/api/agenda/test-email', methods=['POST'])
@login_required
def agenda_test_email():
    """Envoie un email de test pour vérifier la configuration Gmail."""
    now = datetime.now()
    fake_event = {
        'id':             0,
        'user_id':        session['user_id'],
        'title':          '✅ Test de configuration — Kengni Finance Agenda',
        'event_type':     'trading',
        'description':    'Ceci est un email de test envoyé depuis Kengni Finance pour vérifier que la configuration Gmail fonctionne correctement. Si vous recevez cet email, les rappels automatiques sont opérationnels !',
        'start_datetime': (now + timedelta(minutes=30)).strftime('%Y-%m-%dT%H:%M:%S'),
        'end_datetime':   (now + timedelta(minutes=90)).strftime('%Y-%m-%dT%H:%M:%S'),
        'location':       'Kengni Finance — http://localhost:5001',
        'notes':          f'Test effectué le {now.strftime("%d/%m/%Y à %H:%M")} par {session.get("username","admin")}.',
    }
    ok = _send_agenda_email(fake_event, 30)
    if ok:
        return jsonify({
            'success': True,
            'message': f'✅ Email envoyé avec succès à {GMAIL_CONFIG["receiver_email"]} !',
            'from':    GMAIL_CONFIG['sender_email'],
            'to':      GMAIL_CONFIG['receiver_email'],
        })
    else:
        return jsonify({
            'success':  False,
            'message':  '❌ Échec — le mot de passe d\'application Gmail est incorrect ou manquant.',
            'help':     'Générez un mot de passe sur https://myaccount.google.com/apppasswords et mettez-le dans GMAIL_CONFIG dans app.py.',
        }), 500




# ══════════════════════════════════════════════════════════════
# MODULE BLOC-NOTES — Notes, Tableau, Calculatrice & Mémo
# ══════════════════════════════════════════════════════════════

def _get_unread_count(user_id):
    """Retourne le nombre de notifications non lues."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT COUNT(*) as cnt FROM notifications WHERE user_id = ? AND is_read = 0",
            (user_id,)
        )
        result = cursor.fetchone()
        conn.close()
        return result['cnt'] if result else 0
    except Exception:
        return 0


@app.route('/bloc-notes')
@page_access_required('bloc_notes')
def bloc_notes():
    """Page principale du bloc-notes avec tableau et calculatrice."""
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM bloc_notes WHERE user_id = ? ORDER BY entry_date DESC, created_at DESC",
        (user_id,)
    )
    notes = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT memo FROM bloc_notes_memo WHERE user_id = ?", (user_id,))
    memo_row = cursor.fetchone()
    memo = memo_row['memo'] if memo_row else ''
    conn.close()

    return render_template(
        'bloc_notes.html',
        notes=notes,
        memo=memo,
        today=_date.today().isoformat(),
        unread_notifications=_get_unread_count(user_id),
    )


@app.route('/bloc-notes/add', methods=['POST'])
@login_required
def bloc_notes_add():
    """Ajoute une entrée au bloc-notes."""
    user_id    = session['user_id']
    label      = request.form.get('label', '').strip()
    note_text  = request.form.get('note_text', '').strip()
    amount_raw = request.form.get('amount', '').strip()
    entry_type = request.form.get('entry_type', 'info')
    currency   = request.form.get('currency', 'EUR')
    entry_date = request.form.get('entry_date', _date.today().isoformat())

    if not label:
        flash('📝 La description est obligatoire.', 'danger')
        return redirect(url_for('bloc_notes'))

    amount = float(amount_raw) if amount_raw else None

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """INSERT INTO bloc_notes (user_id, label, note_text, amount, currency, entry_type, entry_date)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (user_id, label, note_text or None, amount, currency, entry_type, entry_date)
    )
    conn.commit()
    conn.close()
    flash('✅ Entrée ajoutée avec succès !', 'success')
    return redirect(url_for('bloc_notes'))


@app.route('/bloc-notes/delete/<int:note_id>', methods=['POST'])
@login_required
def bloc_notes_delete(note_id):
    """Supprime une entrée du bloc-notes."""
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute(
        "DELETE FROM bloc_notes WHERE id = ? AND user_id = ?",
        (note_id, user_id)
    )
    conn.commit()
    conn.close()
    flash('🗑️ Entrée supprimée.', 'success')
    return redirect(url_for('bloc_notes'))


@app.route('/bloc-notes/clear', methods=['POST'])
@login_required
def bloc_notes_clear():
    """Efface toutes les entrées du bloc-notes de l'utilisateur."""
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute("DELETE FROM bloc_notes WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash('🗑️ Toutes les entrées ont été supprimées.', 'success')
    return redirect(url_for('bloc_notes'))


@app.route('/bloc-notes/memo', methods=['POST'])
@login_required
def bloc_notes_memo():
    """Sauvegarde le mémo rapide (appel AJAX JSON)."""
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}
    memo    = str(data.get('memo', ''))[:5000]

    conn = get_db_connection()
    conn.execute(
        """INSERT INTO bloc_notes_memo (user_id, memo, updated_at)
           VALUES (?, ?, datetime('now'))
           ON CONFLICT(user_id) DO UPDATE SET memo = excluded.memo, updated_at = excluded.updated_at""",
        (user_id, memo)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/bloc-notes/export')
@login_required
def bloc_notes_export():
    """Exporte les entrées du bloc-notes en CSV."""
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT entry_date, label, note_text, amount, currency, entry_type "
        "FROM bloc_notes WHERE user_id = ? ORDER BY entry_date DESC",
        (user_id,)
    )
    rows = cursor.fetchall()
    conn.close()

    from io import StringIO
    si = StringIO()
    si.write('﻿')  # BOM UTF-8 pour Excel
    writer = csv.writer(si)
    writer.writerow(['Date', 'Description', 'Note', 'Montant', 'Devise', 'Type'])
    for r in rows:
        writer.writerow([
            r['entry_date'], r['label'], r['note_text'] or '',
            r['amount'] or '', r['currency'], r['entry_type']
        ])

    from flask import Response
    return Response(
        si.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=bloc-notes-{_date.today()}.csv'}
    )

# ══════════════════════════════════════════════════════════════
# MODULE CHAT — Messagerie commune Kengni Finance
# ══════════════════════════════════════════════════════════════

CHAT_ENCRYPT_KEY = os.environ.get('CHAT_KEY', 'KengniFinance2024SecretChatKey!!').encode()[:32].ljust(32, b'0')

def _xor_encrypt(text: str) -> str:
    key  = CHAT_ENCRYPT_KEY
    data = text.encode('utf-8')
    enc  = bytes([data[i] ^ key[i % len(key)] for i in range(len(data))])
    return base64.b64encode(enc).decode()

def _xor_decrypt(token: str) -> str:
    try:
        key  = CHAT_ENCRYPT_KEY
        enc  = base64.b64decode(token.encode())
        data = bytes([enc[i] ^ key[i % len(key)] for i in range(len(enc))])
        return data.decode('utf-8')
    except Exception:
        return '🔒 Impossible de déchiffrer'

def init_chat_db():
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS chat_messages (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id      INTEGER NOT NULL,
        content      TEXT    NOT NULL,
        content_enc  TEXT,
        is_encrypted INTEGER DEFAULT 0,
        reply_to_id  INTEGER,
        reactions    TEXT    DEFAULT '{}',
        tags         TEXT    DEFAULT '[]',
        edited       INTEGER DEFAULT 0,
        deleted      INTEGER DEFAULT 0,
        created_at   TEXT    DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS chat_notifications_sent (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        message_id INTEGER NOT NULL,
        user_id    INTEGER NOT NULL,
        method     TEXT    NOT NULL,
        sent_at    TEXT    DEFAULT (datetime('now'))
    )''')
    # Colonne whatsapp sur users si absente
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN whatsapp TEXT")
    except Exception:
        pass
    conn.commit()
    conn.close()

init_chat_db()

def _chat_notify_email(tagged_user: dict, sender_name: str, message_preview: str, msg_id: int):
    cfg = GMAIL_CONFIG
    if not cfg.get('smtp_password') or not tagged_user.get('email'):
        return False
    chat_url = f"https://kengni1234.pythonanywhere.com/chat#{msg_id}"
    html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"/></head>
<body style="margin:0;padding:0;background:#0a0f1a;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:560px;margin:0 auto;padding:24px;">
  <div style="background:linear-gradient(135deg,#0d1b2a,#1a2a3a);border-radius:16px;padding:28px;border:1px solid #1e2a3a;">
    <div style="font-size:2rem;text-align:center;margin-bottom:8px;">💬</div>
    <h2 style="color:#00d4aa;text-align:center;margin:0 0 4px;font-size:18px;">Nouveau message</h2>
    <p style="color:#aaa;text-align:center;font-size:12px;margin:0 0 20px;">Kengni Finance — Messagerie</p>
    <div style="background:#111827;border-radius:12px;padding:16px;border-left:4px solid #00d4aa;margin-bottom:20px;">
      <p style="color:#888;font-size:12px;margin:0 0 6px;">
        <strong style="color:#fff;">{sender_name}</strong> vous a tagué(e)
      </p>
      <p style="color:#e0e0e0;font-size:14px;margin:0;line-height:1.6;">
        {message_preview[:200]}{'…' if len(message_preview)>200 else ''}
      </p>
    </div>
    <div style="text-align:center;">
      <a href="{chat_url}" style="background:linear-gradient(135deg,#00d4aa,#00ff88);color:#000;font-weight:800;
         font-size:14px;padding:12px 28px;border-radius:10px;text-decoration:none;display:inline-block;">
        📨 Voir le message
      </a>
    </div>
    <p style="color:#333;font-size:10px;text-align:center;margin-top:16px;">Kengni Finance · Messagerie interne</p>
  </div>
</div></body></html>'''
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"💬 {sender_name} vous a tagué dans le chat — Kengni Finance"
        msg['From']    = f"Kengni Finance Chat <{cfg['sender_email']}>"
        msg['To']      = tagged_user['email']
        msg.attach(MIMEText(f"{sender_name} vous a tagué : {message_preview}", 'plain', 'utf-8'))
        msg.attach(MIMEText(html, 'html', 'utf-8'))
        for attempt in range(1, 4):
            try:
                with smtplib.SMTP(cfg['smtp_host'], cfg['smtp_port'], timeout=15) as s:
                    s.ehlo(); s.starttls(); s.ehlo()
                    s.login(cfg['sender_email'], cfg['smtp_password'])
                    s.sendmail(cfg['sender_email'], tagged_user['email'], msg.as_string())
                print(f"[Chat] ✅ Email tag envoyé à {tagged_user['email']}")
                return True
            except smtplib.SMTPAuthenticationError:
                return False
            except smtplib.SMTPException:
                if attempt < 3:
                    time.sleep(2 * attempt)
    except Exception as e:
        print(f"[Chat] ❌ Erreur email : {e}")
    return False

def _chat_whatsapp_link(tagged_user: dict, sender_name: str, message_preview: str, msg_id: int) -> str:
    import urllib.parse
    chat_url = f"https://kengni1234.pythonanywhere.com/chat#{msg_id}"
    text = f"💬 *Kengni Finance*\n{sender_name} vous a tagué :\n_{message_preview[:100]}_\n\n👉 {chat_url}"
    phone = (tagged_user.get('whatsapp') or '').replace(' ', '').replace('+', '')
    if phone:
        return f"https://wa.me/{phone}?text={urllib.parse.quote(text)}"
    return f"https://wa.me/?text={urllib.parse.quote(text)}"

def _get_all_members():
    conn = get_db_connection()
    if not conn:
        return []
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, email FROM users WHERE status='active' ORDER BY username")
    members = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return members

if 'chat' not in ALL_USER_PAGES:
    ALL_USER_PAGES.append('chat')

@app.route('/chat')
@page_access_required('chat')
def chat_page():
    members = _get_all_members()
    user_id = session['user_id']
    return render_template('chat.html',
                           members=members,
                           unread_notifications=_get_unread_count(user_id),
                           current_user_id=user_id,
                           current_username=session.get('username', ''))

@app.route('/api/chat/messages')
@login_required
def api_chat_messages():
    since_id = request.args.get('since', 0, type=int)
    limit    = min(request.args.get('limit', 50, type=int), 100)
    conn = get_db_connection()
    if not conn:
        return jsonify({'messages': []})
    cursor = conn.cursor()
    cursor.execute('''
        SELECT m.id, m.content, m.content_enc, m.is_encrypted,
               m.reply_to_id, m.reactions, m.tags, m.edited, m.deleted,
               m.created_at, u.username, u.id as user_id
        FROM chat_messages m
        JOIN users u ON u.id = m.user_id
        WHERE m.id > ? AND m.deleted = 0
        ORDER BY m.id ASC LIMIT ?
    ''', (since_id, limit))
    rows = []
    for r in cursor.fetchall():
        d = dict(r)
        try: d['reactions'] = json.loads(d['reactions'] or '{}')
        except: d['reactions'] = {}
        try: d['tags'] = json.loads(d['tags'] or '[]')
        except: d['tags'] = []
        rows.append(d)
    conn.close()
    return jsonify({'messages': rows})

@app.route('/api/chat/send', methods=['POST'])
@login_required
def api_chat_send():
    data       = request.get_json(force=True, silent=True) or {}
    content    = str(data.get('content', '')).strip()[:2000]
    reply_to   = data.get('reply_to_id')
    encrypt_it = bool(data.get('encrypt', False))
    tags       = data.get('tags', [])
    if not content:
        return jsonify({'success': False, 'error': 'Message vide'}), 400
    user_id     = session['user_id']
    sender_name = session.get('username', 'Membre')
    content_enc = _xor_encrypt(content) if encrypt_it else None
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO chat_messages (user_id, content, content_enc, is_encrypted, reply_to_id, tags)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (user_id,
          content if not encrypt_it else '🔒 Message chiffré',
          content_enc, 1 if encrypt_it else 0,
          reply_to, json.dumps(tags)))
    msg_id  = cursor.lastrowid
    preview = content[:120]
    for tid in tags:
        try:
            cursor.execute('''
                INSERT INTO notifications (user_id, type, title, message, action_url, created_at)
                VALUES (?,?,?,?,?,datetime('now'))
            ''', (int(tid), 'info', f"💬 {sender_name} vous a tagué dans le chat", preview, '/chat'))
        except Exception:
            pass
    conn.commit()
    if tags:
        cursor.execute("SELECT id, username, email, whatsapp FROM users WHERE id IN (%s)" %
                       ','.join('?' * len(tags)), [int(t) for t in tags])
        tagged_users = [dict(r) for r in cursor.fetchall()]
        conn.close()
        def _notify():
            for u in tagged_users:
                _chat_notify_email(u, sender_name, preview, msg_id)
        threading.Thread(target=_notify, daemon=True).start()
    else:
        conn.close()
    return jsonify({'success': True, 'id': msg_id})

@app.route('/api/chat/decrypt', methods=['POST'])
@login_required
def api_chat_decrypt():
    data   = request.get_json(force=True, silent=True) or {}
    msg_id = data.get('id')
    conn   = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute('SELECT content_enc, is_encrypted FROM chat_messages WHERE id=?', (msg_id,))
    row = cursor.fetchone()
    conn.close()
    if not row or not row['is_encrypted']:
        return jsonify({'success': False, 'error': 'Non chiffré'})
    return jsonify({'success': True, 'content': _xor_decrypt(row['content_enc'])})

@app.route('/api/chat/react', methods=['POST'])
@login_required
def api_chat_react():
    data    = request.get_json(force=True, silent=True) or {}
    msg_id  = data.get('id')
    emoji   = str(data.get('emoji', ''))[:8]
    user_id = session['user_id']
    if not msg_id or not emoji:
        return jsonify({'success': False}), 400
    conn   = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT reactions FROM chat_messages WHERE id=?', (msg_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False}), 404
    try: reactions = json.loads(row['reactions'] or '{}')
    except: reactions = {}
    users_reacted = reactions.get(emoji, [])
    if user_id in users_reacted:
        users_reacted.remove(user_id)
    else:
        users_reacted.append(user_id)
    if users_reacted:
        reactions[emoji] = users_reacted
    else:
        reactions.pop(emoji, None)
    cursor.execute('UPDATE chat_messages SET reactions=? WHERE id=?', (json.dumps(reactions), msg_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'reactions': reactions})

@app.route('/api/chat/delete/<int:msg_id>', methods=['POST'])
@login_required
def api_chat_delete(msg_id):
    user_id = session['user_id']
    conn    = get_db_connection()
    cursor  = conn.cursor()
    cursor.execute('SELECT user_id FROM chat_messages WHERE id=?', (msg_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False}), 404
    if row['user_id'] != user_id and session.get('role') not in ('admin', 'superadmin'):
        conn.close()
        return jsonify({'success': False, 'error': 'Interdit'}), 403
    cursor.execute('UPDATE chat_messages SET deleted=1 WHERE id=?', (msg_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/chat/members')
@login_required
def api_chat_members():
    return jsonify({'members': _get_all_members()})

@app.route('/api/chat/whatsapp-tag')
@login_required
def api_chat_whatsapp_tag():
    user_id = request.args.get('user_id', type=int)
    msg_id  = request.args.get('msg_id', type=int)
    preview = request.args.get('preview', '')
    sender  = session.get('username', 'Un membre')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT username, email, whatsapp FROM users WHERE id=?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'link': '#'})
    return jsonify({'link': _chat_whatsapp_link(dict(row), sender, preview, msg_id)})

# ── API Candles SMC Engine ────────────────────────────────────────────────────
import requests as _req_binance

@app.route('/api/candles')
@login_required
def api_candles():
    sym   = request.args.get('symbol', 'BTCUSDT').upper().strip()
    tf    = request.args.get('tf', '1h').lower()
    limit = min(int(request.args.get('limit', 300)), 500)
    if not any(sym.endswith(x) for x in ['USDT','BUSD','BTC','ETH','BNB','USD']):
        sym = sym + 'USDT'
    tf_map = {'1m':'1m','3m':'3m','5m':'5m','15m':'15m','30m':'30m','30':'30m','1h':'1h','60':'1h','2h':'2h','4h':'4h','240':'4h','1d':'1d','D':'1d','1w':'1w','W':'1w'}
    interval = tf_map.get(tf, '1h')
    for url in [
        f'https://fapi.binance.com/fapi/v1/klines?symbol={sym}&interval={interval}&limit={limit}',
        f'https://api.binance.com/api/v3/klines?symbol={sym}&interval={interval}&limit={limit}',
    ]:
        try:
            r = _req_binance.get(url, timeout=8)
            if r.status_code != 200: continue
            data = r.json()
            if not isinstance(data, list) or len(data) < 5: continue
            candles = [{'time':int(k[0])//1000,'open':float(k[1]),'high':float(k[2]),'low':float(k[3]),'close':float(k[4]),'vol':float(k[5])} for k in data]
            return jsonify({'symbol':sym,'tf':tf,'candles':candles})
        except: continue
    return jsonify({'error':f'{sym} introuvable'}), 404

# ── Aliases de routes pour compatibilité templates ────────────────────────────

@app.route('/journal-view')
@login_required
def journal():
    """Alias vers trading_journal pour compatibilité url_for('journal')"""
    return trading_journal()

@app.route('/actualites')
@login_required
def actualites():
    """Page actualités économiques avec données live"""
    import urllib.request as _ur
    import json as _json
    from datetime import datetime as _dt

    # ── BTC prix live (CoinGecko) ──────────────────────────────
    btc = {'price': None, 'change': None, 'high': None, 'low': None, 'vol': 0}
    try:
        url = 'https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd&include_24hr_change=true&include_24hr_vol=true&include_high_low=true'
        req = _ur.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with _ur.urlopen(req, timeout=5) as r:
            d = _json.loads(r.read())['bitcoin']
            btc = {
                'price':  d.get('usd', 0),
                'change': d.get('usd_24h_change', 0),
                'high':   d.get('usd_24h_high', 0),
                'low':    d.get('usd_24h_low', 0),
                'vol':    d.get('usd_24h_vol', 0),
            }
    except Exception:
        pass

    # ── Fear & Greed Index (Alternative.me) ────────────────────
    fear_greed = {'value': 50, 'label': 'Neutral'}
    try:
        req2 = _ur.Request('https://api.alternative.me/fng/?limit=1', headers={'User-Agent': 'Mozilla/5.0'})
        with _ur.urlopen(req2, timeout=5) as r:
            d2 = _json.loads(r.read())['data'][0]
            fear_greed = {
                'value': int(d2.get('value', 50)),
                'label': d2.get('value_classification', 'Neutral')
            }
    except Exception:
        pass

    # ── Articles crypto (CryptoPanic via RSS public) ────────────
    articles = []
    try:
        rss_url = 'https://cryptopanic.com/news/rss/'
        req3 = _ur.Request(rss_url, headers={'User-Agent': 'Mozilla/5.0'})
        with _ur.urlopen(req3, timeout=6) as r:
            import re as _re
            rss = r.read().decode('utf-8', errors='ignore')
            titles = _re.findall(r'<title><!\[CDATA\[(.*?)\]\]></title>', rss)
            links  = _re.findall(r'<link>(https?://[^<]+)</link>', rss)
            dates  = _re.findall(r'<pubDate>(.*?)</pubDate>', rss)
            for i, (t, l) in enumerate(zip(titles[1:], links[1:])):
                sentiment = 'bullish' if any(w in t.lower() for w in ['rise','pump','gain','bull','surge','high','break']) \
                       else 'bearish' if any(w in t.lower() for w in ['drop','fall','bear','crash','low','down','sell']) \
                       else 'neutral'
                articles.append({
                    'id':         i,
                    'title':      t,
                    'source_url': l,
                    'source':     'CryptoPanic',
                    'category':   'crypto',
                    'cat_label':  'Crypto',
                    'sentiment':  sentiment,
                    'score':      '+' if sentiment == 'bullish' else '-' if sentiment == 'bearish' else '~',
                    'prob_bull':  65 if sentiment == 'bullish' else 35,
                    'prob_bear':  35 if sentiment == 'bullish' else 65,
                    'image':      None,
                    'summary':    '',
                    'impact_court_terme':  '',
                    'impact_moyen_terme':  '',
                    'niveaux_cles':        '',
                    'recommandation':      '',
                    'date':       dates[i] if i < len(dates) else '',
                })
                if len(articles) >= 20:
                    break
    except Exception:
        pass

    # ── Calendrier économique (events statiques du jour) ────────
    calendar = []
    try:
        now_h = _dt.now().hour
        events_data = [
            {'time': '08:30', 'event': 'CPI USA (Inflation)', 'currency': 'USD', 'impact': 'high'},
            {'time': '10:00', 'event': 'Indice PMI Zone Euro', 'currency': 'EUR', 'impact': 'medium'},
            {'time': '14:30', 'event': 'Emploi Non-Agricole', 'currency': 'USD', 'impact': 'high'},
            {'time': '16:00', 'event': 'Discours Fed Reserve', 'currency': 'USD', 'impact': 'high'},
            {'time': '20:00', 'event': 'Réserves Pétrole EIA', 'currency': 'USD', 'impact': 'medium'},
        ]
        calendar = events_data
    except Exception:
        pass

    last_update = _dt.now().strftime('%d/%m/%Y %H:%M')

    # Compter les sentiments pour les stats
    bull_count = sum(1 for a in articles if a.get('sentiment') == 'bullish')
    bear_count = sum(1 for a in articles if a.get('sentiment') == 'bearish')
    neut_count = sum(1 for a in articles if a.get('sentiment') == 'neutral')

    return render_template('actualites.html',
        btc=btc,
        fear_greed=fear_greed,
        articles=articles,
        calendar=calendar,
        last_update=last_update,
        bull_count=bull_count,
        bear_count=bear_count,
        neut_count=neut_count,
    )

# ── Fin module Chat ────────────────────────────────────────────────────────────



# ═══════════════════════════════════════════════════════════════════
# MODULE DOCUMENTS & SURVEY — Ajout sans modification de l'existant
# ═══════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────
# INITIALISATION DES TABLES DOCUMENTS & SURVEY
# ─────────────────────────────────────────────────────────────────

def init_documents_db():
    """Crée les tables documents, document_purchases et survey_responses."""
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()

    # ── Table documents ──────────────────────────────────────────
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS documents (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        title          TEXT    NOT NULL,
        description    TEXT,
        category       TEXT    DEFAULT "Général",
        tags           TEXT    DEFAULT "[]",
        thumbnail_url  TEXT,
        doc_type       TEXT    DEFAULT "file"
                               CHECK(doc_type IN ("file","link","both")),
        file_path      TEXT,
        file_url       TEXT,
        link_url       TEXT,
        link_label     TEXT    DEFAULT "Ouvrir le lien",
        price          REAL    DEFAULT 0,
        currency       TEXT    DEFAULT "XAF",
        is_free        INTEGER DEFAULT 1,
        is_published   INTEGER DEFAULT 1,
        view_count     INTEGER DEFAULT 0,
        download_count INTEGER DEFAULT 0,
        created_by     INTEGER,
        created_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
        updated_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (created_by) REFERENCES users(id)
    )
    ''')

    # ── Table achats / demandes d'accès ─────────────────────────
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS document_purchases (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        document_id    INTEGER NOT NULL,
        user_id        INTEGER,
        buyer_email    TEXT,
        buyer_name     TEXT,
        payment_method TEXT,
        payment_ref    TEXT,
        amount         REAL    DEFAULT 0,
        currency       TEXT    DEFAULT "XAF",
        status         TEXT    DEFAULT "pending"
                               CHECK(status IN ("pending","confirmed","rejected")),
        notes          TEXT,
        created_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
        updated_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (document_id) REFERENCES documents(id),
        FOREIGN KEY (user_id)     REFERENCES users(id)
    )
    ''')

    # ── Table sondage ────────────────────────────────────────────
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS survey_responses (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id  TEXT,
        answers     TEXT    NOT NULL,
        ip_address  TEXT,
        created_at  TEXT    DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # ── Migrations optionnelles ──────────────────────────────────
    for col, defn in [
        ('buyer_name',  'TEXT'),
        ('buyer_email', 'TEXT'),
        ('notes',       'TEXT'),
    ]:
        try:
            cursor.execute(f'ALTER TABLE document_purchases ADD COLUMN {col} {defn}')
        except Exception:
            pass

    conn.commit()
    conn.close()
    print("✅ Tables documents & survey initialisées")


# Appel immédiat à la définition du module
init_documents_db()

# ─────────────────────────────────────────────────────────────────
# HELPERS DOCUMENTS
# ─────────────────────────────────────────────────────────────────

DOCS_UPLOAD_FOLDER = os.path.join('static', 'uploads', 'documents')
os.makedirs(DOCS_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_DOC_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
    'txt', 'csv', 'zip', 'rar', 'png', 'jpg', 'jpeg', 'mp4', 'mp3'
}


def _allowed_doc(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_DOC_EXTENSIONS


def _doc_is_admin():
    return 'user_id' in session and session.get('role') in ('admin', 'superadmin')


def _get_purchased_ids(user_id=None):
    if not user_id:
        return []
    conn = get_db_connection()
    if not conn:
        return []
    cursor = conn.cursor()
    cursor.execute(
        "SELECT document_id FROM document_purchases WHERE user_id=? AND status='confirmed'",
        (user_id,)
    )
    ids = [r[0] for r in cursor.fetchall()]
    conn.close()
    return ids


# ═══════════════════════════════════════════════════════════════════
# ROUTES DOCUMENTS — Boutique de ressources trading
# ═══════════════════════════════════════════════════════════════════

@app.route('/documents/access', methods=['GET', 'POST'])
def documents_access():
    """Page d'accès public à la boutique (identification par email)."""
    if request.method == 'POST':
        data  = request.get_json(force=True, silent=True) or request.form
        email = (data.get('email') or '').strip().lower()
        if email and '@' in email:
            session['doc_viewer_email'] = email
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('documents_shop')})
            return redirect(url_for('documents_shop'))
        if request.is_json:
            return jsonify({'success': False, 'message': 'Email invalide'}), 400
        return render_template('document_access.html', error='Email invalide')

    # Si déjà identifié → boutique directement
    if session.get('doc_viewer_email') or _doc_is_admin() or 'user_id' in session:
        return redirect(url_for('documents_shop'))

    return render_template('document_access.html', error=None)


@app.route('/documents/access/logout', methods=['POST'])
def documents_access_logout():
    session.pop('doc_viewer_email', None)
    return jsonify({'success': True})


@app.route('/documents')
def documents_shop():
    """Page principale de la boutique de documents."""
    viewer_mail      = session.get('doc_viewer_email', '')
    is_admin         = _doc_is_admin()
    user_id          = session.get('user_id')
    search_q         = request.args.get('q', '').strip()
    current_category = request.args.get('category', '')

    if not is_admin and not user_id and not viewer_mail:
        return redirect(url_for('documents_access'))

    conn = get_db_connection()
    if not conn:
        return render_template('documents.html',
                               docs=[], categories=[],
                               is_admin=is_admin, purchases=[],
                               purchased_ids=[], stats={},
                               search_q=search_q,
                               current_category=current_category,
                               viewer_mail=viewer_mail,
                               payment_info=PAYMENT_INFO)

    cursor = conn.cursor()

    # ── Requête avec filtres ────────────────────────────────────
    query  = "SELECT * FROM documents WHERE 1=1"
    params = []
    if not is_admin:
        query += " AND is_published=1"
    if search_q:
        query += " AND (title LIKE ? OR description LIKE ? OR tags LIKE ?)"
        like   = f'%{search_q}%'
        params += [like, like, like]
    if current_category:
        query  += " AND category=?"
        params.append(current_category)
    query += " ORDER BY created_at DESC"

    cursor.execute(query, params)
    raw_docs = cursor.fetchall()

    docs = []
    for d in raw_docs:
        d = dict(d)
        try:
            d['tags'] = json.loads(d['tags'] or '[]')
        except Exception:
            d['tags'] = []
        docs.append(d)

    # ── Catégories ──────────────────────────────────────────────
    if is_admin:
        cursor.execute("SELECT DISTINCT category FROM documents ORDER BY category")
    else:
        cursor.execute("SELECT DISTINCT category FROM documents WHERE is_published=1 ORDER BY category")
    categories = [r[0] for r in cursor.fetchall()]

    # ── IDs achetés ─────────────────────────────────────────────
    purchased_ids = _get_purchased_ids(user_id)

    # ── Achats en attente (admin) ───────────────────────────────
    purchases = []
    if is_admin:
        cursor.execute('''
            SELECT dp.*, d.title as doc_title
            FROM document_purchases dp
            JOIN documents d ON d.id = dp.document_id
            ORDER BY dp.created_at DESC
        ''')
        purchases = [dict(r) for r in cursor.fetchall()]

    # ── Stats admin ─────────────────────────────────────────────
    stats = {}
    if is_admin:
        cursor.execute("SELECT COUNT(*) FROM documents")
        stats['total_docs'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM documents WHERE is_published=1")
        stats['published'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM document_purchases WHERE status='pending'")
        stats['pending_purchases'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM document_purchases WHERE status='confirmed'")
        stats['confirmed_purchases'] = cursor.fetchone()[0]
        cursor.execute("SELECT COALESCE(SUM(amount),0) FROM document_purchases WHERE status='confirmed'")
        stats['total_revenue'] = cursor.fetchone()[0]
        stats['active_sessions'] = 1

    conn.close()

    return render_template('documents.html',
                           docs=docs,
                           categories=categories,
                           is_admin=is_admin,
                           purchases=purchases,
                           purchased_ids=purchased_ids,
                           stats=stats,
                           search_q=search_q,
                           current_category=current_category,
                           viewer_mail=viewer_mail,
                           payment_info=PAYMENT_INFO)


@app.route('/documents/add', methods=['POST'])
@admin_required
def documents_add():
    """Ajouter un document (admin)."""
    title       = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    category    = request.form.get('category', 'Général').strip()
    tags_raw    = request.form.get('tags', '').strip()
    thumb       = request.form.get('thumbnail_url', '').strip()
    link_url    = request.form.get('link_url', '').strip()
    link_label  = request.form.get('link_label', 'Ouvrir le lien').strip()
    file_url    = request.form.get('file_url', '').strip()
    price       = float(request.form.get('price', 0) or 0)
    currency    = request.form.get('currency', 'XAF')
    is_pub      = 1 if request.form.get('is_published') else 0
    is_free     = 1 if price == 0 else 0
    tags        = json.dumps([t.strip() for t in tags_raw.split(',') if t.strip()])

    file_path = None
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename and _allowed_doc(f.filename):
            fname     = secure_filename(f"doc_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{f.filename}")
            file_path = os.path.join(DOCS_UPLOAD_FOLDER, fname)
            f.save(file_path)

    has_file = bool(file_path or file_url)
    has_link = bool(link_url)
    if has_file and has_link:
        doc_type = 'both'
    elif has_link:
        doc_type = 'link'
    else:
        doc_type = 'file'

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500

    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO documents
        (title, description, category, tags, thumbnail_url, doc_type,
         file_path, file_url, link_url, link_label,
         price, currency, is_free, is_published, created_by, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        title, description, category, tags, thumb or None, doc_type,
        file_path, file_url or None, link_url or None, link_label,
        price, currency, is_free, is_pub,
        session['user_id'], datetime.now().isoformat()
    ))
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': new_id})


@app.route('/documents/update/<int:doc_id>', methods=['POST'])
@admin_required
def documents_update(doc_id):
    """Modifier un document (admin)."""
    data        = request.get_json(force=True, silent=True) or {}
    title       = data.get('title', '').strip()
    description = data.get('description', '').strip()
    category    = data.get('category', 'Général').strip()
    tags_raw    = data.get('tags', [])
    tags        = json.dumps(tags_raw) if isinstance(tags_raw, list) else json.dumps(
        [t.strip() for t in str(tags_raw).split(',') if t.strip()])
    price       = float(data.get('price', 0) or 0)
    currency    = data.get('currency', 'XAF')
    is_free     = 1 if price == 0 else 0
    is_pub      = 1 if data.get('is_published') else 0
    thumb       = data.get('thumbnail_url', '') or None
    link_url    = data.get('link_url', '') or None
    link_label  = data.get('link_label', 'Ouvrir le lien')
    file_url    = data.get('file_url', '') or None

    has_file = bool(file_url)
    has_link = bool(link_url)
    if has_file and has_link:
        doc_type = 'both'
    elif has_link:
        doc_type = 'link'
    else:
        doc_type = 'file'

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE documents SET
            title=?, description=?, category=?, tags=?, thumbnail_url=?,
            doc_type=?, file_url=?, link_url=?, link_label=?,
            price=?, currency=?, is_free=?, is_published=?, updated_at=?
        WHERE id=?
    ''', (
        title, description, category, tags, thumb,
        doc_type, file_url, link_url, link_label,
        price, currency, is_free, is_pub,
        datetime.now().isoformat(), doc_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/documents/delete/<int:doc_id>', methods=['POST'])
@admin_required
def documents_delete(doc_id):
    """Supprimer un document (admin)."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute("SELECT file_path FROM documents WHERE id=?", (doc_id,))
    row = cursor.fetchone()
    if row and row['file_path'] and os.path.exists(row['file_path']):
        try:
            os.remove(row['file_path'])
        except Exception:
            pass
    cursor.execute("DELETE FROM documents WHERE id=?", (doc_id,))
    cursor.execute("DELETE FROM document_purchases WHERE document_id=?", (doc_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/documents/view/<int:doc_id>')
def documents_view(doc_id):
    """Incrémenter le compteur de vues."""
    conn = get_db_connection()
    if conn:
        conn.execute("UPDATE documents SET view_count=view_count+1 WHERE id=?", (doc_id,))
        conn.commit()
        conn.close()
    return jsonify({'success': True})


@app.route('/documents/download/<int:doc_id>')
def documents_download(doc_id):
    """Télécharger un document (vérifie les droits d'accès)."""
    conn = get_db_connection()
    if not conn:
        return "Erreur DB", 500
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM documents WHERE id=? AND is_published=1", (doc_id,))
    doc = cursor.fetchone()
    if not doc:
        conn.close()
        return "Document introuvable", 404

    doc       = dict(doc)
    is_admin  = _doc_is_admin()
    user_id   = session.get('user_id')
    purchased = _get_purchased_ids(user_id)

    if not doc['is_free'] and not is_admin and doc_id not in purchased:
        conn.close()
        return "Accès refusé — achetez ce document d'abord", 403

    conn.execute("UPDATE documents SET download_count=download_count+1 WHERE id=?", (doc_id,))
    conn.commit()
    conn.close()

    if doc.get('file_path') and os.path.exists(doc['file_path']):
        return send_file(doc['file_path'], as_attachment=True)
    if doc.get('file_url'):
        return redirect(doc['file_url'])
    return "Aucun fichier disponible", 404


@app.route('/documents/purchase/<int:doc_id>', methods=['POST'])
def documents_purchase(doc_id):
    """Créer une demande d'achat."""
    data           = request.get_json(force=True, silent=True) or {}
    payment_method = data.get('payment_method', '')
    payment_ref    = data.get('payment_ref', '')
    user_id        = session.get('user_id')
    buyer_email    = session.get('doc_viewer_email') or session.get('email', '')

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM documents WHERE id=?", (doc_id,))
    doc = cursor.fetchone()
    if not doc:
        conn.close()
        return jsonify({'success': False, 'error': 'Document introuvable'}), 404

    cursor.execute('''
        INSERT INTO document_purchases
        (document_id, user_id, buyer_email, payment_method, payment_ref,
         amount, currency, status, created_at)
        VALUES (?,?,?,?,?,?,?,"pending",?)
    ''', (
        doc_id, user_id, buyer_email,
        payment_method, payment_ref,
        doc['price'], doc['currency'],
        datetime.now().isoformat()
    ))
    purchase_id = cursor.lastrowid

    # Notification in-app à l'admin
    try:
        cursor.execute(
            "SELECT id FROM users WHERE role IN ('admin','superadmin') LIMIT 1"
        )
        admin_row = cursor.fetchone()
        if admin_row:
            cursor.execute('''
                INSERT INTO notifications (user_id, type, title, message, created_at)
                VALUES (?, "info", "💰 Nouvelle demande d''achat", ?, ?)
            ''', (
                admin_row[0],
                f"Demande pour \"{doc['title']}\" — {payment_method} #{payment_ref}",
                datetime.now().isoformat()
            ))
    except Exception:
        pass

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'purchase_id': purchase_id})


@app.route('/documents/purchase/confirm/<int:purchase_id>', methods=['POST'])
@admin_required
def documents_purchase_confirm(purchase_id):
    """Confirmer un achat (admin)."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    conn.execute(
        "UPDATE document_purchases SET status='confirmed', updated_at=? WHERE id=?",
        (datetime.now().isoformat(), purchase_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/documents/purchase/reject/<int:purchase_id>', methods=['POST'])
@admin_required
def documents_purchase_reject(purchase_id):
    """Rejeter un achat (admin)."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    conn.execute(
        "UPDATE document_purchases SET status='rejected', updated_at=? WHERE id=?",
        (datetime.now().isoformat(), purchase_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ─── NAS Admin ────────────────────────────────────────────────
@app.route('/nas/admin')
@admin_required
def nas_admin():
    """Accès au serveur NAS — redirige vers l'URL NAS ou l'admin."""
    nas_url = os.environ.get('NAS_URL', '')
    if nas_url:
        return redirect(nas_url)
    return redirect(url_for('admin_panel'))


# ═══════════════════════════════════════════════════════════════════
# ROUTES SURVEY — Sondage Liberté Financière & Trading
# ═══════════════════════════════════════════════════════════════════

@app.route('/survey')
def survey():
    """Page principale du sondage."""
    return render_template('survey.html')


@app.route('/survey/submit', methods=['POST'])
def survey_submit():
    """Enregistre les réponses du sondage (anonyme, anti-doublon 24h)."""
    data       = request.get_json(force=True, silent=True) or {}
    answers    = json.dumps(data)
    session_id = session.get('survey_sid')
    if not session_id:
        session_id            = secrets.token_hex(16)
        session['survey_sid'] = session_id

    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr or '')

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500

    cursor = conn.cursor()
    cursor.execute('''
        SELECT id FROM survey_responses
        WHERE session_id=? AND created_at >= datetime("now","-1 day")
    ''', (session_id,))
    if cursor.fetchone():
        conn.close()
        return jsonify({
            'success': True,
            'already': True,
            'message': 'Vous avez déjà répondu récemment.',
            'redirect': '/survey/stats'
        })

    cursor.execute(
        'INSERT INTO survey_responses (session_id, answers, ip_address) VALUES (?,?,?)',
        (session_id, answers, ip_address[:64] if ip_address else '')
    )
    conn.commit()
    conn.close()
    return jsonify({
        'success':  True,
        'message':  'Merci pour votre participation !',
        'redirect': '/survey/stats'
    })


@app.route('/survey/stats')
def survey_stats():
    """Page des résultats du sondage."""
    return render_template('survey_stats.html')


@app.route('/survey/api/stats')
def survey_api_stats():
    """Retourne les statistiques agrégées du sondage en JSON."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500

    cursor = conn.cursor()
    cursor.execute("SELECT answers, created_at FROM survey_responses ORDER BY created_at DESC")
    rows  = cursor.fetchall()
    conn.close()

    total = len(rows)
    if total == 0:
        return jsonify({'success': True, 'total': 0, 'stats': {}, 'timeline': []})

    aggregated = {}
    dates      = {}

    for row in rows:
        try:
            answers = json.loads(row['answers'])
        except Exception:
            continue

        day = (row['created_at'] or '')[:10]
        if day:
            dates[day] = dates.get(day, 0) + 1

        for q_id, value in answers.items():
            if q_id not in aggregated:
                aggregated[q_id] = {}
            val_str = str(value)
            aggregated[q_id][val_str] = aggregated[q_id].get(val_str, 0) + 1

    timeline  = [{'date': d, 'count': c} for d, c in sorted(dates.items())]
    stats_out = {}
    for q_id, counts in aggregated.items():
        total_q = sum(counts.values())
        stats_out[q_id] = {
            'counts':      counts,
            'total':       total_q,
            'percentages': {k: round(v / total_q * 100, 1) for k, v in counts.items()},
            'top_answer':  max(counts, key=counts.get) if counts else None,
        }

    return jsonify({
        'success':       True,
        'total':         total,
        'stats':         stats_out,
        'timeline':      timeline,
        'last_response': rows[0]['created_at'][:16] if rows else None,
    })


@app.route('/survey/admin/responses')
@admin_required
def survey_admin_responses():
    """Liste des réponses brutes (admin)."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, session_id, answers, created_at FROM survey_responses ORDER BY created_at DESC LIMIT 500"
    )
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'responses': rows, 'total': len(rows)})


@app.route('/survey/admin/export')
@admin_required
def survey_admin_export():
    """Export CSV des réponses au sondage (admin)."""
    conn = get_db_connection()
    if not conn:
        return "Erreur DB", 500
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM survey_responses ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()

    from io import StringIO
    si = StringIO()
    si.write('\ufeff')
    writer = csv.writer(si)
    writer.writerow(['ID', 'Session', 'Réponses', 'IP', 'Date'])
    for r in rows:
        writer.writerow([r['id'], r['session_id'], r['answers'], r['ip_address'], r['created_at']])

    from flask import Response
    return Response(
        si.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=survey-{datetime.now().strftime("%Y%m%d")}.csv'}
    )

# ═══════════════════════════════════════════════════════════════════
# FIN DU MODULE DOCUMENTS & SURVEY
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# MODULE WORD DOCUMENTS — Sauvegarde serveur + Dashboard
# ═══════════════════════════════════════════════════════════════════

def init_word_docs_db():
    """Crée la table word_documents pour stocker les docs de l'éditeur."""
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS word_documents (
        id          TEXT    NOT NULL,
        user_id     INTEGER NOT NULL,
        title       TEXT    DEFAULT "Sans titre",
        content     TEXT    DEFAULT "",
        word_count  INTEGER DEFAULT 0,
        created_at  TEXT    DEFAULT CURRENT_TIMESTAMP,
        updated_at  TEXT    DEFAULT CURRENT_TIMESTAMP,
        PRIMARY KEY (id, user_id),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    ''')
    conn.commit()
    conn.close()
    print("✅ Table word_documents initialisée")

init_word_docs_db()


@app.route('/api/bloc-notes/save', methods=['POST'])
@login_required
def api_word_doc_save():
    """Sauvegarde ou met à jour un document Word depuis l'éditeur."""
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}
    doc_id  = data.get('id', '').strip()
    title   = (data.get('title') or 'Sans titre').strip()[:255]
    content = data.get('content', '')
    updated = data.get('updatedAt') or datetime.now().isoformat()

    if not doc_id:
        return jsonify({'success': False, 'error': 'ID manquant'}), 400

    # Compter les mots (texte brut)
    import re as _re
    text_plain = _re.sub(r'<[^>]+>', ' ', content)
    word_count = len(text_plain.split()) if text_plain.strip() else 0

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB error'}), 500

    cursor = conn.cursor()
    cursor.execute("SELECT id FROM word_documents WHERE id=? AND user_id=?", (doc_id, user_id))
    exists = cursor.fetchone()

    if exists:
        cursor.execute('''
            UPDATE word_documents SET title=?, content=?, word_count=?, updated_at=?
            WHERE id=? AND user_id=?
        ''', (title, content, word_count, updated, doc_id, user_id))
    else:
        created = data.get('createdAt') or updated
        cursor.execute('''
            INSERT INTO word_documents (id, user_id, title, content, word_count, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?)
        ''', (doc_id, user_id, title, content, word_count, created, updated))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'word_count': word_count})


@app.route('/api/bloc-notes/delete', methods=['POST'])
@login_required
def api_word_doc_delete():
    """Supprime un document Word."""
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}
    doc_id  = data.get('id', '').strip()
    if not doc_id:
        return jsonify({'success': False, 'error': 'ID manquant'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    conn.execute("DELETE FROM word_documents WHERE id=? AND user_id=?", (doc_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/bloc-notes/rename', methods=['POST'])
@login_required
def api_word_doc_rename():
    """Renomme un document Word."""
    user_id = session['user_id']
    data    = request.get_json(force=True, silent=True) or {}
    doc_id  = data.get('id', '').strip()
    title   = (data.get('title') or 'Sans titre').strip()[:255]
    if not doc_id:
        return jsonify({'success': False}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    conn.execute(
        "UPDATE word_documents SET title=?, updated_at=? WHERE id=? AND user_id=?",
        (title, datetime.now().isoformat(), doc_id, user_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/bloc-notes/list', methods=['GET'])
@login_required
def api_word_doc_list():
    """Liste tous les documents Word de l'utilisateur (pour le dashboard)."""
    user_id = session['user_id']
    conn    = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'docs': []}), 500
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, title, word_count, created_at, updated_at,
               SUBSTR(content, 1, 300) as excerpt
        FROM word_documents
        WHERE user_id=?
        ORDER BY updated_at DESC
    ''', (user_id,))
    rows = cursor.fetchall()
    conn.close()

    import re as _re
    docs = []
    for r in rows:
        r = dict(r)
        # Nettoyer l'excerpt HTML
        plain = _re.sub(r'<[^>]+>', ' ', r.get('excerpt') or '')
        plain = ' '.join(plain.split())[:200]
        r['excerpt'] = plain
        docs.append(r)

    return jsonify({'success': True, 'docs': docs, 'total': len(docs)})


@app.route('/api/bloc-notes/get/<doc_id>', methods=['GET'])
@login_required
def api_word_doc_get(doc_id):
    """Récupère le contenu complet d'un document Word."""
    user_id = session['user_id']
    conn    = get_db_connection()
    if not conn:
        return jsonify({'success': False}), 500
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM word_documents WHERE id=? AND user_id=?",
        (doc_id, user_id)
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': 'Document non trouvé'}), 404
    return jsonify({'success': True, 'doc': dict(row)})


# ═══════════════════════════════════════════════════════════════════
# MODULE NEWS — API /api/news/feed et /api/news/analyze
# ═══════════════════════════════════════════════════════════════════

import urllib.request as _ur
import xml.etree.ElementTree as _ET

# Cache mémoire pour éviter trop d'appels externes
_news_cache = {'articles': [], 'btc': {}, 'fear_greed': {}, 'ts': 0}
_NEWS_TTL   = 5 * 60  # 5 minutes

_SOURCES = [
    {
        'url':       'https://feeds.feedburner.com/CoinDesk',
        'source':    'CoinDesk',
        'category':  'crypto',
        'cat_label': 'Crypto',
        'cat_color': '#f7931a',
    },
    {
        'url':       'https://cointelegraph.com/rss',
        'source':    'CoinTelegraph',
        'category':  'crypto',
        'cat_label': 'Crypto',
        'cat_color': '#f7931a',
    },
    {
        'url':       'https://cryptopanic.com/news/rss/',
        'source':    'CryptoPanic',
        'category':  'crypto',
        'cat_label': 'Crypto',
        'cat_color': '#f7931a',
    },
    {
        'url':       'https://feeds.a.dj.com/rss/RSSMarketsMain.xml',
        'source':    'Wall Street Journal',
        'category':  'macro',
        'cat_label': 'Macro',
        'cat_color': '#60a5fa',
    },
    {
        'url':       'https://www.forexfactory.com/ff_calendar_thisweek.xml',
        'source':    'ForexFactory',
        'category':  'forex',
        'cat_label': 'Forex',
        'cat_color': '#a78bfa',
    },
]

_BULL_WORDS = [
    'rise','pump','gain','bull','surge','high','break','rally','ath','record',
    'hausse','montée','record','rebond','signal','achat','buy','moon','profit',
    'growth','adoption','partnership','milestone','launch','upgrade'
]
_BEAR_WORDS = [
    'drop','fall','bear','crash','low','down','sell','dump','decline','fear',
    'baisse','chute','krach','panique','liquidation','hack','ban','warning',
    'loss','risk','concern','plunge','collapse','fraud','scam'
]


def _fetch_rss(src):
    """Récupère et parse un flux RSS, retourne une liste d'articles."""
    articles = []
    try:
        req = _ur.Request(src['url'], headers={
            'User-Agent': 'Mozilla/5.0 (compatible; KengniFinance/2.0)'
        })
        with _ur.urlopen(req, timeout=6) as r:
            raw = r.read()
        root = _ET.fromstring(raw)
        ns   = {'atom': 'http://www.w3.org/2005/Atom'}

        items = root.findall('.//item') or root.findall('.//atom:entry', ns)
        for i, item in enumerate(items[:8]):
            def g(tag):
                el = item.find(tag)
                return (el.text or '').strip() if el is not None else ''

            title = g('title') or g('atom:title')
            link  = g('link')  or g('atom:link')
            pub   = g('pubDate') or g('atom:published') or g('updated')
            desc  = g('description') or g('atom:summary') or g('content')

            if not title:
                continue

            tl = title.lower()
            if any(w in tl for w in _BULL_WORDS):
                sentiment = 'bullish'
                score     = '+' + str(55 + (hash(title) % 30))
                prob_bull = 60 + (hash(title) % 25)
            elif any(w in tl for w in _BEAR_WORDS):
                sentiment = 'bearish'
                score     = '-' + str(55 + (hash(title) % 30))
                prob_bull = 25 + (hash(title) % 20)
            else:
                sentiment = 'neutral'
                score     = '~' + str(45 + (hash(title) % 15))
                prob_bull = 45 + (hash(title) % 15)

            prob_bear = 100 - prob_bull

            # Extraire l'image de la description
            image = None
            try:
                import re as _re2
                m = _re2.search(r'<img[^>]+src=["\']([^"\']+)["\']', desc or '')
                if m:
                    image = m.group(1)
            except Exception:
                pass

            # Nettoyer la description
            import re as _re3
            summary = _re3.sub(r'<[^>]+>', ' ', desc or '')
            summary = ' '.join(summary.split())[:200]

            articles.append({
                'id':        f"{src['source'].lower().replace(' ','_')}_{i}_{abs(hash(title)) % 99999}",
                'title':     title,
                'source':    src['source'],
                'source_url': link,
                'category':  src['category'],
                'cat_label': src['cat_label'],
                'cat_color': src['cat_color'],
                'sentiment': sentiment,
                'score':     score,
                'prob_bull': prob_bull,
                'prob_bear': prob_bear,
                'summary':   summary,
                'image':     image,
                'published': pub[:16] if pub else '',
            })
    except Exception as e:
        print(f"[RSS] Erreur {src['source']}: {e}")
    return articles


def _fetch_btc():
    try:
        req = _ur.Request(
            'https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd'
            '&include_24hr_change=true&include_24hr_vol=true&include_high_low=true',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with _ur.urlopen(req, timeout=5) as r:
            d = json.loads(r.read())['bitcoin']
        return {
            'price':  d.get('usd', 0),
            'change': d.get('usd_24h_change', 0),
            'high':   d.get('usd_24h_high', 0),
            'low':    d.get('usd_24h_low', 0),
        }
    except Exception:
        return {}


def _fetch_fear_greed():
    try:
        req = _ur.Request(
            'https://api.alternative.me/fng/?limit=1',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with _ur.urlopen(req, timeout=5) as r:
            d = json.loads(r.read())['data'][0]
        return {
            'value': int(d.get('value', 50)),
            'label': d.get('value_classification', 'Neutral'),
        }
    except Exception:
        return {'value': 50, 'label': 'Neutral'}


def _get_news_data(force=False):
    """Retourne les données news avec cache 5 min."""
    import time
    global _news_cache
    now = time.time()
    if not force and _news_cache['articles'] and (now - _news_cache['ts']) < _NEWS_TTL:
        return _news_cache

    # Fetch toutes les sources en parallèle (threads)
    import threading
    results = []
    lock    = threading.Lock()

    def fetch_one(src):
        arts = _fetch_rss(src)
        with lock:
            results.extend(arts)

    threads = [threading.Thread(target=fetch_one, args=(s,)) for s in _SOURCES]
    for t in threads:
        t.daemon = True
        t.start()
    for t in threads:
        t.join(timeout=8)

    # Dédupliquer par titre
    seen  = set()
    dedup = []
    for a in results:
        key = a['title'].lower()[:60]
        if key not in seen:
            seen.add(key)
            dedup.append(a)

    # Trier : bullish > bearish > neutral, puis par hash titre
    order = {'bullish': 0, 'bearish': 1, 'neutral': 2}
    dedup.sort(key=lambda a: (order.get(a['sentiment'], 2), hash(a['title'])))

    btc        = _fetch_btc()
    fear_greed = _fetch_fear_greed()

    _news_cache = {
        'articles':   dedup[:40],
        'btc':        btc,
        'fear_greed': fear_greed,
        'ts':         now,
    }
    return _news_cache


@app.route('/api/news/feed')
@login_required
def api_news_feed():
    """Retourne les actualités financières et crypto en JSON."""
    category = request.args.get('category', 'all')
    force    = request.args.get('force', '0') == '1'

    data     = _get_news_data(force=force)
    articles = data['articles']

    if category and category != 'all':
        articles = [a for a in articles if a.get('category') == category]

    return jsonify({
        'success':    True,
        'articles':   articles,
        'btc':        data['btc'],
        'fear_greed': data['fear_greed'],
        'total':      len(articles),
    })


@app.route('/api/news/analyze/<article_id>')
@login_required
def api_news_analyze(article_id):
    """Analyse IA d'un article de news via Claude (si dispo) ou analyse heuristique."""
    # Trouver l'article dans le cache
    data    = _get_news_data(force=False)
    article = next((a for a in data['articles'] if str(a['id']) == str(article_id)), None)

    if not article:
        return jsonify({'success': False, 'error': 'Article non trouvé'}), 404

    title     = article.get('title', '')
    sentiment = article.get('sentiment', 'neutral')
    prob_bull = article.get('prob_bull', 50)
    prob_bear = article.get('prob_bear', 50)

    # Analyse heuristique enrichie (sans appel externe)
    if sentiment == 'bullish':
        analyse_fr       = f"Cette actualité est perçue comme **positive** pour le marché crypto. Le titre indique une dynamique haussière pouvant attirer les acheteurs."
        impact_court     = f"Pression acheteuse probable sur BTC/USD dans les 1 à 4 heures. Volume en hausse possible."
        impact_moyen     = f"Si la tendance se confirme, potentiel de continuation haussière sur 1 à 7 jours."
        niveaux          = f"Support clé : zone des derniers bas. Résistance : ATH récent ou zone de liquidité supérieure."
        conseil          = f"Surveiller la confirmation sur le graphique 1H avant d'entrer. Ratio Risk/Reward minimum 1:2."
    elif sentiment == 'bearish':
        analyse_fr       = f"Cette actualité génère une **pression vendeuse** sur le marché. Risque de liquidations en cascade si les supports cèdent."
        impact_court     = f"Baisse probable à court terme. Surveiller les volumes de vente et les liquidations sur les exchanges."
        impact_moyen     = f"Possible consolidation ou correction de 5 à 15% si le sentiment reste négatif."
        niveaux          = f"Support critique : plancher récent. En cas de rupture, prochain support majeur à surveiller."
        conseil          = f"Prudence recommandée. Réduire l'exposition ou placer des stops serrés. Ne pas acheter la baisse sans confirmation."
    else:
        analyse_fr       = f"Cette actualité a un impact **neutre** sur le marché à court terme. Le marché attend un catalyseur plus fort."
        impact_court     = f"Peu d'impact immédiat attendu. Range probable entre les niveaux de support et résistance actuels."
        impact_moyen     = f"La direction dépendra des prochains événements macro et on-chain."
        niveaux          = f"Range en cours : surveiller les bornes hautes et basses du canal actuel."
        conseil          = f"Attendre une confirmation de direction. Éviter de surtraiter en l'absence de signal clair."

    return jsonify({
        'success': True,
        'article_id': article_id,
        'analysis': {
            'analyse_fr':        analyse_fr,
            'impact_court_terme': impact_court,
            'impact_moyen_terme': impact_moyen,
            'niveaux_cles':      niveaux,
            'conseil':           conseil,
            'prob_bull':         prob_bull,
            'prob_bear':         prob_bear,
            'sentiment':         sentiment,
        }
    })

# ═══════════════════════════════════════════════════════════════════
# FIN MODULE WORD DOCUMENTS + NEWS API
# ═══════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════
# MODULE BOUTIQUE E-COMMERCE — Produits physiques (Alibaba/AliExpress)
# ═══════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════
# MODULE BOUTIQUE E-COMMERCE — Produits physiques (multi-images)
# ═══════════════════════════════════════════════════════════════════

def init_shop_db():
    conn = get_db_connection()
    if not conn: return
    try:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS shop_products (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            name           TEXT    NOT NULL,
            description    TEXT    DEFAULT '',
            features       TEXT    DEFAULT '',
            category       TEXT    NOT NULL DEFAULT 'electronique',
            price          REAL    NOT NULL DEFAULT 0,
            original_price REAL    DEFAULT 0,
            stock          INTEGER DEFAULT 10,
            image_url      TEXT    DEFAULT '',
            images         TEXT    DEFAULT '[]',
            badge          TEXT    DEFAULT '',
            delivery_info  TEXT    DEFAULT 'Livraison 3-7 jours',
            reviews_count  INTEGER DEFAULT 0,
            is_active      INTEGER DEFAULT 1,
            created_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
            updated_at     TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS shop_orders (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_name  TEXT    DEFAULT '',
            customer_phone TEXT    DEFAULT '',
            customer_city  TEXT    DEFAULT '',
            items_summary  TEXT    DEFAULT '',
            items_json     TEXT    DEFAULT '[]',
            total          REAL    NOT NULL DEFAULT 0,
            pay_method     TEXT    DEFAULT 'orange_money',
            note           TEXT    DEFAULT '',
            status         TEXT    DEFAULT 'pending',
            created_at     TEXT    DEFAULT CURRENT_TIMESTAMP,
            updated_at     TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
        # Migration : ajouter colonne images si absente (DB existante)
        try:
            conn.execute("ALTER TABLE shop_products ADD COLUMN images TEXT DEFAULT '[]'")
            conn.commit()
        except Exception:
            pass
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM shop_products")
        if cur.fetchone()[0] == 0:
            demos = [
                ("Écouteurs Bluetooth TWS Pro",
                 "Écouteurs sans fil avec réduction de bruit active. Autonomie 30h avec le boîtier. Compatible iOS & Android.",
                 "✔ Réduction bruit active (ANC)\n✔ Autonomie 8h + 22h boîtier\n✔ Résistant à l'eau IPX5\n✔ Bluetooth 5.3\n✔ Micro HD intégré",
                 "electronique", 8500, 15000, 25, "", "hot", "Livraison 3-5 jours"),
                ("Montre Connectée Sport",
                 "Smartwatch multifonction : fréquence cardiaque, suivi du sommeil, GPS, notifications. Étanche 5ATM.",
                 "✔ Écran AMOLED 1.8 pouces\n✔ GPS intégré\n✔ Fréquence cardiaque 24/7\n✔ +100 modes sport\n✔ Autonomie 7 jours",
                 "electronique", 12000, 22000, 15, "", "hot", "Livraison 3-5 jours"),
                ("Sac à Dos Imperméable 30L",
                 "Sac à dos résistant à l'eau, port USB intégré. Plusieurs compartiments. Idéal école, bureau, voyage.",
                 "✔ Matière Oxford imperméable\n✔ Port USB de charge externe\n✔ Compartiment laptop 15,6\"\n✔ Volume 30 litres",
                 "mode", 7500, 12000, 30, "", "new", "Livraison 4-7 jours"),
                ("Power Bank 20000mAh Ultra-Slim",
                 "Batterie externe grande capacité, ultra-plate. Charge rapide 22.5W. 2 ports USB-A + 1 USB-C.",
                 "✔ Capacité 20000mAh\n✔ Charge rapide 22.5W\n✔ 3 ports simultanés\n✔ Compatible tous appareils",
                 "electronique", 9000, 16000, 35, "", "promo", "Livraison 3-5 jours"),
                ("Chaussures de Course Running",
                 "Chaussures légères et respirantes. Semelle EVA amortissante, mesh technique. Unisexe tailles 36-46.",
                 "✔ Mesh respirant ultra-léger\n✔ Semelle EVA amortissante\n✔ Grip tout terrain\n✔ Poids 280g",
                 "sport", 14000, 25000, 22, "", "promo", "Livraison 5-7 jours"),
            ]
            cur.executemany("""
                INSERT INTO shop_products
                (name,description,features,category,price,original_price,
                 stock,image_url,badge,delivery_info)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, demos)
            conn.commit()
    except Exception as e:
        print(f"[Shop] init error: {e}")
    finally:
        conn.close()

try:
    init_shop_db()
except Exception as _e:
    print(f"[Shop] init skipped: {_e}")


# ── Helpers images Shop ──────────────────────────────────────────────────────

SHOP_IMG_FOLDER = os.path.join('static', 'uploads', 'shop')
os.makedirs(SHOP_IMG_FOLDER, exist_ok=True)

def _shop_parse_images(product: dict) -> list:
    """Retourne la liste consolidée des images d'un produit (galerie + image_url)."""
    imgs = []
    try:
        imgs = json.loads(product.get('images') or '[]')
        if not isinstance(imgs, list):
            imgs = []
    except Exception:
        imgs = []
    # Ajouter image_url principale si pas déjà dans la galerie
    main = (product.get('image_url') or '').strip()
    if main and main not in imgs:
        imgs = [main] + imgs
    return [i for i in imgs if i]


@app.route('/shop')
def shop():
    conn = get_db_connection()
    products, orders_count = [], 0
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT * FROM shop_products WHERE is_active=1
            ORDER BY CASE badge WHEN 'hot' THEN 0 WHEN 'new' THEN 1 WHEN 'promo' THEN 2 ELSE 3 END,
                     created_at DESC
        """)
        raw = cur.fetchall()
        for r in raw:
            p = dict(r)
            p['all_images'] = _shop_parse_images(p)
            p['main_image'] = p['all_images'][0] if p['all_images'] else ''
            products.append(p)
        cur.execute("SELECT COUNT(*) FROM shop_orders")
        orders_count = cur.fetchone()[0] or 0
    except Exception as e:
        print(f"[Shop] {e}")
    finally:
        if conn: conn.close()
    # Récupérer les bannières actives
    banners = []
    try:
        conn2 = get_db_connection()
        cur2 = conn2.cursor()
        cur2.execute("SELECT id,type,content,image_url,icon,color,bg_color,link_url FROM shop_banners WHERE is_active=1 ORDER BY display_order ASC, id ASC")
        banners = [dict(r) for r in cur2.fetchall()]
        conn2.close()
    except Exception:
        pass
    return render_template('shop.html', products=products, orders_count=orders_count, banners=banners)


@app.route('/shop/order', methods=['POST'])
def shop_create_order():
    import json as _json
    data = request.get_json(force=True, silent=True) or {}
    items          = data.get('items', [])
    total          = float(data.get('total', 0) or 0)
    customer_name  = (data.get('name') or '').strip()
    customer_phone = (data.get('phone') or '').strip()
    customer_city  = (data.get('city') or '').strip()
    pay_method     = data.get('pay_method', 'orange_money')
    note           = data.get('note', '')

    if not items or not customer_name:
        return jsonify({'success': False, 'error': 'Données manquantes'})

    items_summary = ', '.join(f"{i.get('name','?').strip()} ×{i.get('qty',1)}" for i in items)

    conn = get_db_connection()
    order_id = None
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO shop_orders
            (customer_name, customer_phone, customer_city,
             items_summary, items_json, total, pay_method, note, status)
            VALUES (?,?,?,?,?,?,?,?,'pending')
        """, (customer_name, customer_phone, customer_city,
              items_summary, _json.dumps(items, ensure_ascii=False),
              total, pay_method, note))
        conn.commit()
        order_id = cur.lastrowid
        for item in items:
            pid = item.get('id')
            qty = int(item.get('qty', 1))
            if pid:
                cur.execute("""
                    UPDATE shop_products
                    SET stock = MAX(0, stock - ?), updated_at = CURRENT_TIMESTAMP
                    WHERE id = ? AND stock > 0
                """, (qty, pid))
        conn.commit()
        try:
            cur.execute("""
                INSERT INTO notifications (user_id,type,title,message,is_read,created_at)
                SELECT id,'shop','🛒 Nouvelle commande',?,0,CURRENT_TIMESTAMP
                FROM users WHERE role IN ('admin','superadmin') LIMIT 3
            """, (f"#{order_id} — {customer_name} — {items_summary} — {total:,.0f} XAF",))
            conn.commit()
        except Exception:
            pass
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()
    return jsonify({'success': True, 'order_id': order_id})


@app.route('/shop/admin')
@login_required
def shop_admin():
    role = session.get('role')
    uid  = session.get('user_id')
    # Admins toujours autorisés ; autres utilisateurs si au moins 1 permission boutique
    if role not in ('admin', 'superadmin'):
        perms = get_shop_perms(uid)
        if not any(perms.values()):
            return redirect(url_for('shop'))
    conn = get_db_connection()
    products, orders = [], []
    total_revenue = pending_count = active_count = 0
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM shop_products ORDER BY created_at DESC")
        for r in cur.fetchall():
            p = dict(r)
            p['all_images'] = _shop_parse_images(p)
            p['main_image'] = p['all_images'][0] if p['all_images'] else ''
            products.append(p)
        cur.execute("SELECT * FROM shop_orders ORDER BY created_at DESC LIMIT 200")
        orders = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT COALESCE(SUM(total),0) FROM shop_orders WHERE status='paid'")
        total_revenue = cur.fetchone()[0] or 0
        pending_count = sum(1 for o in orders if o['status'] == 'pending')
        active_count  = sum(1 for p in products if p['is_active'])
    except Exception as e:
        print(f"[ShopAdmin] {e}")
    finally:
        if conn: conn.close()
    # Récupérer tous les utilisateurs pour la section Accès (tous les rôles sauf admin/superadmin)
    all_users = []
    try:
        conn2 = get_db_connection()
        if conn2:
            cur2 = conn2.cursor()
            cur2.execute(
                "SELECT id, username, email, role, shop_permissions, shop_access, last_login, status FROM users ORDER BY username ASC"
            )
            all_users = [dict(r) for r in cur2.fetchall()]
            # Calculer inactivité : > 8 jours = inactif automatiquement
            from datetime import datetime as _dt
            for u in all_users:
                ll = u.get('last_login')
                if ll:
                    try:
                        delta = (_dt.now() - _dt.fromisoformat(ll[:19])).days
                        u['days_inactive'] = delta
                        u['auto_inactive'] = delta > 8
                    except Exception:
                        u['days_inactive'] = None
                        u['auto_inactive'] = False
                else:
                    u['days_inactive'] = None
                    u['auto_inactive'] = True  # Jamais connecté
            conn2.close()
    except Exception as e:
        print(f"[ShopAdmin users] {e}")

    # Données activité (courbes évolution 30 jours)
    activity_data = {}
    try:
        conn3 = get_db_connection()
        if conn3:
            cur3 = conn3.cursor()
            cur3.execute("""
                SELECT handled_by, DATE(COALESCE(handled_at,updated_at)) as day, COUNT(*) as cnt
                FROM shop_orders
                WHERE handled_by IS NOT NULL AND DATE(COALESCE(handled_at,updated_at)) >= DATE('now','-30 days')
                GROUP BY handled_by, day ORDER BY day ASC
            """)
            for row in cur3.fetchall():
                uid2 = row["handled_by"]
                if uid2 not in activity_data:
                    activity_data[uid2] = {}
                activity_data[uid2][row["day"]] = row["cnt"]
            conn3.close()
    except Exception as e:
        print(f"[ShopAdmin activity] {e}")

    current_user_perms = get_shop_perms(session.get("user_id"))
    return render_template("shop_admin.html",
        products=products, orders=orders,
        total_revenue=total_revenue,
        pending_count=pending_count, active_count=active_count,
        all_users=all_users,
        activity_data=activity_data,
        user_perms=current_user_perms)


@app.route('/shop/api/product', methods=['POST'])
@login_required
def shop_create_product():
    if not can_shop('add'):
        return jsonify({'success': False, 'error': 'Non autorisé — permission Ajouter requise'}), 403
    d = request.get_json(force=True, silent=True) or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Nom requis'})
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        images_json = json.dumps(d.get('images', []) if isinstance(d.get('images'), list) else [])
        cur.execute("""
            INSERT INTO shop_products
            (name,description,features,category,price,original_price,
             stock,image_url,images,badge,delivery_info,is_active)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (name, d.get('description', ''), d.get('features', ''),
              d.get('category', 'electronique'),
              float(d.get('price', 0) or 0), float(d.get('original_price', 0) or 0),
              int(d.get('stock', 10)), d.get('image_url', ''),
              images_json,
              d.get('badge', ''), d.get('delivery_info', 'Livraison 3-7 jours'),
              1 if d.get('is_active', True) else 0))
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/product/<int:pid>', methods=['PUT'])
@login_required
def shop_update_product(pid):
    if not can_shop('edit'):
        return jsonify({'success': False, 'error': 'Non autorisé — permission Modifier requise'}), 403
    d = request.get_json(force=True, silent=True) or {}
    images_json = json.dumps(d.get('images', []) if isinstance(d.get('images'), list) else [])
    conn = get_db_connection()
    try:
        conn.execute("""
            UPDATE shop_products SET
                name=?,description=?,features=?,category=?,price=?,
                original_price=?,stock=?,image_url=?,images=?,badge=?,
                delivery_info=?,is_active=?,updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, ((d.get('name') or '').strip(), d.get('description', ''),
              d.get('features', ''), d.get('category', 'electronique'),
              float(d.get('price', 0) or 0), float(d.get('original_price', 0) or 0),
              int(d.get('stock', 10)), d.get('image_url', ''),
              images_json,
              d.get('badge', ''), d.get('delivery_info', 'Livraison 3-7 jours'),
              1 if d.get('is_active', True) else 0, pid))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/product/<int:pid>/toggle', methods=['POST'])
@login_required
def shop_toggle_product(pid):
    if not can_shop('edit'):
        return jsonify({'success': False, 'error': 'Non autorisé — permission Modifier requise'}), 403
    conn = get_db_connection()
    try:
        conn.execute("UPDATE shop_products SET is_active=1-is_active, updated_at=CURRENT_TIMESTAMP WHERE id=?", (pid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/product/<int:pid>', methods=['DELETE'])
@login_required
def shop_delete_product(pid):
    if not can_shop('delete'):
        return jsonify({'success': False, 'error': 'Non autorisé — permission Supprimer requise'}), 403
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT image_url, images FROM shop_products WHERE id=?", (pid,))
        row = cur.fetchone()
        if row:
            all_imgs = _shop_parse_images(dict(row))
            for img_path in all_imgs:
                if '/uploads/shop/' in img_path:
                    local = os.path.join('static', img_path.lstrip('/static/'))
                    try:
                        if os.path.exists(local): os.remove(local)
                    except Exception:
                        pass
        conn.execute("DELETE FROM shop_products WHERE id=?", (pid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/product/upload-image', methods=['POST'])
@login_required
def shop_upload_product_image():
    """Upload une ou plusieurs images pour un produit.
    Form-data : image (fichier), pid (optionnel, int).
    Retourne l'URL et met à jour la galerie si pid fourni."""
    if not can_shop('add') and not can_shop('edit'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    if 'image' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier reçu'})
    f = request.files['image']
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'Fichier vide'})
    if not allowed_file(f.filename):
        return jsonify({'success': False, 'error': 'Extension invalide (png, jpg, jpeg, gif, webp)'})
    try:
        ext   = f.filename.rsplit('.', 1)[1].lower()
        fname = secure_filename(f'shop_{datetime.now().strftime("%Y%m%d_%H%M%S%f")}.{ext}')
        dest  = os.path.join(app.config['UPLOAD_FOLDER'], 'shop')
        os.makedirs(dest, exist_ok=True)
        f.save(os.path.join(dest, fname))
        url = f'/static/uploads/shop/{fname}'

        # Si pid fourni → ajouter à la galerie du produit
        pid = request.form.get('pid', type=int)
        if pid:
            conn = get_db_connection()
            try:
                cur = conn.cursor()
                cur.execute("SELECT image_url, images FROM shop_products WHERE id=?", (pid,))
                row = cur.fetchone()
                if row:
                    row = dict(row)
                    try:
                        imgs = json.loads(row['images'] or '[]')
                        if not isinstance(imgs, list): imgs = []
                    except Exception:
                        imgs = []
                    # Définir comme image principale si aucune encore
                    new_main = row['image_url'] or url
                    imgs.append(url)
                    conn.execute(
                        "UPDATE shop_products SET image_url=?, images=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (new_main, json.dumps(imgs), pid)
                    )
                    conn.commit()
            except Exception as db_e:
                print(f"[ShopUpload] DB error: {db_e}")
            finally:
                conn.close()

        return jsonify({'success': True, 'url': url})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/shop/api/product/<int:pid>/remove-image', methods=['POST'])
@login_required
def shop_remove_product_image(pid):
    """Supprime une image de la galerie d'un produit (+ fichier local si uploadé)."""
    if not can_shop('edit'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    data = request.get_json(force=True, silent=True) or {}
    url_to_remove = (data.get('url') or '').strip()
    if not url_to_remove:
        return jsonify({'success': False, 'error': 'URL manquante'})
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT image_url, images FROM shop_products WHERE id=?", (pid,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Produit introuvable'}), 404
        row = dict(row)
        try:
            imgs = json.loads(row['images'] or '[]')
            if not isinstance(imgs, list): imgs = []
        except Exception:
            imgs = []
        imgs = [i for i in imgs if i != url_to_remove]
        new_main = row['image_url']
        if new_main == url_to_remove:
            new_main = imgs[0] if imgs else ''
        conn.execute(
            "UPDATE shop_products SET image_url=?, images=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (new_main, json.dumps(imgs), pid)
        )
        conn.commit()
        # Supprimer le fichier local si c'est un upload local
        if '/uploads/shop/' in url_to_remove:
            local = os.path.join('static', url_to_remove.lstrip('/'))
            try:
                if os.path.exists(local): os.remove(local)
            except Exception:
                pass
        return jsonify({'success': True, 'images': imgs, 'main_image': new_main})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/product/<int:pid>/set-main-image', methods=['POST'])
@login_required
def shop_set_main_image(pid):
    """Définit l'image principale d'un produit depuis sa galerie."""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    data = request.get_json(force=True, silent=True) or {}
    url  = (data.get('url') or '').strip()
    if not url:
        return jsonify({'success': False, 'error': 'URL manquante'})
    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE shop_products SET image_url=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (url, pid)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/order/<int:oid>/status', methods=['POST'])
@login_required
def shop_update_order_status(oid):
    """Met à jour le statut et enregistre qui a géré la commande."""
    perms = get_shop_perms(session.get('user_id'))
    if not (perms.get('access') or perms.get('edit')):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    status = (request.get_json(force=True, silent=True) or {}).get('status', 'pending')
    if status not in ('pending', 'paid', 'shipped', 'delivered', 'cancelled'):
        return jsonify({'success': False, 'error': 'Statut invalide'})
    conn = get_db_connection()
    try:
        conn.execute(
            """UPDATE shop_orders SET status=?, handled_by=?, handled_at=CURRENT_TIMESTAMP,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (status, session.get('user_id'), oid)
        )
        conn.commit()
        try:
            conn.execute(
                "INSERT INTO shop_activity_log (user_id, action, entity_type, entity_id, details) VALUES (?,?,?,?,?)",
                (session.get('user_id'), 'order_status', 'order', oid, f'status={status}')
            )
            conn.commit()
        except Exception:
            pass
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/order/<int:oid>', methods=['DELETE'])
@login_required
def shop_delete_order(oid):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM shop_orders WHERE id=?", (oid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()

# ── FIN MODULE BOUTIQUE ─────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════
# MODULE BOUTIQUE — EXTENSIONS SPRINT 1
# Import/Export CSV · Notifications SSE · Kanban
# ═══════════════════════════════════════════════════════════════════

import io as _io
import csv as _csv

# ── SSE : clients connectés ─────────────────────────────────────────
_sse_clients = []   # list of queue.Queue()
import queue as _queue

def _sse_broadcast(data: dict):
    """Diffuse un événement JSON à tous les clients SSE connectés."""
    import json as _j
    msg = f"data: {_j.dumps(data, ensure_ascii=False)}\n\n"
    dead = []
    for q in _sse_clients:
        try:
            q.put_nowait(msg)
        except Exception:
            dead.append(q)
    for q in dead:
        try: _sse_clients.remove(q)
        except ValueError: pass


@app.route('/shop/stream')
def shop_sse_stream():
    """SSE endpoint — nouvelles commandes en temps réel."""
    from flask import Response, stream_with_context
    if not session.get('user_id'):
        return jsonify({'error': 'Non autorisé'}), 403

    q = _queue.Queue(maxsize=20)
    _sse_clients.append(q)

    def generate():
        yield "data: {\"type\":\"connected\"}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield msg
                except _queue.Empty:
                    yield ": keepalive\n\n"
        except GeneratorExit:
            pass
        finally:
            try: _sse_clients.remove(q)
            except ValueError: pass

    resp = Response(stream_with_context(generate()),
                    mimetype='text/event-stream',
                    headers={
                        'Cache-Control': 'no-cache',
                        'X-Accel-Buffering': 'no',
                        'Connection': 'keep-alive',
                    })
    return resp


# Patch shop_create_order pour broadcaster via SSE ──────────────────
_orig_shop_create_order = None

def _patch_shop_create_order():
    global _orig_shop_create_order
    orig = app.view_functions.get('shop_create_order')
    if orig is None or _orig_shop_create_order is not None:
        return

    @wraps(orig)
    def patched(*a, **kw):
        resp = orig(*a, **kw)
        try:
            import json as _j
            data = request.get_json(force=True, silent=True) or {}
            _sse_broadcast({
                'type': 'new_order',
                'customer': data.get('name', ''),
                'total': data.get('total', 0),
                'items': data.get('items', []),
                'ts': datetime.utcnow().strftime('%H:%M'),
            })
        except Exception:
            pass
        return resp

    _orig_shop_create_order = orig
    app.view_functions['shop_create_order'] = patched

with app.app_context():
    try:
        _patch_shop_create_order()
    except Exception:
        pass


# ── Export CSV ──────────────────────────────────────────────────────
@app.route('/shop/api/products/export')
@login_required
def shop_export_products():
    """Exporte le catalogue produits en CSV ou Excel."""
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403

    fmt = request.args.get('format', 'csv').lower()
    cat = request.args.get('category', '')

    conn = get_db_connection()
    try:
        q = "SELECT * FROM shop_products"
        params = []
        if cat:
            q += " WHERE category=?"
            params.append(cat)
        q += " ORDER BY category, name"
        rows = conn.execute(q, params).fetchall()

        COLS = ['id','name','category','brand','price','original_price',
                'stock','sku','description','features','badge','is_active',
                'image_url','delivery_info','created_at']

        if fmt == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Catalogue'
                hdr_font  = Font(bold=True, color='FFFFFF')
                hdr_fill  = PatternFill(fill_type='solid', fgColor='059669')
                ws.append(COLS)
                for cell in ws[1]:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                for row in rows:
                    ws.append([row[c] if c in row.keys() else '' for c in COLS])
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = max(
                        len(str(col[0].value or '')), 12)
                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return send_file(buf,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'kni_catalogue_{datetime.now().strftime("%Y%m%d")}.xlsx')
            except ImportError:
                pass  # fall back to CSV

        # CSV
        buf = _io.StringIO()
        writer = _csv.DictWriter(buf, fieldnames=COLS, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow({c: (row[c] if c in row.keys() else '') for c in COLS})

        output = _io.BytesIO(buf.getvalue().encode('utf-8-sig'))
        return send_file(output, mimetype='text/csv', as_attachment=True,
                         download_name=f'kni_catalogue_{datetime.now().strftime("%Y%m%d")}.csv')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Import CSV ──────────────────────────────────────────────────────
@app.route('/shop/api/products/import', methods=['POST'])
@login_required
def shop_import_products():
    """Importe des produits depuis un fichier CSV (mode upsert par SKU)."""
    if not can_shop('add'):
        return jsonify({'success': False, 'error': 'Permission refusée'}), 403

    f = request.files.get('file')
    if not f:
        return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('csv', 'xlsx', 'xls'):
        return jsonify({'success': False, 'error': 'Format non supporté (CSV ou Excel)'}), 400

    results = {'created': 0, 'updated': 0, 'errors': [], 'total': 0}
    REQUIRED = {'name'}
    FLOAT_FIELDS  = {'price', 'original_price'}
    INT_FIELDS    = {'stock', 'is_active'}

    def parse_rows(content_bytes):
        """Parse CSV ou Excel, retourne liste de dicts."""
        if ext == 'csv':
            text = content_bytes.decode('utf-8-sig', errors='replace')
            reader = _csv.DictReader(_io.StringIO(text))
            return list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows: return []
            headers = [str(h).strip() if h else '' for h in rows[0]]
            return [dict(zip(headers, r)) for r in rows[1:]]

    try:
        raw_rows = parse_rows(f.read())
    except Exception as e:
        return jsonify({'success': False, 'error': f'Lecture fichier échouée : {e}'}), 400

    conn = get_db_connection()
    try:
        for i, row in enumerate(raw_rows, 1):
            row = {k.strip().lower(): str(v).strip() if v is not None else '' for k, v in row.items()}
            results['total'] += 1

            # Validation
            if not row.get('name'):
                results['errors'].append({'line': i, 'error': 'Champ name manquant', 'data': row})
                continue

            # Nettoyage des types
            for ff in FLOAT_FIELDS:
                if row.get(ff):
                    try: row[ff] = float(str(row[ff]).replace(' ', '').replace(',', '.'))
                    except: row[ff] = 0.0
                else:
                    row[ff] = 0.0
            for fi in INT_FIELDS:
                if row.get(fi):
                    try: row[fi] = int(float(row[fi]))
                    except: row[fi] = 0
                else:
                    row[fi] = 0 if fi == 'is_active' else 0

            sku = row.get('sku', '').strip() or None

            try:
                existing = None
                if sku:
                    existing = conn.execute(
                        "SELECT id FROM shop_products WHERE sku=?", (sku,)
                    ).fetchone()

                if existing:
                    conn.execute("""
                        UPDATE shop_products SET
                          name=?, category=?, brand=?, price=?, original_price=?,
                          stock=?, description=?, features=?, badge=?, is_active=?,
                          delivery_info=?, updated_at=CURRENT_TIMESTAMP
                        WHERE id=?""",
                        (row.get('name'), row.get('category',''), row.get('brand',''),
                         row['price'], row['original_price'], row['stock'],
                         row.get('description',''), row.get('features',''),
                         row.get('badge',''), row.get('is_active', 1),
                         row.get('delivery_info',''), existing['id']))
                    results['updated'] += 1
                else:
                    conn.execute("""
                        INSERT INTO shop_products
                          (name, category, brand, price, original_price, stock, sku,
                           description, features, badge, is_active, delivery_info,
                           created_at, updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)""",
                        (row.get('name'), row.get('category',''), row.get('brand',''),
                         row['price'], row['original_price'], row['stock'],
                         sku, row.get('description',''), row.get('features',''),
                         row.get('badge',''), row.get('is_active', 1),
                         row.get('delivery_info','')))
                    results['created'] += 1

            except Exception as e:
                results['errors'].append({'line': i, 'error': str(e), 'name': row.get('name','')})

        conn.commit()
        return jsonify({'success': True, **results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Kanban : ordre drag & drop ──────────────────────────────────────
@app.route('/shop/api/orders/kanban', methods=['GET'])
@login_required
def shop_kanban_orders():
    """Retourne les commandes groupées par statut pour la vue Kanban."""
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403

    STATUSES = ['pending', 'paid', 'shipped', 'delivered', 'cancelled']
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT o.*, u.username as handler_name
            FROM shop_orders o
            LEFT JOIN users u ON u.id = o.handled_by
            ORDER BY o.created_at DESC
            LIMIT 300
        """).fetchall()
        grouped = {s: [] for s in STATUSES}
        for r in rows:
            s = r['status'] if r['status'] in STATUSES else 'pending'
            try:
                items = json.loads(r['items'] or '[]')
            except Exception:
                items = []
            grouped[s].append({
                'id': r['id'], 'customer_name': r['customer_name'],
                'customer_phone': r['customer_phone'],
                'total': r['total'], 'payment_method': r['payment_method'],
                'items': items, 'created_at': r['created_at'],
                'handler': r['handler_name'] or '',
            })
        return jsonify({'success': True, 'kanban': grouped})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn: conn.close()




# ═══════════════════════════════════════════════════════════════════
# MODULE STAFF ASSISTANCE — Techniciens & Commerciaux
# ═══════════════════════════════════════════════════════════════════

def init_staff_db():
    conn = get_db_connection()
    if not conn: return
    try:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS shop_staff (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    NOT NULL,
            role_type     TEXT    NOT NULL DEFAULT 'commercial',
            bio           TEXT    DEFAULT '',
            phone         TEXT    DEFAULT '',
            whatsapp      TEXT    DEFAULT '',
            email         TEXT    DEFAULT '',
            avatar_url    TEXT    DEFAULT '',
            specialties   TEXT    DEFAULT '[]',
            rating        REAL    DEFAULT 5.0,
            reviews       INTEGER DEFAULT 0,
            response_time TEXT    DEFAULT '< 5 min',
            is_active     INTEGER DEFAULT 1,
            display_order INTEGER DEFAULT 0,
            created_at    TEXT    DEFAULT CURRENT_TIMESTAMP,
            updated_at    TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
    except Exception as e:
        print(f"[Staff] init error: {e}")
    finally:
        conn.close()

try:
    init_staff_db()
except Exception as _e:
    print(f"[Staff] init skipped: {_e}")


@app.route('/shop/api/staff', methods=['GET'])
def shop_get_staff():
    """Liste publique des agents actifs (pour le widget boutique)."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id,name,role_type,bio,phone,whatsapp,avatar_url,
                   specialties,rating,reviews,response_time
            FROM shop_staff WHERE is_active=1
            ORDER BY display_order ASC, name ASC
        """)
        staff = [dict(r) for r in cur.fetchall()]
        for s in staff:
            try: s['specialties'] = json.loads(s['specialties'] or '[]')
            except: s['specialties'] = []
        return jsonify({'success': True, 'staff': staff})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'staff': []})
    finally:
        if conn: conn.close()


@app.route('/shop/api/staff/all', methods=['GET'])
@login_required
def shop_get_staff_all():
    """Liste complète des agents (admin)."""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM shop_staff ORDER BY display_order ASC, name ASC")
        staff = [dict(r) for r in cur.fetchall()]
        for s in staff:
            try: s['specialties'] = json.loads(s['specialties'] or '[]')
            except: s['specialties'] = []
        return jsonify({'success': True, 'staff': staff})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/staff', methods=['POST'])
@login_required
def shop_create_staff():
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Nom requis'})
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        specs = json.dumps(d.get('specialties', []) if isinstance(d.get('specialties'), list) else [])
        cur.execute("""
            INSERT INTO shop_staff
            (name,role_type,bio,phone,whatsapp,email,avatar_url,
             specialties,rating,reviews,response_time,is_active,display_order)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (name, d.get('role_type','commercial'), d.get('bio',''),
              d.get('phone',''), d.get('whatsapp',''), d.get('email',''),
              d.get('avatar_url',''), specs,
              float(d.get('rating', 5.0) or 5.0),
              int(d.get('reviews', 0) or 0),
              d.get('response_time','< 5 min'),
              1 if d.get('is_active', True) else 0,
              int(d.get('display_order', 0) or 0)))
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/staff/<int:sid>', methods=['PUT'])
@login_required
def shop_update_staff(sid):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    specs = json.dumps(d.get('specialties', []) if isinstance(d.get('specialties'), list) else [])
    conn = get_db_connection()
    try:
        conn.execute("""
            UPDATE shop_staff SET
                name=?,role_type=?,bio=?,phone=?,whatsapp=?,email=?,
                avatar_url=?,specialties=?,rating=?,reviews=?,
                response_time=?,is_active=?,display_order=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, ((d.get('name') or '').strip(), d.get('role_type','commercial'),
              d.get('bio',''), d.get('phone',''), d.get('whatsapp',''),
              d.get('email',''), d.get('avatar_url',''), specs,
              float(d.get('rating', 5.0) or 5.0),
              int(d.get('reviews', 0) or 0),
              d.get('response_time','< 5 min'),
              1 if d.get('is_active', True) else 0,
              int(d.get('display_order', 0) or 0), sid))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/staff/<int:sid>', methods=['DELETE'])
@login_required
def shop_delete_staff(sid):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        # Supprimer l'avatar local si uploadé
        cur = conn.cursor()
        cur.execute("SELECT avatar_url FROM shop_staff WHERE id=?", (sid,))
        row = cur.fetchone()
        if row and row['avatar_url'] and '/uploads/staff/' in row['avatar_url']:
            local = os.path.join('static', row['avatar_url'].lstrip('/'))
            try:
                if os.path.exists(local): os.remove(local)
            except Exception: pass
        conn.execute("DELETE FROM shop_staff WHERE id=?", (sid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/staff/upload-avatar', methods=['POST'])
@login_required
def shop_upload_staff_avatar():
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    if 'avatar' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier'})
    f = request.files['avatar']
    if not f or not f.filename or not allowed_file(f.filename):
        return jsonify({'success': False, 'error': 'Fichier invalide'})
    try:
        ext   = f.filename.rsplit('.', 1)[1].lower()
        fname = secure_filename(f'staff_{datetime.now().strftime("%Y%m%d_%H%M%S%f")}.{ext}')
        dest  = os.path.join(app.config['UPLOAD_FOLDER'], 'staff')
        os.makedirs(dest, exist_ok=True)
        f.save(os.path.join(dest, fname))
        url = f'/static/uploads/staff/{fname}'
        # Si sid fourni, mettre à jour directement
        sid = request.form.get('sid', type=int)
        if sid:
            conn = get_db_connection()
            try:
                conn.execute("UPDATE shop_staff SET avatar_url=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (url, sid))
                conn.commit()
            finally:
                conn.close()
        return jsonify({'success': True, 'url': url})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/shop/api/access/set', methods=['POST'])
@login_required
def shop_set_access():
    """Définit les permissions boutique granulaires d'un utilisateur.
    Body JSON : { user_id, add, edit, delete }  (booléens)
    Réservé aux admins/superadmins.
    """
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403

    data    = request.get_json(force=True, silent=True) or {}
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'error': 'user_id manquant'})

    perms = {
        'add':    bool(data.get('add',    False)),
        'edit':   bool(data.get('edit',   False)),
        'delete': bool(data.get('delete', False)),
    }

    conn = get_db_connection()
    try:
        # Vérification : on ne modifie pas un admin/superadmin
        row = conn.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Utilisateur introuvable'})
        if row['role'] in ('admin', 'superadmin'):
            return jsonify({'success': False, 'error': 'Impossible de modifier un administrateur'})

        conn.execute(
            "UPDATE users SET shop_permissions=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (json.dumps(perms), user_id)
        )
        conn.commit()
        return jsonify({'success': True, 'permissions': perms})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()

# ── FIN MODULE STAFF ASSISTANCE ──────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════
# MODULE BANNIÈRES DÉFILANTES — Top Bar Shop (texte + image)
# ═══════════════════════════════════════════════════════════════════

def init_banner_db():
    conn = get_db_connection()
    if not conn: return
    try:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS shop_banners (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            type         TEXT    NOT NULL DEFAULT 'text',
            content      TEXT    NOT NULL DEFAULT '',
            image_url    TEXT    DEFAULT '',
            icon         TEXT    DEFAULT '',
            color        TEXT    DEFAULT '#ffffff',
            bg_color     TEXT    DEFAULT '',
            link_url     TEXT    DEFAULT '',
            display_order INTEGER DEFAULT 0,
            is_active    INTEGER DEFAULT 1,
            created_at   TEXT    DEFAULT CURRENT_TIMESTAMP,
            updated_at   TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
        # Données de démo si vide
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM shop_banners")
        if cur.fetchone()[0] == 0:
            demos = [
                ('text','🚚 Livraison gratuite pour toute commande > 15 000 XAF','','fas fa-truck','#ffffff','','',0,1),
                ('text','🟠 Orange Money · 🟡 MTN MoMo · Paiement 100% sécurisé','','fas fa-shield-alt','#ffffff','','',1,1),
                ('text',"🔥 Jusqu'à -50% sur les produits électroniques cette semaine !",'','fas fa-fire','#fde68a','','',2,1),
                ('text','📦 Livraison partout au Cameroun en 3-7 jours ouvrables','','fas fa-map-marker-alt','#ffffff','','',3,1),
                ('text','💬 Support WhatsApp disponible 8h-20h du Lundi au Samedi','','fab fa-whatsapp','#bbf7d0','','',4,1),
            ]
            cur.executemany("""
                INSERT INTO shop_banners
                (type,content,image_url,icon,color,bg_color,link_url,display_order,is_active)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, demos)
            conn.commit()
    except Exception as e:
        print(f"[Banner] init error: {e}")
    finally:
        conn.close()

try:
    init_banner_db()
except Exception as _e:
    print(f"[Banner] init skipped: {_e}")


@app.route('/shop/api/banners', methods=['GET'])
def shop_get_banners():
    """Retourne les bannières actives pour la boutique publique."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id,type,content,image_url,icon,color,bg_color,link_url
            FROM shop_banners WHERE is_active=1
            ORDER BY display_order ASC, id ASC
        """)
        banners = [dict(r) for r in cur.fetchall()]
        return jsonify({'success': True, 'banners': banners})
    except Exception as e:
        return jsonify({'success': False, 'banners': [], 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/all', methods=['GET'])
@login_required
def shop_get_banners_all():
    """Retourne toutes les bannières (admin)."""
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM shop_banners ORDER BY display_order ASC, id ASC")
        banners = [dict(r) for r in cur.fetchall()]
        return jsonify({'success': True, 'banners': banners})
    except Exception as e:
        return jsonify({'success': False, 'banners': [], 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners', methods=['POST'])
@login_required
def shop_create_banner():
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    content = (d.get('content') or '').strip()
    if not content and not d.get('image_url'):
        return jsonify({'success':False,'error':'Contenu ou image requis'})
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO shop_banners
            (type,content,image_url,icon,color,bg_color,link_url,display_order,is_active)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (d.get('type','text'), content, d.get('image_url',''),
              d.get('icon',''), d.get('color','#ffffff'), d.get('bg_color',''),
              d.get('link_url',''), int(d.get('display_order',0)),
              1 if d.get('is_active',True) else 0))
        conn.commit()
        return jsonify({'success':True, 'id':cur.lastrowid})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/<int:bid>', methods=['PUT'])
@login_required
def shop_update_banner(bid):
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db_connection()
    try:
        conn.execute("""
            UPDATE shop_banners SET
                type=?,content=?,image_url=?,icon=?,color=?,
                bg_color=?,link_url=?,display_order=?,is_active=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (d.get('type','text'), (d.get('content') or '').strip(),
              d.get('image_url',''), d.get('icon',''), d.get('color','#ffffff'),
              d.get('bg_color',''), d.get('link_url',''),
              int(d.get('display_order',0)),
              1 if d.get('is_active',True) else 0, bid))
        conn.commit()
        return jsonify({'success':True})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/<int:bid>/toggle', methods=['POST'])
@login_required
def shop_toggle_banner(bid):
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        conn.execute("UPDATE shop_banners SET is_active=1-is_active,updated_at=CURRENT_TIMESTAMP WHERE id=?", (bid,))
        conn.commit()
        return jsonify({'success':True})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/<int:bid>/order', methods=['POST'])
@login_required
def shop_reorder_banner(bid):
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db_connection()
    try:
        conn.execute("UPDATE shop_banners SET display_order=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                     (int(d.get('order',0)), bid))
        conn.commit()
        return jsonify({'success':True})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/<int:bid>', methods=['DELETE'])
@login_required
def shop_delete_banner(bid):
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        # Supprimer image locale si uploadée
        cur = conn.cursor()
        cur.execute("SELECT image_url FROM shop_banners WHERE id=?", (bid,))
        row = cur.fetchone()
        if row and row['image_url'] and '/uploads/banners/' in row['image_url']:
            local = os.path.join('static', row['image_url'].lstrip('/'))
            try:
                if os.path.exists(local): os.remove(local)
            except Exception: pass
        conn.execute("DELETE FROM shop_banners WHERE id=?", (bid,))
        conn.commit()
        return jsonify({'success':True})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/banners/upload-image', methods=['POST'])
@login_required
def shop_upload_banner_image():
    if session.get('role') not in ('admin','superadmin'):
        return jsonify({'success':False,'error':'Non autorisé'}), 403
    if 'image' not in request.files:
        return jsonify({'success':False,'error':'Aucun fichier'})
    f = request.files['image']
    if not f or not f.filename or not allowed_file(f.filename):
        return jsonify({'success':False,'error':'Fichier invalide'})
    try:
        ext   = f.filename.rsplit('.',1)[1].lower()
        fname = secure_filename(f'banner_{datetime.now().strftime("%Y%m%d_%H%M%S%f")}.{ext}')
        dest  = os.path.join(app.config['UPLOAD_FOLDER'], 'banners')
        os.makedirs(dest, exist_ok=True)
        f.save(os.path.join(dest, fname))
        url = f'/static/uploads/banners/{fname}'
        return jsonify({'success':True, 'url':url})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})

# ── FIN MODULE BANNIÈRES ─────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════
# NOUVELLES ROUTES : accès étendu, inactivité, logs activité, factures
# ═══════════════════════════════════════════════════════════════════

@app.route('/shop/api/access/set-v2', methods=['POST'])
@login_required
def shop_set_access_v2():
    """Définit accès boutique + permissions granulaires pour N'IMPORTE quel utilisateur.
    Body JSON : { user_id, access, add, edit, delete }
    Réservé aux admins/superadmins.
    """
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    data = request.get_json(force=True, silent=True) or {}
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'error': 'user_id manquant'})
    perms = {
        'add':    bool(data.get('add',    False)),
        'edit':   bool(data.get('edit',   False)),
        'delete': bool(data.get('delete', False)),
    }
    shop_access = 1 if data.get('access', False) else 0
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Utilisateur introuvable'})
        # Même les admins peuvent modifier d'autres utilisateurs (sauf superadmin)
        if row['role'] == 'superadmin' and session.get('role') != 'superadmin':
            return jsonify({'success': False, 'error': 'Seul le superadmin peut modifier ce compte'})
        conn.execute(
            "UPDATE users SET shop_permissions=?, shop_access=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (json.dumps(perms), shop_access, user_id)
        )
        conn.commit()
        return jsonify({'success': True, 'permissions': perms, 'access': bool(shop_access)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/users/check-inactivity', methods=['POST'])
@login_required
def shop_check_inactivity():
    """Passe en statut 'inactive' les users n'ayant pas connecté depuis > 8 jours.
    Retourne la liste des utilisateurs mis à jour.
    """
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    updated = []
    try:
        threshold = (datetime.now() - timedelta(days=8)).isoformat()
        rows = conn.execute(
            """SELECT id, username, last_login, status FROM users
               WHERE (last_login IS NULL OR last_login < ?) AND status = 'active'
               AND role NOT IN ('admin','superadmin')""",
            (threshold,)
        ).fetchall()
        for r in rows:
            conn.execute(
                "UPDATE users SET status='inactive', updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (r['id'],)
            )
            updated.append({'id': r['id'], 'username': r['username'], 'last_login': r['last_login']})
        conn.commit()
        return jsonify({'success': True, 'updated': updated, 'count': len(updated)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/activity/log', methods=['GET'])
@login_required
def shop_get_activity():
    """Retourne les données d'activité (commandes gérées) par utilisateur sur 30 jours."""
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        # Commandes gérées par utilisateur
        cur.execute("""
            SELECT u.id, u.username, u.email,
                   COUNT(o.id) as total_orders,
                   SUM(CASE WHEN o.status='delivered' THEN 1 ELSE 0 END) as delivered,
                   SUM(CASE WHEN o.status='cancelled' THEN 1 ELSE 0 END) as cancelled,
                   DATE(COALESCE(o.handled_at, o.updated_at)) as day
            FROM users u
            LEFT JOIN shop_orders o ON o.handled_by = u.id
                AND DATE(COALESCE(o.handled_at, o.updated_at)) >= DATE('now','-30 days')
            WHERE u.shop_access = 1 OR u.role IN ('admin','superadmin')
            GROUP BY u.id, day
            ORDER BY u.id, day
        """)
        rows = cur.fetchall()
        # Structurer par utilisateur
        users_data = {}
        for r in rows:
            uid2 = r['id']
            if uid2 not in users_data:
                users_data[uid2] = {
                    'id': uid2, 'username': r['username'], 'email': r['email'],
                    'days': {}
                }
            if r['day']:
                users_data[uid2]['days'][r['day']] = {
                    'orders': r['total_orders'] or 0,
                    'delivered': r['delivered'] or 0,
                    'cancelled': r['cancelled'] or 0,
                }
        # Totaux globaux
        cur.execute("""
            SELECT handled_by, COUNT(*) as total,
                   SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as delivered
            FROM shop_orders WHERE handled_by IS NOT NULL
            GROUP BY handled_by
        """)
        for r in cur.fetchall():
            uid2 = r['handled_by']
            if uid2 in users_data:
                users_data[uid2]['total_all_time'] = r['total']
                users_data[uid2]['total_delivered'] = r['delivered']
        return jsonify({'success': True, 'data': list(users_data.values())})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()



# ── FACTURES ─────────────────────────────────────────────────────────

def _next_invoice_number():
    """Génère le prochain numéro de facture : FAC-YYYY-NNNN"""
    conn = get_db_connection()
    try:
        year = datetime.now().year
        row = conn.execute(
            "SELECT COUNT(*) as cnt FROM shop_invoices WHERE invoice_number LIKE ?",
            (f'FAC-{year}-%',)
        ).fetchone()
        n = (row['cnt'] if row else 0) + 1
        return f'FAC-{year}-{n:04d}'
    except Exception:
        return f'FAC-{datetime.now().year}-{random.randint(1000,9999)}'
    finally:
        if conn: conn.close()


@app.route('/shop/api/invoice', methods=['POST'])
@login_required
def shop_create_invoice():
    """Crée une facture manuelle ou depuis une commande."""
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    items = d.get('items', [])
    subtotal = sum(float(i.get('price', 0)) * int(i.get('qty', 1)) for i in items)
    discount = float(d.get('discount', 0) or 0)
    tax_rate = float(d.get('tax_rate', 0) or 0)
    tax_amount = round((subtotal - discount) * tax_rate / 100, 2)
    total = round(subtotal - discount + tax_amount, 2)
    inv_num = _next_invoice_number()
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO shop_invoices
            (invoice_number, order_id, type, customer_name, customer_phone,
             customer_city, customer_email, items_json, subtotal, discount,
             tax_rate, tax_amount, total, pay_method, status, notes, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (inv_num, d.get('order_id'), d.get('type', 'manual'),
              d.get('customer_name', ''), d.get('customer_phone', ''),
              d.get('customer_city', ''), d.get('customer_email', ''),
              json.dumps(items, ensure_ascii=False),
              subtotal, discount, tax_rate, tax_amount, total,
              d.get('pay_method', ''), d.get('status', 'draft'),
              d.get('notes', ''), session.get('user_id')))
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid, 'invoice_number': inv_num, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/invoice/all', methods=['GET'])
@login_required
def shop_get_invoices():
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM shop_invoices ORDER BY created_at DESC LIMIT 200"
        ).fetchall()
        invoices = []
        for r in rows:
            inv = dict(r)
            try: inv['items'] = json.loads(inv['items_json'] or '[]')
            except: inv['items'] = []
            invoices.append(inv)
        return jsonify({'success': True, 'invoices': invoices})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/invoice/<int:iid>', methods=['PUT'])
@login_required
def shop_update_invoice(iid):
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db_connection()
    try:
        items = d.get('items', [])
        subtotal = sum(float(i.get('price', 0)) * int(i.get('qty', 1)) for i in items)
        discount = float(d.get('discount', 0) or 0)
        tax_rate = float(d.get('tax_rate', 0) or 0)
        tax_amount = round((subtotal - discount) * tax_rate / 100, 2)
        total = round(subtotal - discount + tax_amount, 2)
        conn.execute("""
            UPDATE shop_invoices SET
                customer_name=?, customer_phone=?, customer_city=?, customer_email=?,
                items_json=?, subtotal=?, discount=?, tax_rate=?, tax_amount=?, total=?,
                pay_method=?, status=?, notes=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (d.get('customer_name',''), d.get('customer_phone',''),
              d.get('customer_city',''), d.get('customer_email',''),
              json.dumps(items, ensure_ascii=False),
              subtotal, discount, tax_rate, tax_amount, total,
              d.get('pay_method',''), d.get('status','draft'),
              d.get('notes',''), iid))
        conn.commit()
        return jsonify({'success': True, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/invoice/<int:iid>', methods=['DELETE'])
@login_required
def shop_delete_invoice(iid):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM shop_invoices WHERE id=?", (iid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/invoice/from-order/<int:oid>', methods=['POST'])
@login_required
def shop_invoice_from_order(oid):
    """Génère automatiquement une facture à partir d'une commande."""
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        # Vérifier si facture existe déjà
        existing = conn.execute(
            "SELECT id, invoice_number FROM shop_invoices WHERE order_id=?", (oid,)
        ).fetchone()
        if existing:
            return jsonify({'success': True, 'id': existing['id'],
                            'invoice_number': existing['invoice_number'], 'already_exists': True})
        order = conn.execute("SELECT * FROM shop_orders WHERE id=?", (oid,)).fetchone()
        if not order:
            return jsonify({'success': False, 'error': 'Commande introuvable'})
        order = dict(order)
        items = json.loads(order.get('items_json') or '[]')
        inv_num = _next_invoice_number()
        total = float(order.get('total', 0))
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO shop_invoices
            (invoice_number, order_id, type, customer_name, customer_phone,
             customer_city, items_json, subtotal, total, pay_method, status, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (inv_num, oid, 'order',
              order.get('customer_name',''), order.get('customer_phone',''),
              order.get('customer_city',''), order.get('items_json','[]'),
              total, total, order.get('pay_method',''),
              'issued', session.get('user_id')))
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid, 'invoice_number': inv_num, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()


@app.route('/shop/api/products/stock-value', methods=['GET'])
@login_required
def shop_stock_value():
    """Calcule la valeur totale du stock."""
    perms = get_shop_perms(session.get('user_id'))
    if not perms.get('access'):
        return jsonify({'success': False, 'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    try:
        rows = conn.execute(
            "SELECT id, name, price, stock, category FROM shop_products WHERE is_active=1 AND stock > 0"
        ).fetchall()
        items = []
        total_value = 0
        for r in rows:
            val = float(r['price']) * int(r['stock'])
            total_value += val
            items.append({'id': r['id'], 'name': r['name'], 'price': r['price'],
                          'stock': r['stock'], 'category': r['category'], 'value': val})
        items.sort(key=lambda x: x['value'], reverse=True)
        return jsonify({'success': True, 'items': items, 'total_value': total_value})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn: conn.close()

# ── FIN NOUVELLES ROUTES ──────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════
# MODULE IA BOUTIQUE — Agent Client & Recommandations Personnalisées
# Proxy sécurisé vers l'API Anthropic — clé jamais exposée côté client
# ═══════════════════════════════════════════════════════════════════

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
ANTHROPIC_URL     = 'https://api.anthropic.com/v1/messages'
CLAUDE_MODEL      = 'claude-sonnet-4-6'

# Rate limiting simple en mémoire
import collections as _collections
_ai_rate  = _collections.defaultdict(list)
AI_MAX_RPM = 20  # requêtes par minute par IP

def _ai_rate_ok(ip: str) -> bool:
    now = time.time()
    window = [t for t in _ai_rate[ip] if now - t < 60]
    _ai_rate[ip] = window
    if len(window) >= AI_MAX_RPM:
        return False
    _ai_rate[ip].append(now)
    return True


def _call_anthropic(system: str, messages: list, max_tokens: int = 800) -> dict:
    """Appelle l'API Anthropic et retourne la réponse parsée."""
    import json as _j
    import urllib.request as _ur
    import urllib.error as _ue

    if not ANTHROPIC_API_KEY:
        return {'error': 'Clé API Anthropic non configurée (variable ANTHROPIC_API_KEY)'}

    payload = _j.dumps({
        'model': CLAUDE_MODEL,
        'max_tokens': max_tokens,
        'system': system,
        'messages': messages,
    }).encode('utf-8')

    req = _ur.Request(ANTHROPIC_URL, data=payload, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('x-api-key', ANTHROPIC_API_KEY)
    req.add_header('anthropic-version', '2023-06-01')

    try:
        with _ur.urlopen(req, timeout=30) as resp:
            data = _j.loads(resp.read().decode('utf-8'))
            text = ''.join(b.get('text', '') for b in data.get('content', []) if b.get('type') == 'text')
            return {'text': text}
    except _ue.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        return {'error': f'Anthropic HTTP {e.code}: {body[:300]}'}
    except Exception as ex:
        return {'error': str(ex)}


# ── Route : Agent Chat Client ────────────────────────────────────────
@app.route('/shop/api/ai/chat', methods=['POST'])
def shop_ai_chat():
    """
    Agent conversationnel Claude pour les clients de k-ni Store.
    Répond aux questions produits, livraison, paiement.
    Escalade intelligente vers WhatsApp / support humain.
    Aucune authentification requise (public).
    """
    ip = request.remote_addr or '0.0.0.0'
    if not _ai_rate_ok(ip):
        return jsonify({'error': 'Trop de requêtes. Attendez une minute.'}), 429

    data    = request.get_json(force=True, silent=True) or {}
    history = data.get('history', [])   # [{role, content}, ...]
    message = data.get('message', '').strip()
    catalog = data.get('catalog', [])   # [{id, name, category, price, stock, badge}, ...]

    if not message:
        return jsonify({'error': 'Message vide'}), 400

    # Limite de sécurité : max 20 messages dans l'historique
    history = history[-20:]

    # Résumé catalogue (max 40 produits pour ne pas dépasser le contexte)
    cat_lines = '\n'.join(
        f"- [{p.get('id','')}] {p.get('name','')} | {p.get('category','')} | "
        f"{p.get('price',0):,} XAF | stock:{p.get('stock',0)} | badge:{p.get('badge','')}"
        for p in catalog[:40]
    ) or '(catalogue non chargé)'

    SYSTEM = f"""Tu es Keni, l'assistante IA de k-ni Store — boutique en ligne de qualité basée à Yaoundé, Cameroun.
Tu parles français. Tu es chaleureuse, efficace et professionnelle.

CATALOGUE ACTUEL (produits disponibles) :
{cat_lines}

INFORMATIONS BOUTIQUE :
- Paiements : Orange Money 695 072 759 | MTN MoMo 670 695 946 (nom : Fabrice Kengni)
- Livraison : 3-7 jours ouvrés partout au Cameroun. Offerte sur toute commande.
- Retours : 7 jours si produit défectueux, avec preuve d'achat.
- Support WhatsApp : +237 695 072 759 (réponse en moins d'1h)
- Garantie constructeur sur tous les produits électroniques.

TES RÈGLES :
1. Réponds toujours en français, sois concise (max 3 phrases par réponse sauf si on demande des détails).
2. Si un client demande un produit, vérifie dans le catalogue et propose le(s) meilleur(s) choix avec le prix.
3. Pour les questions de paiement/livraison, donne les infos directement.
4. Si le client a un problème (retour, réclamation, article manquant, défectueux), collecte :
   - Nom du produit concerné
   - Numéro de commande si disponible
   - Description du problème
   Puis indique qu'un conseiller va le contacter.
5. Après avoir collecté les infos d'un problème, génère un résumé WhatsApp avec le préfixe [ESCALADE].
6. Si le client dit "parler à quelqu'un", "agent", "humain", "support" → génère directement [ESCALADE].
7. N'invente jamais de prix ou de produits absents du catalogue.

FORMAT ESCALADE (quand nécessaire) :
[ESCALADE]
CLIENT: [nom si connu, sinon "Client"]
SUJET: [en 1 ligne]
DÉTAILS: [résumé en 2-3 lignes]
[FIN ESCALADE]

Réponds UNIQUEMENT avec le texte de ta réponse (sans préfixe ni balise sauf [ESCALADE])."""

    messages = [*history, {'role': 'user', 'content': message}]

    result = _call_anthropic(SYSTEM, messages, max_tokens=600)

    if 'error' in result:
        # Fallback si API non configurée : réponse statique utile
        fb = _ai_chat_fallback(message, catalog)
        return jsonify({'reply': fb, 'escalade': False, 'fallback': True})

    reply = result['text'].strip()
    escalade = '[ESCALADE]' in reply

    # Extraire le bloc escalade pour construire le lien WhatsApp
    wa_text = None
    if escalade:
        try:
            bloc = reply.split('[ESCALADE]')[1].split('[FIN ESCALADE]')[0].strip()
            wa_text = f"Bonjour k-ni Store ! Je suis transféré(e) depuis le chat.\n\n{bloc}"
        except Exception:
            wa_text = "Bonjour, j'ai besoin d'une assistance humaine depuis le chat k-ni."
        # Nettoyer la réponse affichée
        reply = reply.replace('[ESCALADE]', '').split('[FIN ESCALADE]')[0].strip()
        if not reply:
            reply = "Je vais vous mettre en contact avec un conseiller k-ni qui pourra vous aider directement. 😊"

    return jsonify({
        'reply': reply,
        'escalade': escalade,
        'wa_text': wa_text,
    })


def _ai_chat_fallback(message: str, catalog: list) -> str:
    """Réponses de fallback si l'API Claude n'est pas configurée."""
    msg = message.lower()
    if any(w in msg for w in ['livraison', 'délai', 'délais']):
        return "Nous livrons partout au Cameroun en 3 à 7 jours ouvrés. La livraison est offerte sur toute commande. 🚚"
    if any(w in msg for w in ['paiement', 'payer', 'orange', 'mtn']):
        return "Vous pouvez payer par Orange Money (695 072 759) ou MTN MoMo (670 695 946) au nom de Fabrice Kengni. 💳"
    if any(w in msg for w in ['retour', 'remboursement', 'défectueux', 'problème']):
        return "Pour tout retour ou problème, contactez-nous sur WhatsApp +237 695 072 759. Retours acceptés sous 7 jours. 🔄"
    if any(w in msg for w in ['prix', 'combien', 'coût']):
        return "Les prix varient selon le produit. Consultez notre catalogue ou posez-moi une question sur un article précis ! 💰"
    # Chercher un produit par mot-clé
    for p in catalog[:20]:
        name = (p.get('name') or '').lower()
        if any(w in name for w in msg.split() if len(w) > 3):
            price = p.get('price', 0)
            return f"Nous avons **{p['name']}** à {price:,} XAF. Stock disponible. Souhaitez-vous commander ?"
    return "Bonjour ! Je suis Keni, votre assistante k-ni Store. Posez-moi vos questions sur nos produits, livraisons ou paiements ! 😊"


# ── Route : Recommandations IA ───────────────────────────────────────
@app.route('/shop/api/ai/recommend', methods=['POST'])
def shop_ai_recommend():
    """
    Génère des recommandations produits personnalisées basées sur le
    comportement de navigation (recherches, produits vus, panier).
    """
    ip = request.remote_addr or '0.0.0.0'
    if not _ai_rate_ok(ip):
        return jsonify({'error': 'Trop de requêtes.'}), 429

    data     = request.get_json(force=True, silent=True) or {}
    viewed   = data.get('viewed', [])      # [{id, name, category, price}, ...]
    searches = data.get('searches', [])    # ['écouteurs', 'sport', ...]
    cart     = data.get('cart', [])        # [{id, name, price}, ...]
    catalog  = data.get('catalog', [])     # catalogue complet

    if not catalog:
        return jsonify({'recommendations': [], 'reason': 'Catalogue vide'})

    # Si pas assez de signaux, retourner les produits hot/new
    if not viewed and not searches and not cart:
        popular = [p for p in catalog if p.get('badge') in ('hot', 'new')][:4]
        return jsonify({
            'recommendations': popular,
            'reason': 'Produits populaires du moment',
            'fallback': True
        })

    # Construire le profil d'intérêt
    profile_lines = []
    if searches:
        profile_lines.append(f"Recherches : {', '.join(searches[-6:])}")
    if viewed:
        viewed_desc = ', '.join(f"{p.get('name','')} ({p.get('price',0):,} XAF)" for p in viewed[-5:])
        profile_lines.append(f"Produits consultés : {viewed_desc}")
    if cart:
        cart_desc = ', '.join(f"{p.get('name','')}" for p in cart[:3])
        profile_lines.append(f"Dans le panier : {cart_desc}")

    profile = '\n'.join(profile_lines)

    # Catalogue pour l'IA (max 50 produits)
    cat_json = json.dumps([{
        'id': p.get('id'), 'name': p.get('name'),
        'category': p.get('category'), 'price': p.get('price'),
        'stock': p.get('stock'), 'badge': p.get('badge')
    } for p in catalog[:50] if p.get('stock', 0) > 0], ensure_ascii=False)

    SYSTEM = """Tu es un moteur de recommandations pour k-ni Store (boutique Cameroun).
Tu analyses le comportement d'un visiteur et tu sélectionnes les produits les plus pertinents.
Tu réponds UNIQUEMENT en JSON valide, sans aucun texte autour."""

    prompt = f"""Profil du visiteur :
{profile}

Catalogue disponible (JSON) :
{cat_json}

Sélectionne exactement 4 produits parmi le catalogue qui correspondent au mieux à ce visiteur.
Réponds en JSON pur (aucun texte autour) :
{{
  "reason": "phrase courte expliquant la sélection (ex: Vous semblez intéressé par...)",
  "ids": [id1, id2, id3, id4],
  "explanations": {{
    "id1": "pourquoi ce produit en 1 phrase max",
    "id2": "...",
    "id3": "...",
    "id4": "..."
  }}
}}"""

    result = _call_anthropic(SYSTEM, [{'role': 'user', 'content': prompt}], max_tokens=400)

    if 'error' in result:
        # Fallback : produits de la même catégorie la plus vue
        cats = [p.get('category') for p in viewed if p.get('category')]
        fav_cat = max(set(cats), key=cats.count) if cats else None
        recs = [p for p in catalog if p.get('stock', 0) > 0 and
                (not fav_cat or p.get('category') == fav_cat)][:4]
        return jsonify({
            'recommendations': recs,
            'reason': f'Basé sur votre intérêt pour {fav_cat or "nos produits"}',
            'fallback': True
        })

    # Parser le JSON retourné par Claude
    try:
        import re as _re
        raw = result['text'].strip()
        # Extraire uniquement le bloc JSON
        m = _re.search(r'\{.*\}', raw, _re.DOTALL)
        if not m:
            raise ValueError('Pas de JSON trouvé')
        parsed = json.loads(m.group(0))
        ids = [int(i) for i in parsed.get('ids', [])]
        reason = parsed.get('reason', 'Sélectionné pour vous')
        explanations = parsed.get('explanations', {})

        # Récupérer les produits correspondants
        prod_map = {p['id']: p for p in catalog if 'id' in p}
        recs = []
        for pid in ids:
            p = prod_map.get(pid)
            if p:
                p2 = dict(p)
                p2['ai_reason'] = explanations.get(str(pid), '')
                recs.append(p2)

        return jsonify({'recommendations': recs, 'reason': reason})

    except Exception as ex:
        # Fallback
        recs = [p for p in catalog if p.get('badge') == 'hot' and p.get('stock', 0) > 0][:4]
        return jsonify({'recommendations': recs, 'reason': 'Produits populaires', 'fallback': True})


if __name__ == '__main__':

    # Démarrer le scheduler d'agenda (rappels email Gmail)
    start_agenda_scheduler()
    
    # Run the application
    print("=" * 60)
    print("🚀 Kengni Finance v2.0 - Démarrage")
    print("=" * 60)
    print("📊 Application de gestion financière et trading avec IA")
    print("🌐 URL: http://localhost:5001")
    print("👤 Email: fabrice.kengni@icloud.com")
    print("📅 Agenda: http://localhost:5001/agenda")
    print("=" * 60)
    
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))

else:
    # Exécuté par Gunicorn (PythonAnywhere, Fly.io, etc.)
    init_db()
    start_agenda_scheduler()